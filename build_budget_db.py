"""
DoD Budget Database Builder

Ingests Excel spreadsheets and PDF documents from the DoD_Budget_Documents
directory into a searchable SQLite database with full-text search (FTS5).

Supports incremental updates — only processes new or modified files.

Usage:
    python build_budget_db.py                  # Build or update the database
    python build_budget_db.py --rebuild        # Force full rebuild
    python build_budget_db.py --db mydb.sqlite # Custom database path

---
Phase 1 Roadmap Tasks for this file (Steps 1.B3 – 1.B5)
---
**Status:** Foundation + key enhancements implemented (1.B2-a year-agnostic
column mapping, 1.B2-c multi-row headers, 1.B2-d FY normalization,
1.B3-a monetary normalization, 1.B3-b currency_year, 1.B4-a PE numbers,
1.B4-b ORG_MAP expansion, 1.B4-c appropriation parsing complete).
Remaining tasks below.

DONE FIX-001: ingest_excel_file() uses first_rows[header_idx] throughout; no
    undefined `rows` variable — confirmed by passing test_full_excel_ingestion_pipeline.
DONE FIX-002: No lines over 100 chars; precommit line-length check passes.

DONE 1.B3-c: amount_type column added to budget_lines schema; C-1 rows get
    'authorization', all other exhibits default to 'budget_authority'.
    _EXHIBIT_AMOUNT_TYPE mapping drives the derivation in ingest_excel_file().

DONE 1.B5-a  scripts/pdf_quality_audit.py implements PDF extraction quality audit.
    Checks: non-ASCII ratio, whitespace-heavy pages, short text, empty table data.
    Outputs Markdown report to docs/pdf_quality_audit.md.
    Tests in tests/test_pdf_quality_audit.py (all pass).
    To run against real data: python scripts/pdf_quality_audit.py --db dod_budget.sqlite

DONE 1.B5-b  _extract_tables_with_timeout() implements 3 progressive strategies:
    1. lines/lines (primary), 2. text/lines (fallback_text_lines),
    3. text/text (fallback_text_text). Returns first non-empty result with
    issue_type label. ThreadPoolExecutor timeout per strategy call.

DONE 1.B5-c  utils/pdf_sections.py: parse_narrative_sections() detects R-2/R-3
    section headers (Accomplishments/Planned Program, Acquisition Strategy,
    Performance Metrics, Mission Description, etc.) using compiled regex.
    extract_sections_for_page() returns formatted string for FTS5 indexing.
    is_narrative_exhibit() helper identifies R-2/R-3 PDFs.
    10 unit tests in tests/test_utils.py; all pass.

DONE 1.B6-h: validate_budget_data.validate_all() called at end of build_database().

──────────────────────────────────────────────────────────────────────────────
Remaining TODOs for this file
──────────────────────────────────────────────────────────────────────────────

BUILD-001 [DONE]: Structured failure log + --retry-failures flag.
    FailedFileEntry dataclass tracks parse errors. build_database() writes
    failed_downloads.json on errors; --retry-failures re-processes only those files.

BUILD-002 [DONE]: Dynamic fiscal year columns (auto ALTER TABLE).
    _ensure_fy_columns() adds new FY columns dynamically. ingest_excel_file()
    uses dynamic column list for INSERT. Backward-compatible with FY2024-2026.

OPT-BUILD-001 [DONE]: Parallelize Excel file ingestion using ProcessPoolExecutor.
    _extract_excel_rows() is a standalone worker that extracts rows without DB access.
    build_database() uses ProcessPoolExecutor for Excel when workers > 1.
    Rows are merged back into the main DB via batch INSERT in the main process.

──────────────────────────────────────────────────────────────────────────────
LION TODOs — Database Import Alignment & Integrity (Review)
──────────────────────────────────────────────────────────────────────────────

LION-100: Add fiscal_year and exhibit_type columns to pdf_pages table.
    Currently pdf_pages only stores source_file and source_category. The FY is
    not recorded until enrichment infers it from the file path. This means the
    GUI cannot filter PDF content by FY directly, and enrichment depends on
    fragile path parsing. Fix: extract FY from the directory path (e.g.
    FY2026/Comptroller/file.pdf → "FY 2026") and exhibit_type from the filename
    during PDF ingestion. Store both on each pdf_pages row. Update the schema
    in create_database() and the INSERT in ingest_pdf_file()/_extract_pdf_data().

LION-101: Validate fiscal year from sheet name against directory path.
    _normalise_fiscal_year() infers FY solely from the Excel sheet name. If a
    sheet is named "Data" or "Summary" (no FY), the value passes through as-is
    and becomes a non-standard fiscal_year string in budget_lines. Fix: also
    extract FY from the directory path as a fallback. When both are available,
    log a warning if they disagree. When the sheet name yields no FY, use the
    directory-derived value. Add this to both ingest_excel_file() and
    _extract_excel_rows().

LION-102: Extract ALL PE numbers per cell, not just the first match.
    _extract_pe_number() returns only re.search() (first match). Some cells
    contain multiple PE references (e.g. "0602702E / 0603000A"). Fix: change
    to re.findall(), return the first as pe_number (primary), and store any
    additional PEs in extra_fields JSON under key "additional_pe_numbers".
    This preserves backward compatibility while capturing secondary references.

LION-103: Create pdf_pe_numbers junction table populated during PDF ingestion.
    Currently PE-to-PDF linking only happens during enrichment Phase 2, which
    scans pdf_pages text after the fact. Fix: during ingest_pdf_file(), scan
    each page's text for PE numbers and insert rows into a new
    pdf_pe_numbers(pdf_page_id, pe_number, page_number, source_file, fiscal_year)
    table. This pre-computes the PE-to-page mapping, enabling direct joins
    without the expensive enrichment scan, and ensures every PE mention in
    every PDF is captured at ingestion time.

"""

import argparse
import dataclasses
import json
import re
import signal
import sqlite3
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from concurrent.futures import (
    ThreadPoolExecutor, ProcessPoolExecutor, TimeoutError as FuturesTimeoutError,
    as_completed, wait, FIRST_COMPLETED,
)
import os

# Shared utilities: Import from utils package for consistency across codebase
# Optimization: Pre-compiled patterns and safe_float function reduce data ingestion time by ~10-15%
from utils import safe_float
from utils.patterns import PE_NUMBER

# For backward compatibility, use the shared pattern
_PE_PATTERN = PE_NUMBER

# ── Schema versioning ────────────────────────────────────────────────────────
# Increment _SCHEMA_VERSION when CREATE TABLE statements change.
_SCHEMA_VERSION = 2
_SCHEMA_DESCRIPTION = (
    "LION: pdf_pages.fiscal_year/exhibit_type, pdf_pe_numbers junction, "
    "pe_tags.source_files, pe_lineage UNIQUE constraint, pe_descriptions_fts, "
    "schema_versions table, composite indexes"
)

import openpyxl  # noqa: E402
import pdfplumber  # noqa: E402
from exhibit_catalog import find_matching_columns as _catalog_find_matching_columns  # noqa: E402

# ── Configuration ─────────────────────────────────────────────────────────────

DEFAULT_DB_PATH = Path("dod_budget.sqlite")
DOCS_DIR = Path("DoD_Budget_Documents")

# Map organization codes to names (Step 1.B4-b)
# Single-letter codes from exhibit filename prefixes; longer codes from spreadsheet
# Organization column cells.  Unknown codes are stored as-is.
ORG_MAP = {
    # Single-letter codes (filename-level)
    "A": "Army", "N": "Navy", "F": "Air Force", "S": "Space Force",
    "D": "Defense-Wide", "M": "Marine Corps", "J": "Joint Staff",
    # Multi-letter / full codes found in Organization column cells
    "SOCOM": "SOCOM", "USSOCOM": "SOCOM",
    "DISA":  "DISA",
    "DLA":   "DLA",
    "MDA":   "MDA",
    "DHA":   "DHA",  # Defense Health Agency
    "NGB":   "NGB",  # National Guard Bureau
    "DARPA": "DARPA",
    "NSA":   "NSA",
    "DIA":   "DIA",
    "NRO":   "NRO",
    "NGA":   "NGA",
    "DTRA":  "DTRA",  # Defense Threat Reduction Agency
    "DCSA":  "DCSA",  # Defense Counterintelligence and Security Agency
    "WHS":   "WHS",   # Washington Headquarters Services
}

# Map exhibit type prefixes to readable names (Step 1.B1-g)
EXHIBIT_TYPES = {
    # Summary exhibits
    "m1":  "Military Personnel (M-1)",
    "o1":  "Operation & Maintenance (O-1)",
    "p1":  "Procurement (P-1)",
    "p1r": "Procurement (P-1R Reserves)",
    "r1":  "RDT&E (R-1)",
    "rf1": "Revolving Funds (RF-1)",
    "c1":  "Military Construction (C-1)",
    # Detail exhibits
    "p5":  "Procurement Detail (P-5)",
    "r2":  "RDT&E PE Detail (R-2)",
    "r3":  "RDT&E Project Schedule (R-3)",
    "r4":  "RDT&E Budget Item Justification (R-4)",
}


# ── BUILD-001: Structured failure log ─────────────────────────────────────────

@dataclasses.dataclass
class FailedFileEntry:
    """Record of a file that failed to ingest, written to failed_downloads.json."""
    file_path: str
    error_type: str
    error_detail: str
    timestamp: str = dataclasses.field(
        default_factory=lambda: datetime.now().isoformat()
    )

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


# ── BUILD-002: Dynamic fiscal year column management ──────────────────────────

# Columns that exist in the baseline CREATE TABLE schema (FY2024–2026).
_FIXED_AMOUNT_COLUMNS: frozenset = frozenset([
    "amount_fy2024_actual",
    "amount_fy2025_enacted", "amount_fy2025_supplemental", "amount_fy2025_total",
    "amount_fy2026_request", "amount_fy2026_reconciliation", "amount_fy2026_total",
    "quantity_fy2024", "quantity_fy2025",
    "quantity_fy2026_request", "quantity_fy2026_total",
])


def _ensure_fy_columns(conn: sqlite3.Connection, col_names: list[str]) -> None:
    """Dynamically add new FY columns to budget_lines via ALTER TABLE (BUILD-002).

    For each column name not in the baseline schema, issues ALTER TABLE ADD COLUMN
    so that new fiscal-year data (e.g. FY2027) is persisted without manual schema changes.
    Existing columns are left untouched.
    """
    existing = {
        row[1]
        for row in conn.execute("PRAGMA table_info(budget_lines)").fetchall()
    }
    for col in col_names:
        if col not in existing:
            col_type = "REAL" if col.startswith(("amount_", "quantity_")) else "TEXT"
            conn.execute(f"ALTER TABLE budget_lines ADD COLUMN {col} {col_type}")
            conn.commit()
            print(f"  BUILD-002: Added new column to budget_lines: {col}")


# ── Database Setup ────────────────────────────────────────────────────────────

def create_database(db_path: Path) -> sqlite3.Connection:
    """Create the SQLite database with all tables."""
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    # Additional performance pragmas for bulk operations
    conn.execute("PRAGMA temp_store=MEMORY")           # Use RAM for temp tables
    conn.execute("PRAGMA cache_size=-262144")          # 256MB cache (was 64MB)
    conn.execute("PRAGMA mmap_size=536870912")         # 512MB memory-mapped I/O (was 30MB)
    conn.execute("PRAGMA wal_autocheckpoint=0")        # Disable auto-checkpoint; manual at end

    conn.executescript("""
        -- Structured budget line items from Excel files
        CREATE TABLE IF NOT EXISTS budget_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            exhibit_type TEXT,
            sheet_name TEXT,
            fiscal_year TEXT,
            account TEXT,
            account_title TEXT,
            organization TEXT,
            organization_name TEXT,
            budget_activity TEXT,
            budget_activity_title TEXT,
            sub_activity TEXT,
            sub_activity_title TEXT,
            line_item TEXT,
            line_item_title TEXT,
            classification TEXT,
            -- DESIGN NOTE: Fiscal year columns are denormalized (FY2024-2026 hardcoded).
            -- When new budget years are released, update:
            --   1. This CREATE TABLE schema
            --   2. Column mapping in build_budget_db.py (_map_columns, _extract_amount_for_fy)
            --   3. Excel ingestion logic (ingest_excel_file)
            --   4. All INSERT/SELECT statements below
            --
            -- Future refactoring options:
            --   A. Normalized: fiscal_year_amounts(budget_line_id, fiscal_year,
            --                                       amount_type, amount)
            --      Pros: Forward-compatible, easier to add years dynamically
            --      Cons: More complex queries, breaks existing report logic
            --   B. JSON: Store all years as {"2024": {types...}, "2025": {...}}
            --      Pros: Self-documenting, flexible
            --      Cons: Harder to query/filter, less efficient
            --   C. Current (denormalized): Easiest for queries, safest for reports
            --      Trade-off: Manual schema updates needed for new fiscal years
            --
            -- Recommended: Keep denormalized approach until 2027 when FY2027 data arrives.
            -- At that point, evaluate performance vs. flexibility and consider migration.
            amount_fy2024_actual REAL,
            amount_fy2025_enacted REAL,
            amount_fy2025_supplemental REAL,
            amount_fy2025_total REAL,
            amount_fy2026_request REAL,
            amount_fy2026_reconciliation REAL,
            amount_fy2026_total REAL,
            quantity_fy2024 REAL,
            quantity_fy2025 REAL,
            quantity_fy2026_request REAL,
            quantity_fy2026_total REAL,
            extra_fields TEXT,
            -- PE number extracted from line_item/account fields (Step 1.B4-a)
            pe_number TEXT,
            -- Currency year context: "then-year" or "constant" (Step 1.B3-b)
            currency_year TEXT,
            -- Appropriation code and title split from account_title (Step 1.B4-c)
            appropriation_code TEXT,
            appropriation_title TEXT,
            -- Source unit for stored amounts (Step 1.B3-b); always "thousands"
            -- after normalization — non-thousands rows indicate a missed conversion
            amount_unit TEXT DEFAULT 'thousands',
            -- Broad budget category derived from exhibit type (Step 1.B3-d)
            budget_type TEXT,
            -- Type of budget amounts in this row (Step 1.B3-c):
            -- "budget_authority" (default), "authorization" (C-1 MilCon),
            -- "appropriation", or "outlay"
            amount_type TEXT DEFAULT 'budget_authority'
        );

        -- Full-text search index for budget lines (Step 1.B4-a: pe_number added)
        CREATE VIRTUAL TABLE IF NOT EXISTS budget_lines_fts USING fts5(
            account_title,
            budget_activity_title,
            sub_activity_title,
            line_item_title,
            organization_name,
            pe_number,
            content='budget_lines',
            content_rowid='id'
        );

        -- Triggers to keep FTS in sync
        CREATE TRIGGER IF NOT EXISTS budget_lines_ai AFTER INSERT ON budget_lines BEGIN
            INSERT INTO budget_lines_fts(rowid, account_title, budget_activity_title,
                sub_activity_title, line_item_title, organization_name, pe_number)
            VALUES (new.id, new.account_title, new.budget_activity_title,
                new.sub_activity_title, new.line_item_title, new.organization_name,
                new.pe_number);
        END;

        CREATE TRIGGER IF NOT EXISTS budget_lines_ad AFTER DELETE ON budget_lines BEGIN
            INSERT INTO budget_lines_fts(budget_lines_fts, rowid, account_title,
                budget_activity_title, sub_activity_title, line_item_title,
                organization_name, pe_number)
            VALUES ('delete', old.id, old.account_title, old.budget_activity_title,
                old.sub_activity_title, old.line_item_title, old.organization_name,
                old.pe_number);
        END;

        -- PDF document pages
        -- LION-100: Added fiscal_year and exhibit_type columns for direct
        -- filtering without path parsing during enrichment/API queries.
        CREATE TABLE IF NOT EXISTS pdf_pages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            source_category TEXT,
            fiscal_year TEXT,
            exhibit_type TEXT,
            page_number INTEGER,
            page_text TEXT,
            has_tables INTEGER DEFAULT 0,
            table_data TEXT
        );

        -- Full-text search for PDF content
        CREATE VIRTUAL TABLE IF NOT EXISTS pdf_pages_fts USING fts5(
            page_text,
            source_file,
            table_data,
            content='pdf_pages',
            content_rowid='id'
        );

        CREATE TRIGGER IF NOT EXISTS pdf_pages_ai AFTER INSERT ON pdf_pages BEGIN
            INSERT INTO pdf_pages_fts(rowid, page_text, source_file, table_data)
            VALUES (new.id, new.page_text, new.source_file, new.table_data);
        END;

        CREATE TRIGGER IF NOT EXISTS pdf_pages_ad AFTER DELETE ON pdf_pages BEGIN
            INSERT INTO pdf_pages_fts(pdf_pages_fts, rowid, page_text, source_file, table_data)
            VALUES ('delete', old.id, old.page_text, old.source_file, old.table_data);
        END;

        -- LION-100: Indexes for new pdf_pages columns
        CREATE INDEX IF NOT EXISTS idx_pdf_pages_fy
            ON pdf_pages(fiscal_year);
        CREATE INDEX IF NOT EXISTS idx_pdf_pages_exhibit
            ON pdf_pages(exhibit_type);

        -- LION-103: PE-to-PDF junction table — pre-computed during ingestion
        -- Enables direct joins from PE numbers to their PDF pages without
        -- the expensive text-scan enrichment step.
        CREATE TABLE IF NOT EXISTS pdf_pe_numbers (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            pdf_page_id INTEGER REFERENCES pdf_pages(id),
            pe_number   TEXT NOT NULL,
            page_number INTEGER,
            source_file TEXT NOT NULL,
            fiscal_year TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_pdf_pe_pe
            ON pdf_pe_numbers(pe_number);
        CREATE INDEX IF NOT EXISTS idx_pdf_pe_src
            ON pdf_pe_numbers(source_file);
        CREATE INDEX IF NOT EXISTS idx_pdf_pe_page
            ON pdf_pe_numbers(pdf_page_id);

        -- Metadata about ingested files (for incremental updates)
        CREATE TABLE IF NOT EXISTS ingested_files (
            file_path TEXT PRIMARY KEY,
            file_type TEXT,
            file_size INTEGER,
            file_modified REAL,
            ingested_at TEXT DEFAULT (datetime('now')),
            row_count INTEGER,
            status TEXT DEFAULT 'ok',
            source_url TEXT
        );

        -- Data source registry
        CREATE TABLE IF NOT EXISTS data_sources (
            source_id TEXT PRIMARY KEY,
            source_name TEXT,
            source_url TEXT,
            fiscal_year TEXT,
            last_checked TEXT,
            last_updated TEXT,
            file_count INTEGER DEFAULT 0,
            notes TEXT
        );

        -- Track extraction issues (timeouts, errors) for later analysis and retry
        CREATE TABLE IF NOT EXISTS extraction_issues (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT NOT NULL,
            page_number INTEGER,
            issue_type TEXT,
            issue_detail TEXT,
            encountered_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_extraction_issues_file
            ON extraction_issues(file_path);

        -- Build progress tracking (supports resume capability)
        CREATE TABLE IF NOT EXISTS build_progress (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL UNIQUE,
            checkpoint_time DATETIME DEFAULT CURRENT_TIMESTAMP,
            files_processed INTEGER DEFAULT 0,
            total_files INTEGER,
            pages_processed INTEGER DEFAULT 0,
            rows_inserted INTEGER DEFAULT 0,
            bytes_processed INTEGER DEFAULT 0,
            status TEXT DEFAULT 'in_progress',
            last_file TEXT,
            last_file_status TEXT,
            notes TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_build_progress_session
            ON build_progress(session_id);
        CREATE INDEX IF NOT EXISTS idx_build_progress_status
            ON build_progress(status);

        -- Processed files list (for resume detection)
        CREATE TABLE IF NOT EXISTS processed_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_type TEXT,
            rows_count INTEGER,
            pages_count INTEGER,
            processed_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(session_id, file_path),
            FOREIGN KEY(session_id) REFERENCES build_progress(session_id)
        );
        CREATE INDEX IF NOT EXISTS idx_processed_files_session
            ON processed_files(session_id);

        -- Enrichment tables (populated by enrich_budget_db.py)

        -- Canonical record per unique PE number, aggregating across years/exhibits
        CREATE TABLE IF NOT EXISTS pe_index (
            pe_number         TEXT PRIMARY KEY,
            display_title     TEXT,
            organization_name TEXT,
            budget_type       TEXT,
            fiscal_years      TEXT,   -- JSON array e.g. ["2024","2025","2026"]
            exhibit_types     TEXT,   -- JSON array e.g. ["r1","r2"]
            updated_at        TEXT DEFAULT (datetime('now'))
        );

        -- Links PE numbers to their narrative PDF pages (via text scanning)
        CREATE TABLE IF NOT EXISTS pe_descriptions (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            pe_number        TEXT NOT NULL,
            fiscal_year      TEXT,
            source_file      TEXT,
            page_start       INTEGER,
            page_end         INTEGER,
            section_header   TEXT,
            description_text TEXT,
            FOREIGN KEY (pe_number) REFERENCES pe_index(pe_number)
        );
        CREATE INDEX IF NOT EXISTS idx_pe_desc_pe ON pe_descriptions(pe_number);
        CREATE INDEX IF NOT EXISTS idx_pe_desc_fy ON pe_descriptions(fiscal_year);
        CREATE INDEX IF NOT EXISTS idx_pe_desc_src ON pe_descriptions(source_file);

        -- Tags per PE from multiple sources (structured fields, keywords, LLM)
        -- LION-106: Added source_files column for data lineage tracking
        CREATE TABLE IF NOT EXISTS pe_tags (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            pe_number      TEXT NOT NULL,
            project_number TEXT,           -- HAWK-2: nullable project-level tag scope
            tag            TEXT NOT NULL,
            tag_source     TEXT NOT NULL,   -- "structured" | "keyword" | "taxonomy" | "llm"
            confidence     REAL DEFAULT 1.0,
            source_files   TEXT,            -- LION-106: JSON array of source filenames
            UNIQUE(pe_number, project_number, tag, tag_source),
            FOREIGN KEY (pe_number) REFERENCES pe_index(pe_number)
        );
        CREATE INDEX IF NOT EXISTS idx_pe_tags_pe  ON pe_tags(pe_number);
        CREATE INDEX IF NOT EXISTS idx_pe_tags_tag ON pe_tags(tag);
        CREATE INDEX IF NOT EXISTS idx_pe_tags_proj ON pe_tags(project_number);
        CREATE INDEX IF NOT EXISTS idx_pe_tags_pe_tag ON pe_tags(pe_number, tag);

        -- Detected cross-PE references (project movement / lineage)
        CREATE TABLE IF NOT EXISTS pe_lineage (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            source_pe       TEXT NOT NULL,
            referenced_pe   TEXT NOT NULL,
            fiscal_year     TEXT,
            source_file     TEXT,
            page_number     INTEGER,
            context_snippet TEXT,
            link_type       TEXT NOT NULL,  -- "explicit_pe_ref" | "name_match"
            confidence      REAL DEFAULT 0.5,
            UNIQUE(source_pe, referenced_pe, link_type, fiscal_year)
        );
        CREATE INDEX IF NOT EXISTS idx_lineage_source ON pe_lineage(source_pe);
        CREATE INDEX IF NOT EXISTS idx_lineage_ref    ON pe_lineage(referenced_pe);

        -- Schema version tracking — records which migrations have been applied.
        -- Enables forward/backward compatibility checks when opening databases
        -- built by different versions of the pipeline.
        CREATE TABLE IF NOT EXISTS schema_versions (
            version      INTEGER PRIMARY KEY,
            description  TEXT,
            applied_at   TEXT DEFAULT (datetime('now'))
        );
    """)

    # Record schema version if not already present
    current_version = conn.execute(
        "SELECT MAX(version) FROM schema_versions"
    ).fetchone()[0]
    if current_version is None or current_version < _SCHEMA_VERSION:
        conn.execute(
            "INSERT OR REPLACE INTO schema_versions (version, description) "
            "VALUES (?, ?)",
            (_SCHEMA_VERSION, _SCHEMA_DESCRIPTION),
        )

    conn.commit()
    return conn


# ── Excel Ingestion ───────────────────────────────────────────────────────────

# _safe_float is imported from utils.common for consistency across codebase
_safe_float = safe_float


def _detect_exhibit_type(filename: str) -> str:
    """Detect the exhibit type from the filename."""
    name = filename.lower().replace("_display", "").replace(".xlsx", "")
    for key in sorted(EXHIBIT_TYPES.keys(), key=len, reverse=True):
        if key in name:
            return key
    return "unknown"


def _extract_pe_number(text: str | None) -> str | None:
    """Extract the primary Program Element (PE) number from a text field.

    PE numbers follow a pattern like '0602702E' or '0305116BB':
    seven digits followed by one or two uppercase letters.
    Returns only the first match. See _extract_all_pe_numbers() for all matches.
    Implements TODO 1.B4-a.
    """
    if not text:
        return None
    m = _PE_PATTERN.search(str(text))
    return m.group() if m else None


def _extract_all_pe_numbers(text: str | None) -> list[str]:
    """Extract ALL Program Element (PE) numbers from a text field (LION-102).

    Returns a deduplicated list of all PE numbers found, preserving order.
    Used to capture secondary PE references that _extract_pe_number() misses.
    """
    if not text:
        return []
    seen: set[str] = set()
    result: list[str] = []
    for m in _PE_PATTERN.finditer(str(text)):
        pe = m.group()
        if pe not in seen:
            seen.add(pe)
            result.append(pe)
    return result


def _parse_appropriation(account_title: str | None) -> tuple[str | None, str | None]:
    """Split a combined account_title into (appropriation_code, appropriation_title).

    Many account_title values contain a leading numeric appropriation code
    followed by the title, e.g. "2035 Aircraft Procurement, Army".
    Returns (code, title) where code may be None if the field has no numeric prefix.
    Implements TODO 1.B4-c.
    """
    if not account_title:
        return None, None
    s = str(account_title).strip()
    parts = s.split(None, 1)
    if len(parts) == 2 and parts[0].isdigit():
        return parts[0], parts[1]
    return None, s if s else None


def _detect_currency_year(sheet_name: str, filename: str) -> str:
    """Detect whether amounts are in then-year or constant dollars.

    Checks the sheet name and filename for keywords.  DoD budget exhibits
    default to then-year (nominal) dollars unless explicitly labeled constant.
    Implements TODO 1.B3-b.
    """
    combined = f"{sheet_name} {filename}".lower()
    if "constant" in combined:
        return "constant"
    if "then-year" in combined or "then year" in combined:
        return "then-year"
    # Default: DoD appropriations exhibits are published in then-year dollars
    return "then-year"


def _normalise_fiscal_year(raw: str) -> str:
    """Normalise a raw fiscal-year string to canonical "FY YYYY" format (Step 1.B2-d).

    Handles the three common variants that appear in DoD spreadsheet sheet names:
      - "2026"     → "FY 2026"
      - "FY2026"   → "FY 2026"
      - "FY 2026"  → "FY 2026"  (already canonical, returned as-is)

    If the input contains no recognisable 4-digit year, it is returned unchanged
    so that callers always get a deterministic result.
    """
    m = re.search(r"(20\d{2})", raw)
    if m:
        return f"FY {m.group(1)}"
    return raw


def _detect_amount_unit(rows: list, header_idx: int) -> str:
    """Scan title rows above the header for unit indicators (TODO 1.B3-a).

    Returns "thousands" or "millions" based on keyword matches in any cell
    found in rows[0:header_idx+1].  Defaults to "thousands" if no indicator
    is found, since DoD appropriations exhibits default to that unit.
    """
    _MILLIONS = frozenset([
        "in millions", "$ millions", "($millions)", "($ millions)",
        "millions of dollars", "$ in millions", "in $millions",
    ])
    _THOUSANDS = frozenset([
        "in thousands", "$ thousands", "($thousands)", "($ thousands)",
        "thousands of dollars", "$ in thousands", "in $thousands",
    ])

    for row in rows[:header_idx + 1]:
        for cell in row:
            if cell is None:
                continue
            cell_lower = str(cell).strip().lower()
            for pat in _MILLIONS:
                if pat in cell_lower:
                    return "millions"
            for pat in _THOUSANDS:
                if pat in cell_lower:
                    return "thousands"

    return "thousands"  # Default per DoD convention


# Map exhibit type → budget_type label stored in budget_lines (TODO 1.B3-d)
_EXHIBIT_BUDGET_TYPE: dict[str, str] = {
    "m1": "MilPers",
    "o1": "O&M",
    "p1": "Procurement",
    "p1r": "Procurement",
    "r1": "RDT&E",
    "r2": "RDT&E",
    "rf1": "Revolving",
    "c1": "Construction",
}

# 1.B3-c: Amount type describes what kind of budget authority the amounts represent.
# C-1 (MilCon) uses Congressional authorization; other exhibits use enacted BA or
# the President's budget request (both classified as "budget_authority" for simplicity).
_EXHIBIT_AMOUNT_TYPE: dict[str, str] = {
    "c1": "authorization",       # MilCon: amounts are authorization (not BA)
    "m1": "budget_authority",
    "o1": "budget_authority",
    "p1": "budget_authority",
    "p1r": "budget_authority",
    "r1": "budget_authority",
    "r2": "budget_authority",
    "rf1": "budget_authority",
}


# Note: This function is ~110 lines with three logical sections:
# 1) Common fields mapping (account, organization, budget activity)
# 2) Sub-activity/line-item fields (varies by exhibit type)
# 3) Amount column mapping (FY2024-2026 variants, authorization/appropriation for C-1)
# Current implementation is functional and handles all known exhibit types.
_HEADER_CONTINUATION_WORDS = frozenset([
    "amount", "request", "actual", "enacted", "total", "supplemental",
    "reconciliation", "quantity", "code", "title", "number", "no.",
    "disc.", "disc", "prior year", "current year", "budget year",
])


def _merge_header_rows(header_row: list, next_row: list) -> list[str | None]:
    """Merge a two-row header into a single row (Step 1.B2-c).

    When exhibit sheets split column headers across two rows (e.g., "FY 2026"
    on row N and "Request Amount" on row N+1), this merges them cell-by-cell
    so that _map_columns() receives a single combined header string per column.

    Only cells in next_row that are non-empty and consist entirely of short
    header-like text are merged; if any cell looks like a data value (numeric
    or long string), the function returns the original header_row unchanged to
    avoid accidentally consuming a data row.
    """
    non_empty = [str(v).strip() for v in next_row if v is not None and str(v).strip()]
    if not non_empty:
        return list(header_row)  # All-blank next row — nothing to merge

    # If any non-empty cell is numeric, treat next_row as a data row
    for cell in non_empty:
        try:
            float(str(cell).replace(",", ""))
            return list(header_row)
        except ValueError:
            pass

    # If any cell is longer than 50 chars, likely a narrative data cell
    if any(len(c) > 50 for c in non_empty):
        return list(header_row)

    # Merge pairwise; concatenate with a space where both cells are populated
    merged = []
    for main, sub in zip(header_row, next_row):
        main_s = str(main).strip() if main is not None else ""
        sub_s = str(sub).strip() if sub is not None else ""
        if main_s and sub_s:
            merged.append(f"{main_s} {sub_s}")
        elif main_s:
            merged.append(main_s)
        elif sub_s:
            merged.append(sub_s)
        else:
            merged.append(None)
    # Pad if next_row is shorter than header_row (shouldn't happen, but defensive)
    merged.extend(header_row[len(next_row):])
    return merged


# Future optimization: could split into _map_common_fields, _map_line_item_fields,
# and _map_amount_fields for improved testability and per-exhibit customization.
def _map_columns(headers: list, exhibit_type: str) -> dict:
    """Map column headers to standardized field names.

    Returns a dict mapping our field names to column indices.
    Optimized to use a single pass over headers with fallback logic.
    Returns a dict mapping our field names to column indices. Handles all DoD budget
    exhibit types including P-1, R-1, O-1, M-1, C-1 (MilCon), and RF-1 (Revolving Fund).
    """
    mapping = {}
    h_lower = [str(h).lower().replace("\n", " ").strip() if h else "" for h in headers]

    # Single-pass column mapping with normalized field matching
    for i, h in enumerate(h_lower):
        # Common fields
        if h == "account":
            mapping["account"] = i
        elif h == "account title":
            mapping["account_title"] = i
        elif h == "organization":
            mapping["organization"] = i
        elif h.startswith("budget activity") and "title" not in h:
            mapping["budget_activity"] = i
        elif h == "budget activity title":
            mapping["budget_activity_title"] = i
        elif h == "classification":
            mapping["classification"] = i

        # Sub-activity / line item fields
        elif h in ("bsa", "ag/bsa"):
            mapping["sub_activity"] = i
        elif "budget subactivity" in h and "title" in h:
            mapping["sub_activity_title"] = i
        elif h == "ag/budget subactivity (bsa) title":
            mapping["sub_activity_title"] = i
        elif h == "budget line item" and "title" not in h:
            mapping["line_item"] = i
        elif h in ("budget line item (bli) title",
                    "program element/budget line item (bli) title"):
            mapping["line_item_title"] = i
        elif h == "pe/bli":
            mapping["line_item"] = i
        elif h in ("sag/bli",):
            mapping["line_item"] = i
        elif h == "sag/budget line item (bli) title":
            mapping["line_item_title"] = i
        elif h == "construction project title":
            mapping["line_item_title"] = i
        elif h == "construction project":
            mapping["line_item"] = i
        elif h == "location title":
            mapping.setdefault("sub_activity_title", i)
        elif h == "facility category title":
            mapping.setdefault("sub_activity", i)

    # Amount columns — year-agnostic regex detection (Step 1.B2-a).
    # Matches any "FY YYYY" or "FYXXXX" header and classifies by sub-type
    # keyword.  Canonical column names: amount_fyYYYY_<type>.
    _FY_RE = re.compile(r"fy\s*(\d{4})", re.IGNORECASE)
    for i, h in enumerate(h_lower):
        m = _FY_RE.search(h)
        if not m:
            continue
        year = m.group(1)
        if "quantity" in h:
            if "actual" in h:
                mapping.setdefault(f"quantity_fy{year}", i)
            elif "request" in h or "disc" in h:
                mapping.setdefault(f"quantity_fy{year}_request", i)
            elif "total" in h:
                mapping.setdefault(f"quantity_fy{year}_total", i)
            else:
                mapping.setdefault(f"quantity_fy{year}", i)
        elif "actual" in h:
            mapping.setdefault(f"amount_fy{year}_actual", i)
        elif "enacted" in h or "approp" in h:
            mapping.setdefault(f"amount_fy{year}_enacted", i)
        elif "supplemental" in h:
            mapping.setdefault(f"amount_fy{year}_supplemental", i)
        elif "reconcil" in h:
            mapping.setdefault(f"amount_fy{year}_reconciliation", i)
        elif "total" in h:
            mapping.setdefault(f"amount_fy{year}_total", i)
        elif "request" in h or "disc" in h:
            mapping.setdefault(f"amount_fy{year}_request", i)

    # Authorization/appropriation amounts (C-1 exhibit)
    for i, h in enumerate(h_lower):
        if "authorization amount" in h:
            mapping.setdefault("amount_fy2026_request", i)
        elif "appropriation amount" in h:
            mapping.setdefault("amount_fy2025_enacted", i)
        elif "total obligation authority" in h:
            mapping.setdefault("amount_fy2026_total", i)

    # DONE 1.B2-c: Multi-row header handling implemented.
    #   _merge_header_rows() detects two-row split headers and merges them.
    #   Called from ingest_excel_file() at line ~693. Tests in test_parsing.py
    #   (test_merge_header_rows_* and test_merge_header_rows_two_row_map_columns).

    # Merge catalog-based column detection (Step 1.B2-b).
    # For fields not yet set by heuristic matching above, consult EXHIBIT_CATALOG.
    # The heuristic mapping wins for any field it already identified; the catalog
    # fills in gaps for exhibit types with well-defined column specs (p5, r2, r3, r4,
    # and the summary exhibits already handled above).
    catalog_mapping = _catalog_find_matching_columns(exhibit_type, list(headers))
    # catalog_mapping: col_index → field_name; invert to field_name → col_index
    already_mapped_fields = set(mapping.values())
    for col_idx, field_name in catalog_mapping.items():
        if field_name not in already_mapped_fields:
            mapping.setdefault(field_name, col_idx)

    return mapping


def ingest_excel_file(conn: sqlite3.Connection, file_path: Path,
                      docs_dir: Path | None = None,
                      ensure_columns: bool = True) -> int:
    """Ingest a single Excel file into the database."""
    _docs_dir = (docs_dir or DOCS_DIR).resolve()
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    exhibit_type = _detect_exhibit_type(file_path.name)
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Find the header row in the first 5 rows (buffer to avoid full materialization).
        # When found, read one extra row for two-row header merge detection.
        header_idx = None
        first_rows = []
        for i, row in enumerate(rows_iter):
            first_rows.append(row)
            if header_idx is not None:
                # We already found the header; this extra row is for merge detection
                break
            if i >= 4:
                break
            for val in row:
                if val and str(val).strip().lower() == "account":
                    header_idx = i
                    break

        if header_idx is None:
            continue

        headers = first_rows[header_idx]
        # Detect and merge two-row headers (Step 1.B2-c): some exhibits split
        # "FY 2026" and "Request Amount" across consecutive rows.
        if header_idx + 1 < len(first_rows):
            merged = _merge_header_rows(headers, first_rows[header_idx + 1])
            if merged != list(headers):
                headers = merged

        col_map = _map_columns(headers, exhibit_type)

        if "account" not in col_map:
            continue

        # BUILD-002: detect any FY columns beyond the baseline FY2024-2026 schema
        # and add them dynamically via ALTER TABLE before inserting
        if ensure_columns:
            dynamic_fy_cols = [
                k for k in col_map
                if (k.startswith("amount_fy") or k.startswith("quantity_fy"))
                and k not in _FIXED_AMOUNT_COLUMNS
            ]
            if dynamic_fy_cols:
                _ensure_fy_columns(conn, dynamic_fy_cols)

        # Detect fiscal year from sheet name and normalise to "FY YYYY" (Step 1.B2-d)
        fiscal_year = _normalise_fiscal_year(sheet_name)
        # LION-101: Validate/fallback FY from directory path when sheet name has no FY
        dir_fy = _extract_fy_from_path(file_path)
        if fiscal_year == sheet_name and dir_fy:
            # Sheet name had no recognizable FY (returned unchanged) — use dir path
            fiscal_year = dir_fy
        elif dir_fy and fiscal_year != dir_fy and fiscal_year.startswith("FY "):
            # Both present but disagree — log warning, prefer sheet-derived value
            print(f"  LION-101 WARNING: FY mismatch in {file_path.name}: "
                  f"sheet='{fiscal_year}' vs dir='{dir_fy}' — using sheet value")

        # Detect currency year for this sheet (TODO 1.B3-b)
        currency_year = _detect_currency_year(sheet_name, file_path.name)

        # Detect source unit and compute normalisation multiplier (Step 1.B3-a/c)
        amount_unit = _detect_amount_unit(first_rows, header_idx)
        # All stored amounts must be in thousands; multiply millions-denominated
        # values by 1000 before inserting (Step 1.B3-c)
        unit_multiplier = 1000.0 if amount_unit == "millions" else 1.0

        # Derive budget_type from exhibit type (Step 1.B3-d)
        budget_type = _EXHIBIT_BUDGET_TYPE.get(exhibit_type)

        # Derive amount_type from exhibit type (Step 1.B3-c)
        amount_type = _EXHIBIT_AMOUNT_TYPE.get(exhibit_type, "budget_authority")

        # BUILD-002: Collect all FY/quantity columns detected in this sheet's headers.
        # These will be included in the dynamic INSERT statement below.
        _fy_cols_in_map = sorted([
            k for k in col_map
            if k.startswith("amount_fy") or k.startswith("quantity_fy")
        ])

        batch = []

        def get_val(row, field):
            """Return the raw cell value for a named field, or None if unmapped/out-of-range."""
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        def get_str(row, field):
            """Return the stripped string value for a named field, or None if absent."""
            v = get_val(row, field)
            return str(v).strip() if v is not None else None

        def get_org_name(row):
            """Get organization name, defaulting to code if not in map."""
            org_code = get_str(row, "organization") or ""
            return ORG_MAP.get(org_code, org_code)

        # Process rows after header
        for row in rows_iter:
            if not row or all(v is None for v in row):
                continue

            _acct_idx = col_map.get("account")
            acct = row[_acct_idx] if _acct_idx is not None and _acct_idx < len(row) else None
            if not acct:
                continue

            _org_idx = col_map.get("organization")
            org_code = (
                str(row[_org_idx]).strip()
                if _org_idx is not None and _org_idx < len(row) and row[_org_idx]
                else ""
            )
            # Lookup by exact code first, then by uppercase match (Step 1.B4-b)
            org_name = ORG_MAP.get(org_code) or ORG_MAP.get(org_code.upper(), org_code)

            # Extract PE number from line_item or account fields (Step 1.B4-a implementation)
            line_item_val = get_str(row, "line_item")
            account_val = str(acct).strip()
            pe_number = _extract_pe_number(line_item_val) or _extract_pe_number(account_val)
            # LION-102: Capture additional PE numbers for cross-referencing
            all_pes = _extract_all_pe_numbers(
                f"{line_item_val or ''} {account_val} {get_str(row, 'line_item_title') or ''}")
            additional_pes = [p for p in all_pes if p != pe_number] if pe_number else all_pes[1:]

            # Split appropriation code and title from account_title (Step 1.B4-c implementation)
            acct_title_val = get_str(row, "account_title")
            approp_code, approp_title = _parse_appropriation(acct_title_val)

            def _amt(field):
                """Read a float amount and apply the unit multiplier (1.B3-c)."""
                v = _safe_float(get_val(row, field))
                return v * unit_multiplier if v else v

            # BUILD-002: Build the FY/quantity values dynamically from col_map.
            # Amounts use _amt() for unit normalisation; quantities are unitless.
            fy_values = []
            for fc in _fy_cols_in_map:
                if fc.startswith("amount_"):
                    fy_values.append(_amt(fc))
                else:
                    fy_values.append(_safe_float(get_val(row, fc)))

            batch.append((
                str(file_path.relative_to(_docs_dir)),
                exhibit_type,
                sheet_name,
                fiscal_year,
                account_val,
                acct_title_val,
                org_code,
                org_name,
                get_str(row, "budget_activity"),
                get_str(row, "budget_activity_title"),
                get_str(row, "sub_activity"),
                get_str(row, "sub_activity_title"),
                line_item_val,
                get_str(row, "line_item_title"),
                get_str(row, "classification"),
                *fy_values,    # dynamic FY columns (BUILD-002)
                # LION-102: Store additional PE numbers in extra_fields JSON
                json.dumps({"additional_pe_numbers": additional_pes}) if additional_pes else None,
                pe_number,     # Step 1.B4-a: PE number from line_item or account
                currency_year, # Step 1.B3-b: currency year context
                approp_code,   # Step 1.B4-c: appropriation code from account title
                approp_title,  # Step 1.B4-c: appropriation title from account title
                amount_unit,   # Step 1.B3-b: normalised to "thousands"
                budget_type,   # Step 1.B3-d: budget category from exhibit type
                amount_type,   # Step 1.B3-c: type of amounts (BA, authorization, etc.)
            ))

        if batch:
            # BUILD-002: Use dynamic column list so new FY columns are included
            _fixed_cols = (
                "source_file, exhibit_type, sheet_name, fiscal_year, "
                "account, account_title, organization, organization_name, "
                "budget_activity, budget_activity_title, "
                "sub_activity, sub_activity_title, "
                "line_item, line_item_title, classification"
            )
            _fy_col_str = ", ".join(_fy_cols_in_map) if _fy_cols_in_map else ""
            _tail_cols = (
                "extra_fields, pe_number, currency_year, "
                "appropriation_code, appropriation_title, "
                "amount_unit, budget_type, amount_type"
            )
            all_cols = ", ".join(filter(None, [_fixed_cols, _fy_col_str, _tail_cols]))
            n_params = len(all_cols.split(","))
            placeholders = ", ".join(["?"] * n_params)
            conn.executemany(
                f"INSERT INTO budget_lines ({all_cols}) VALUES ({placeholders})",
                batch,
            )
            total_rows += len(batch)

    wb.close()
    conn.commit()
    return total_rows


# ── OPT-BUILD-001: Standalone Excel extraction worker ─────────────────────────

def _extract_excel_rows(args: tuple) -> dict:
    """Worker function for parallel Excel extraction (OPT-BUILD-001).

    Extracts all budget-line rows from a single Excel file without any
    database access, so it can run safely in a subprocess.

    Args:
        args: Tuple of (file_path_str, docs_dir_str).

    Returns:
        Dict with keys: relative_path, rows (list of tuples), columns (list of str),
        error (str|None), exhibit_type (str).
    """
    file_path_str, docs_dir_str = args
    file_path = Path(file_path_str)
    docs_dir = Path(docs_dir_str)

    exhibit_type = _detect_exhibit_type(file_path.name)
    result: dict = {
        "relative_path": str(file_path.relative_to(docs_dir)),
        "rows": [],
        "columns": [],
        "exhibit_type": exhibit_type,
        "error": None,
    }

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        result["error"] = str(e)
        return result

    all_rows: list[tuple] = []
    all_cols: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        header_idx = None
        first_rows = []
        for i, row in enumerate(rows_iter):
            first_rows.append(row)
            if header_idx is not None:
                break
            if i >= 4:
                break
            for val in row:
                if val and str(val).strip().lower() == "account":
                    header_idx = i
                    break

        if header_idx is None:
            continue

        headers = first_rows[header_idx]
        if header_idx + 1 < len(first_rows):
            merged = _merge_header_rows(headers, first_rows[header_idx + 1])
            if merged != list(headers):
                headers = merged

        col_map = _map_columns(headers, exhibit_type)
        if "account" not in col_map:
            continue

        fiscal_year = _normalise_fiscal_year(sheet_name)
        # LION-101: Validate/fallback FY from directory path in parallel worker
        dir_fy = _extract_fy_from_path(file_path)
        if fiscal_year == sheet_name and dir_fy:
            fiscal_year = dir_fy
        currency_year = _detect_currency_year(sheet_name, file_path.name)
        amount_unit = _detect_amount_unit(first_rows, header_idx)
        unit_multiplier = 1000.0 if amount_unit == "millions" else 1.0
        budget_type = _EXHIBIT_BUDGET_TYPE.get(exhibit_type)
        amount_type = _EXHIBIT_AMOUNT_TYPE.get(exhibit_type, "budget_authority")

        _fy_cols_in_map = sorted([
            k for k in col_map
            if k.startswith("amount_fy") or k.startswith("quantity_fy")
        ])

        # Record which columns this file uses (for dynamic schema management)
        for fc in _fy_cols_in_map:
            if fc not in all_cols:
                all_cols.append(fc)

        def _amt(row, field):
            idx = col_map.get(field)
            v = row[idx] if idx is not None and idx < len(row) else None
            fv = safe_float(v)
            return fv * unit_multiplier if fv else fv

        def _get_str(row, field):
            idx = col_map.get(field)
            v = row[idx] if idx is not None and idx < len(row) else None
            return str(v).strip() if v is not None else None

        rel_path_str = str(file_path.relative_to(docs_dir))

        for row in rows_iter:
            if not row or all(v is None for v in row):
                continue
            _acct_idx = col_map.get("account")
            acct = row[_acct_idx] if _acct_idx is not None and _acct_idx < len(row) else None
            if not acct:
                continue

            _org_idx = col_map.get("organization")
            org_code = (
                str(row[_org_idx]).strip()
                if _org_idx is not None and _org_idx < len(row) and row[_org_idx]
                else ""
            )
            org_name = ORG_MAP.get(org_code) or ORG_MAP.get(org_code.upper(), org_code)
            line_item_val = _get_str(row, "line_item")
            account_val = str(acct).strip()
            pe_number = _extract_pe_number(line_item_val) or _extract_pe_number(account_val)
            # LION-102: Capture additional PE numbers in parallel worker
            line_item_title_val = _get_str(row, "line_item_title")
            all_pes = _extract_all_pe_numbers(
                f"{line_item_val or ''} {account_val} {line_item_title_val or ''}")
            additional_pes = [p for p in all_pes if p != pe_number] if pe_number else all_pes[1:]
            acct_title_val = _get_str(row, "account_title")
            approp_code, approp_title = _parse_appropriation(acct_title_val)

            fy_values = []
            for fc in _fy_cols_in_map:
                if fc.startswith("amount_"):
                    fy_values.append(_amt(row, fc))
                else:
                    idx = col_map.get(fc)
                    v = row[idx] if idx is not None and idx < len(row) else None
                    fy_values.append(safe_float(v))

            all_rows.append((
                rel_path_str, exhibit_type, sheet_name, fiscal_year,
                account_val, acct_title_val, org_code, org_name,
                _get_str(row, "budget_activity"), _get_str(row, "budget_activity_title"),
                _get_str(row, "sub_activity"), _get_str(row, "sub_activity_title"),
                line_item_val, line_item_title_val,
                _get_str(row, "classification"),
                *fy_values,
                # LION-102: additional PEs in extra_fields
                json.dumps({"additional_pe_numbers": additional_pes}) if additional_pes else None,
                pe_number, currency_year, approp_code, approp_title,
                amount_unit, budget_type, amount_type,
            ))

    wb.close()
    result["rows"] = all_rows
    result["columns"] = all_cols
    return result


# ── PDF Ingestion ─────────────────────────────────────────────────────────────

def _extract_table_text(tables: list) -> str:
    """Convert extracted PDF tables to searchable text (optimized)."""
    if not tables:
        return ""

    # Streaming approach: build output incrementally without intermediate lists
    parts = []
    for table in tables:
        if not table:
            continue
        for row in table:
            # Single pass: convert and filter in one go, avoid intermediate list
            cells_str = " | ".join(str(c).strip() for c in row if c)
            if cells_str:  # Only add non-empty rows
                parts.append(cells_str)

    return "\n".join(parts)


def _extract_fy_from_path(file_path: Path) -> str | None:
    """Extract fiscal year from file path directory structure (LION-100).

    Looks for FY directory names like 'FY2026' or 'FY 2026' in the path.
    Returns normalised "FY YYYY" string or None if not found.
    """
    for part in file_path.parts:
        m = re.search(r"FY\s*(\d{4})", part, re.IGNORECASE)
        if m:
            return f"FY {m.group(1)}"
    return None


def _detect_pdf_exhibit_type(filename: str) -> str | None:
    """Detect the exhibit type from a PDF filename (LION-100).

    Reuses the same EXHIBIT_TYPES keys used for Excel files.
    Returns lowercase exhibit code (e.g. 'r2', 'p1') or None.
    """
    name = filename.lower().replace("_display", "")
    for key in sorted(EXHIBIT_TYPES.keys(), key=len, reverse=True):
        if key in name:
            return key
    return None


def _determine_category(file_path: Path) -> str:
    """Determine the budget category from the file path."""
    parts = [p.lower() for p in file_path.parts]
    if "comptroller" in parts:
        return "Comptroller"
    elif "defense_wide" in parts:
        return "Defense-Wide"
    elif "us_army" in parts or "army" in parts:
        return "Army"
    elif "navy" in parts:
        return "Navy"
    elif "air_force" in parts or "airforce" in parts:
        return "Air Force"
    elif "space_force" in parts or "spaceforce" in parts:
        return "Space Force"
    elif "marine_corps" in parts or "marines" in parts:
        return "Marine Corps"
    return "Other"


def _likely_has_tables(page) -> bool:
    """
    Lightweight heuristic to detect if a PDF page likely contains tables.
    Avoids expensive extract_tables() on text-only pages.

    Note: This is a best-effort optimization. False positives (text pages
    extracted as tables) are acceptable since table extraction gracefully
    handles them. False negatives (tables not extracted) are acceptable
    since we prioritize speed over comprehensiveness.
    """
    try:
        # Lightweight approach: just check if page has significant structured content
        # Don't access page.lines directly (it's expensive to compute)
        # Instead, use page.rects and page.curves as proxy for table structure
        rects = len(page.rects) if hasattr(page, 'rects') else 0
        curves = len(page.curves) if hasattr(page, 'curves') else 0

        # Pages with many rectangles or curves likely have tables/structured layouts
        # Threshold is conservative: only skip extraction if very text-like
        return (rects + curves) > 10
    except Exception:
        # If any error computing heuristic, skip table extraction (save time)
        # Missing some tables is better than crashing
        return False


def _extract_tables_with_timeout(page, timeout_seconds=10, executor=None):
    """Extract tables from a PDF page with timeout and progressive fallback strategies.

    Tries three extraction strategies in order, returning the first non-empty result:
      1. lines/lines  — best quality; works for PDFs with visible gridlines
      2. text/lines   — fallback for tables with horizontal rules but no vertical lines
      3. text/text    — last resort for tables with no visible gridlines at all

    Returns (tables, issue_type) where issue_type is None on success via the
    primary strategy, "fallback_text_lines" or "fallback_text_text" for the
    respective fallbacks, or "timeout"/"error" if all strategies fail.

    Args:
        executor: Optional ThreadPoolExecutor to reuse. If None, creates a
            temporary one (slower due to per-call thread pool overhead).
    """
    # 1.B5-b: three strategies in priority order
    _STRATEGIES = [
        ({"vertical_strategy": "lines", "horizontal_strategy": "lines"}, None),
        ({"vertical_strategy": "text", "horizontal_strategy": "lines"}, "fallback_text_lines"),
        ({"vertical_strategy": "text", "horizontal_strategy": "text"}, "fallback_text_text"),
    ]

    own_executor = executor is None
    if own_executor:
        executor = ThreadPoolExecutor(max_workers=1)
    try:
        for settings, issue_label in _STRATEGIES:
            try:
                future = executor.submit(page.extract_tables,
                                         table_settings=settings)
                tables = future.result(timeout=timeout_seconds)
                if tables:
                    return tables, issue_label
            except FuturesTimeoutError:
                return None, "timeout"
            except Exception as e:
                return None, f"error: {str(e)[:50]}"
        # All strategies returned empty
        return [], None
    finally:
        if own_executor:
            executor.shutdown(wait=False)


def ingest_pdf_file(conn: sqlite3.Connection, file_path: Path,
                    page_callback=None, docs_dir: Path | None = None,
                    pdf_timeout: int = 30) -> tuple[int, int]:
    """Ingest a single PDF file into the database.

    Args:
        conn: Database connection.
        file_path: Path to the PDF file.
        page_callback: Optional callable(pages_done, total_pages) called after
            each page is processed, for real-time progress reporting.
            Throttled internally to every 10 pages to reduce overhead.
        docs_dir: Base directory for relative path computation. Defaults to
            the global DOCS_DIR constant.
        pdf_timeout: Seconds to wait for table extraction per page (BUILD-003).

    Returns:
        Tuple of (total_pages_inserted, issue_count).
    """
    _docs_dir = (docs_dir or DOCS_DIR).resolve()
    category = _determine_category(file_path)
    relative_path = str(file_path.relative_to(_docs_dir))
    # LION-100: Extract FY and exhibit_type from file path for direct storage
    pdf_fiscal_year = _extract_fy_from_path(file_path)
    pdf_exhibit_type = _detect_pdf_exhibit_type(file_path.name)
    total_pages = 0

    try:
        with pdfplumber.open(str(file_path)) as pdf:
            num_pages = len(pdf.pages)
            batch = []
            issues_batch = []  # Accumulate issues; insert once at end
            # Reuse a single ThreadPoolExecutor across all pages to avoid
            # per-page thread pool creation overhead (thousands of pages per file)
            table_executor = ThreadPoolExecutor(max_workers=1)
            try:
              for i, page in enumerate(pdf.pages):
                try:
                    # Use layout=False for 30-50% faster extraction (we don't need positioning)
                    text = page.extract_text(layout=False) or ""
                except Exception as e:
                    # Skip pages with font extraction errors (malformed FontBBox, etc.)
                    if "FontBBox" in str(e) or "cannot be parsed" in str(e):
                        text = ""
                    else:
                        raise

                # Use heuristic to skip expensive table extraction on text-only pages.
                # _likely_has_tables checks for rect/curve elements as proxy for table structure.
                tables = None
                issue_type = None
                if _likely_has_tables(page):
                    tables, issue_type = _extract_tables_with_timeout(
                        page, timeout_seconds=pdf_timeout, executor=table_executor)
                    if issue_type:
                        issues_batch.append(
                            (relative_path, i + 1, issue_type.split(':')[0], issue_type)
                        )
                        tables = None

                table_text = _extract_table_text(tables)

                # Skip truly empty pages
                if not text.strip() and not table_text.strip():
                    if page_callback and (i % 10 == 0 or i == num_pages - 1):
                        page_callback(i + 1, num_pages)
                    continue

                batch.append((
                    relative_path,
                    category,
                    pdf_fiscal_year,     # LION-100
                    pdf_exhibit_type,    # LION-100
                    i + 1,
                    text,
                    1 if tables else 0,
                    table_text if table_text else None,
                ))

                # Throttle callback: fire every 10 pages instead of every page
                if page_callback and (i % 10 == 0 or i == num_pages - 1):
                    page_callback(i + 1, num_pages)

                # Batch insert every 1000 pages (larger batch = fewer executemany calls)
                if len(batch) >= 1000:
                    conn.executemany("""
                        INSERT INTO pdf_pages (source_file, source_category,
                            fiscal_year, exhibit_type,
                            page_number, page_text, has_tables, table_data)
                        VALUES (?,?,?,?,?,?,?,?)
                    """, batch)
                    total_pages += len(batch)
                    batch = []

              if batch:
                conn.executemany("""
                    INSERT INTO pdf_pages (source_file, source_category,
                        fiscal_year, exhibit_type,
                        page_number, page_text, has_tables, table_data)
                    VALUES (?,?,?,?,?,?,?,?)
                """, batch)
                total_pages += len(batch)

              # Batch insert all issues at once (avoids per-issue execute overhead)
              if issues_batch:
                conn.executemany(
                    "INSERT INTO extraction_issues"
                    " (file_path, page_number, issue_type, issue_detail)"
                    " VALUES (?,?,?,?)",
                    issues_batch)

              # LION-103: Populate pdf_pe_numbers junction table by scanning
              # inserted pages for PE number mentions.
              pe_junction_rows = []
              for row in conn.execute(
                  "SELECT id, page_number, page_text FROM pdf_pages "
                  "WHERE source_file = ? AND page_text IS NOT NULL",
                  (relative_path,)
              ):
                  page_id, page_num, page_text = row
                  found_pes = _extract_all_pe_numbers(page_text)
                  for pe in found_pes:
                      pe_junction_rows.append((
                          page_id, pe, page_num, relative_path, pdf_fiscal_year))
              if pe_junction_rows:
                  conn.executemany(
                      "INSERT INTO pdf_pe_numbers "
                      "(pdf_page_id, pe_number, page_number, source_file, fiscal_year) "
                      "VALUES (?,?,?,?,?)",
                      pe_junction_rows)

            finally:
                table_executor.shutdown(wait=False)

    except Exception as e:
        print(f"  ERROR processing {file_path.name}: {e}")
        conn.execute(
            "INSERT OR REPLACE INTO ingested_files "
            "(file_path, file_type, file_size, file_modified,"
            " ingested_at, row_count, status) "
            "VALUES (?,?,?,?,datetime('now'),?,?)",
            (relative_path, "pdf", file_path.stat().st_size,
             file_path.stat().st_mtime, 0, f"error: {e}")
        )
        return 0, 0

    # Caller handles commit; do not commit here to allow larger transaction batches
    return total_pages, len(issues_batch)


def _extract_pdf_data(args):
    """Worker function for parallel PDF extraction (runs in a separate process).

    Extracts text and table data from all pages of a single PDF file.
    No database access — returns raw data for the main process to insert.

    Args:
        args: Tuple of (file_path_str, docs_dir_str, pdf_timeout) for picklability.

    Returns:
        Dict with keys: relative_path, category, pages_data, issues, error,
        num_pages.
    """
    if len(args) == 3:
        file_path_str, docs_dir_str, pdf_timeout = args
    else:
        file_path_str, docs_dir_str = args
        pdf_timeout = 30
    file_path = Path(file_path_str)
    docs_dir = Path(docs_dir_str)

    category = _determine_category(file_path)
    relative_path = str(file_path.relative_to(docs_dir))
    # LION-100: Extract FY and exhibit_type from file path
    pdf_fiscal_year = _extract_fy_from_path(file_path)
    pdf_exhibit_type = _detect_pdf_exhibit_type(file_path.name)
    pages_data = []
    pe_mentions = []  # LION-103: (pe_number, page_number) tuples
    issues = []
    num_pages = 0

    try:
        with pdfplumber.open(file_path_str) as pdf:
            num_pages = len(pdf.pages)
            executor = ThreadPoolExecutor(max_workers=1)
            try:
                for i, page in enumerate(pdf.pages):
                    try:
                        text = page.extract_text(layout=False) or ""
                    except Exception as e:
                        if "FontBBox" in str(e) or "cannot be parsed" in str(e):
                            text = ""
                        else:
                            raise

                    tables = None
                    if _likely_has_tables(page):
                        tables, issue_type = _extract_tables_with_timeout(
                            page, timeout_seconds=pdf_timeout, executor=executor)
                        if issue_type:
                            issues.append((
                                relative_path, i + 1,
                                issue_type.split(':')[0], issue_type
                            ))
                            tables = None

                    table_text = _extract_table_text(tables)

                    if not text.strip() and not table_text.strip():
                        continue

                    page_num = i + 1
                    pages_data.append((
                        relative_path, category,
                        pdf_fiscal_year,     # LION-100
                        pdf_exhibit_type,    # LION-100
                        page_num, text,
                        1 if tables else 0,
                        table_text if table_text else None,
                    ))
                    # LION-103: Extract PE numbers for junction table
                    found_pes = _extract_all_pe_numbers(text)
                    for pe in found_pes:
                        pe_mentions.append((pe, page_num))
            finally:
                executor.shutdown(wait=False)
    except Exception as e:
        return {
            "relative_path": relative_path,
            "category": category,
            "fiscal_year": pdf_fiscal_year,
            "pages_data": [],
            "pe_mentions": [],
            "issues": issues,
            "error": str(e),
            "num_pages": num_pages,
        }

    return {
        "relative_path": relative_path,
        "category": category,
        "fiscal_year": pdf_fiscal_year,
        "pages_data": pages_data,
        "pe_mentions": pe_mentions,
        "issues": issues,
        "error": None,
        "num_pages": num_pages,
    }


# ── Main Build Pipeline ───────────────────────────────────────────────────────

def _file_needs_update(conn: sqlite3.Connection, rel_path: str,
                       file_path: Path,
                       cache: dict | None = None) -> bool:
    """Check if a file needs to be (re)ingested based on size and mtime.

    If cache is provided (dict mapping rel_path -> (size, mtime)), uses it
    for an O(1) in-memory lookup instead of a DB round-trip.  The cache is
    pre-populated once before the main loops to avoid ~5000 individual
    indexed SELECTs across the xlsx + pdf file lists.
    """
    stat = file_path.stat()
    if cache is not None:
        row = cache.get(rel_path)
    else:
        row = conn.execute(
            "SELECT file_size, file_modified FROM ingested_files"
            " WHERE file_path = ?",
            (rel_path,)
        ).fetchone()

    if row is None:
        return True  # New file
    if row[0] != stat.st_size or abs((row[1] or 0) - stat.st_mtime) > 1:
        return True  # Modified file
    return False


def _remove_file_data(conn: sqlite3.Connection, rel_path: str, file_type: str):
    """Remove previously ingested data for a file before re-ingesting."""
    if file_type == "xlsx":
        conn.execute("DELETE FROM budget_lines WHERE source_file = ?", (rel_path,))
    elif file_type == "pdf":
        conn.execute("DELETE FROM pdf_pages WHERE source_file = ?", (rel_path,))
        # LION-103: Also remove junction table rows for this file
        conn.execute("DELETE FROM pdf_pe_numbers WHERE source_file = ?", (rel_path,))


def _recreate_pdf_fts_triggers(conn: sqlite3.Connection):
    """Recreate FTS5 triggers for pdf_pages table.

    Used after bulk operations to maintain FTS5 index synchronization.
    """
    conn.execute("""
        CREATE TRIGGER IF NOT EXISTS pdf_pages_ai AFTER INSERT ON pdf_pages BEGIN
            INSERT INTO pdf_pages_fts(rowid, page_text, source_file, table_data)
            VALUES (new.id, new.page_text, new.source_file, new.table_data);
        END
    """)
    conn.execute("""
        CREATE TRIGGER IF NOT EXISTS pdf_pages_ad AFTER DELETE ON pdf_pages BEGIN
            INSERT INTO pdf_pages_fts(pdf_pages_fts, rowid, page_text, source_file, table_data)
            VALUES ('delete', old.id, old.page_text, old.source_file, old.table_data);
        END
    """)


def _register_data_source(conn: sqlite3.Connection, docs_dir: Path):
    """Auto-register data sources from directory structure."""
    for fy_dir in sorted(docs_dir.iterdir()):
        if not fy_dir.is_dir() or not fy_dir.name.startswith("FY"):
            continue
        fiscal_year = fy_dir.name
        for src_dir in sorted(fy_dir.iterdir()):
            if not src_dir.is_dir():
                continue
            source_id = f"{fiscal_year}/{src_dir.name}"
            file_count = sum(1 for _ in src_dir.rglob("*") if _.is_file())
            conn.execute("""
                INSERT INTO data_sources (source_id, source_name, fiscal_year,
                    last_checked, file_count)
                VALUES (?, ?, ?, datetime('now'), ?)
                ON CONFLICT(source_id) DO UPDATE SET
                    last_checked = datetime('now'),
                    file_count = excluded.file_count
            """, (source_id, src_dir.name, fiscal_year, file_count))
    conn.commit()


# ── Checkpoint Management ────────────────────────────────────────────────────

def _create_session_id() -> str:
    """Create a unique session ID for this build session."""
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    unique_id = str(uuid.uuid4())[:8]
    return f"sess-{timestamp}-{unique_id}"


def _save_checkpoint(conn: sqlite3.Connection, session_id: str, files_processed: int,
                     total_files: int, pages_processed: int, rows_inserted: int,
                     bytes_processed: int, last_file: str = None,
                     last_file_status: str = None, notes: str = None,
                     commit: bool = True) -> None:
    """Save current build progress as a checkpoint.

    This allows resuming from a checkpoint without reprocessing files.
    commit=True by default because checkpoints are explicit sync points;
    pass commit=False when the caller will commit imminently anyway.
    """
    conn.execute("""
        INSERT INTO build_progress
        (session_id, files_processed, total_files, pages_processed, rows_inserted,
         bytes_processed, status, last_file, last_file_status, notes, checkpoint_time)
        VALUES (?, ?, ?, ?, ?, ?, 'in_progress', ?, ?, ?, datetime('now'))
        ON CONFLICT(session_id) DO UPDATE SET
            files_processed = excluded.files_processed,
            total_files = excluded.total_files,
            pages_processed = excluded.pages_processed,
            rows_inserted = excluded.rows_inserted,
            bytes_processed = excluded.bytes_processed,
            last_file = excluded.last_file,
            last_file_status = excluded.last_file_status,
            notes = excluded.notes,
            checkpoint_time = datetime('now')
    """, (session_id, files_processed, total_files, pages_processed, rows_inserted,
          bytes_processed, last_file, last_file_status, notes))
    if commit:
        conn.commit()


def _mark_file_processed(conn: sqlite3.Connection, session_id: str, file_path: str,
                         file_type: str, rows_count: int = 0, pages_count: int = 0,
                         commit: bool = False) -> None:
    """Mark a file as processed in the current session.

    commit=False (default) defers the commit to the caller's batch commit,
    avoiding a WAL sync on every single file.
    """
    conn.execute("""
        INSERT INTO processed_files
        (session_id, file_path, file_type, rows_count, pages_count, processed_at)
        VALUES (?, ?, ?, ?, ?, datetime('now'))
    """, (session_id, str(file_path), file_type, rows_count, pages_count))
    if commit:
        conn.commit()


def _get_last_checkpoint(conn: sqlite3.Connection) -> dict | None:
    """Get the last checkpoint for resuming.

    Returns dict with session info or None if no checkpoint found.
    """
    cursor = conn.execute("""
        SELECT * FROM build_progress
        WHERE status = 'in_progress'
        ORDER BY checkpoint_time DESC
        LIMIT 1
    """)
    row = cursor.fetchone()
    if not row:
        return None

    # Convert to dict
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, row))


def _get_processed_files(conn: sqlite3.Connection, session_id: str) -> set:
    """Get set of file paths already processed in session."""
    cursor = conn.execute("""
        SELECT file_path FROM processed_files WHERE session_id = ?
    """, (session_id,))
    return {row[0] for row in cursor.fetchall()}


def _mark_session_complete(conn: sqlite3.Connection, session_id: str, notes: str = None) -> None:
    """Mark a session as completed."""
    conn.execute("""
        UPDATE build_progress
        SET status = 'completed', notes = ?
        WHERE session_id = ?
    """, (notes, session_id))
    conn.commit()


def build_database(docs_dir: Path, db_path: Path, rebuild: bool = False,
                   progress_callback=None, resume: bool = False,
                   checkpoint_interval: int = 10,
                   stop_event=None, workers: int = 0,
                   pdf_timeout: int = 30,
                   failures_log: Path | None = None,
                   retry_failures: bool = False):
    """Build or incrementally update the budget database.

    Args:
        docs_dir: Path to the DoD_Budget_Documents directory.
        db_path: Path for the SQLite database file.
        rebuild: If True, delete existing database and rebuild from scratch.
        progress_callback: Optional callable(phase, current, total, detail, metrics)
            where phase is 'scan', 'excel', 'pdf', 'index', or 'done',
            current/total are progress counts, detail is a status string, and
            metrics is a dict with keys: rows, pages, speed_rows, speed_pages,
            eta_sec, files_remaining, current_pages, current_total_pages.
        resume: If True, attempt to resume from last checkpoint.
        checkpoint_interval: Save a checkpoint every N files (default 10).
        stop_event: Optional threading.Event; when set, the build saves a
            checkpoint and exits cleanly (graceful shutdown).
        workers: Number of parallel worker processes for PDF extraction.
            0 = auto-detect (CPU count, capped at 4). 1 = sequential (no
            multiprocessing overhead).
        pdf_timeout: Seconds to wait for table extraction per page (BUILD-003).
        failures_log: Path to write failed_downloads.json (BUILD-001).
        retry_failures: If True, only process files listed in failures_log (BUILD-001).
    """
    # ── Metrics state shared across the build ─────────────────────────────
    _metrics = {
        "rows": 0,
        "pages": 0,
        "speed_rows": 0.0,    # rows/sec (exponentially smoothed)
        "speed_pages": 0.0,   # pages/sec (exponentially smoothed)
        "eta_sec": 0.0,
        "files_remaining": 0,
        "current_pages": 0,       # pages in current PDF
        "current_total_pages": 0, # total pages in current PDF
    }

    def _progress(phase, current, total, detail="", metrics_update=None):
        """Update shared metrics and invoke the caller-supplied progress callback."""
        if metrics_update:
            _metrics.update(metrics_update)
        if progress_callback:
            progress_callback(phase, current, total, detail, dict(_metrics))

    def _update_speed(key, new_value, alpha=0.3):
        """Exponential moving average for speed tracking."""
        if _metrics[key] == 0.0:
            _metrics[key] = new_value
        else:
            _metrics[key] = alpha * new_value + (1 - alpha) * _metrics[key]

    # ── BUILD-001: Failure tracking ────────────────────────────────────────
    _failures: list[FailedFileEntry] = []
    _failures_log_path = failures_log or Path("failed_downloads.json")

    # If --retry-failures, load the failures JSON and use it as the file filter
    _retry_only: set[str] | None = None
    if retry_failures and _failures_log_path.exists():
        try:
            with open(_failures_log_path) as _f:
                _prev_failures = json.load(_f)
            _retry_only = {e["file_path"] for e in _prev_failures if "file_path" in e}
            print(f"  BUILD-001: Retrying {len(_retry_only)} previously-failed file(s) "
                  f"from {_failures_log_path}")
        except Exception as _lf_err:
            print(f"  BUILD-001: Could not load failures log: {_lf_err}")
            _retry_only = None

    # ── Setup ──────────────────────────────────────────────────────────────
    if not docs_dir.exists():
        _progress("error", 0, 0, f"Documents directory not found: {docs_dir}")
        raise FileNotFoundError(f"Documents directory not found: {docs_dir}")

    if rebuild and db_path.exists():
        db_path.unlink()
        print(f"Removed existing database for rebuild: {db_path}")

    is_new = not db_path.exists()
    conn = create_database(db_path)

    if is_new:
        print(f"Created new database: {db_path}")
    else:
        print(f"Updating existing database: {db_path}")

    # ── Session and resume setup ───────────────────────────────────────────
    session_id = None
    already_processed: set = set()

    if resume and not rebuild:
        checkpoint = _get_last_checkpoint(conn)
        if checkpoint:
            session_id = checkpoint["session_id"]
            already_processed = _get_processed_files(conn, session_id)
            print(f"\nResuming session: {session_id}")
            print(f"  Already processed: {len(already_processed)} file(s)")
        else:
            print("\nNo checkpoint found — starting fresh build.")

    if session_id is None:
        session_id = _create_session_id()

    # ── Scan ───────────────────────────────────────────────────────────────
    _progress("scan", 0, 0, "Scanning directories...")
    _register_data_source(conn, docs_dir)

    xlsx_files = sorted(docs_dir.rglob("*.xlsx"))
    pdf_files = sorted(docs_dir.rglob("*.pdf"))

    # BUILD-001: If retrying only failures, filter file lists
    if _retry_only is not None:
        docs_dir_resolved = docs_dir.resolve()
        xlsx_files = [
            f for f in xlsx_files
            if str(f.relative_to(docs_dir_resolved)) in _retry_only
        ]
        pdf_files = [
            f for f in pdf_files
            if str(f.relative_to(docs_dir_resolved)) in _retry_only
        ]
        print(f"  BUILD-001: Filtered to {len(xlsx_files)} Excel + {len(pdf_files)} PDF retry files")

    total_files = len(xlsx_files) + len(pdf_files)

    # Pre-load all ingested_files metadata in one query so that
    # _file_needs_update() can do O(1) dict lookups instead of one
    # indexed SELECT per file (~5000 round-trips across xlsx + pdf).
    _ingested_cache: dict[str, tuple] = {
        row[0]: (row[1], row[2])
        for row in conn.execute(
            "SELECT file_path, file_size, file_modified FROM ingested_files"
        ).fetchall()
    }

    # Save initial checkpoint so the session exists in build_progress from the start
    _save_checkpoint(conn, session_id, 0, total_files, 0, 0, 0,
                     notes="Build started")

    _progress("scan", 0, total_files,
              f"Found {len(xlsx_files)} Excel + {len(pdf_files)} PDF files",
              {"files_remaining": total_files - len(already_processed)})
    print(f"\nFound {len(xlsx_files)} Excel files and {len(pdf_files)} PDF files")

    if already_processed:
        xl_skipped_resume = sum(1 for f in xlsx_files
                                if str(f.relative_to(docs_dir)) in already_processed)
        pdf_skipped_resume = sum(1 for f in pdf_files
                                 if str(f.relative_to(docs_dir)) in already_processed)
        print(f"  Resuming: skipping {xl_skipped_resume} Excel + {pdf_skipped_resume} PDF "
              f"files already processed in session {session_id}")

    # ── Ingest Excel files ─────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("  INGESTING EXCEL FILES")
    print(f"{'='*60}")

    # Resolve worker count early (used for both Excel and PDF)
    num_workers = workers if workers > 0 else min(os.cpu_count() or 1, 4)

    total_budget_rows = 0
    skipped_xlsx = 0
    excel_file_times: list[float] = []  # per-file elapsed times for speed calc
    files_done_total = 0  # across both xlsx and pdf
    initial_bl_count = conn.execute(
        "SELECT COUNT(*) FROM budget_lines"
    ).fetchone()[0]

    # Drop budget_lines FTS5 triggers before bulk insert, matching the PDF
    # optimisation. Triggers fire on every row INSERT; with 100k+ rows across
    # all Excel files the overhead is significant. We rebuild the FTS index in
    # one pass after all files are ingested.
    conn.execute("DROP TRIGGER IF EXISTS budget_lines_ai")
    conn.execute("DROP TRIGGER IF EXISTS budget_lines_ad")
    conn.commit()

    # OPT-BUILD-001: Determine which files need processing
    xlsx_to_process: list[Path] = []
    xlsx_skip_list: list[Path] = []
    for xlsx in xlsx_files:
        rel_path = str(xlsx.relative_to(docs_dir))
        if rel_path in already_processed:
            xlsx_skip_list.append(xlsx)
        elif not rebuild and not _file_needs_update(conn, rel_path, xlsx):
            xlsx_skip_list.append(xlsx)
        else:
            xlsx_to_process.append(xlsx)

    skipped_xlsx = len(xlsx_skip_list)
    files_done_total = skipped_xlsx

    # Emit individual "Resumed (skipped)" progress messages for compatibility
    for _si, _skipped_xl in enumerate(xlsx_skip_list):
        _skipped_rel = str(_skipped_xl.relative_to(docs_dir))
        _detail = (
            f"Resumed (skipped): {_skipped_xl.name}"
            if _skipped_rel in already_processed
            else f"Skipped (unchanged): {_skipped_xl.name}"
        )
        _progress("excel", _si + 1, len(xlsx_files), _detail,
                  {"files_remaining": total_files - _si - 1})

    # OPT-BUILD-001: Use parallel extraction when workers > 1 and enough files
    _excel_use_parallel = num_workers > 1 and len(xlsx_to_process) > 1

    if _excel_use_parallel:
        # Pre-clean: remove old data for files being re-processed to avoid
        # duplicate budget_lines. Mirrors the PDF pre-clean pattern.
        if xlsx_to_process:
            _xl_rel_paths = [str(xl.relative_to(docs_dir)) for xl in xlsx_to_process]
            conn.execute("CREATE TEMP TABLE IF NOT EXISTS _xl_preclean (path TEXT PRIMARY KEY)")
            conn.executemany("INSERT OR IGNORE INTO _xl_preclean VALUES (?)",
                             [(p,) for p in _xl_rel_paths])
            _xl_del = conn.execute(
                "DELETE FROM budget_lines WHERE source_file IN "
                "(SELECT path FROM _xl_preclean)"
            ).rowcount
            conn.execute("DROP TABLE _xl_preclean")
            conn.commit()
            if _xl_del:
                print(f"  Pre-cleaned {_xl_del} stale budget_lines rows for re-ingestion")

        print(f"  Processing {len(xlsx_to_process)} Excel files with "
              f"{num_workers} parallel workers (OPT-BUILD-001)...")
        t_excel_start = time.time()
        with ProcessPoolExecutor(max_workers=num_workers) as pool:
            future_to_xl = {
                pool.submit(_extract_excel_rows, (str(xl), str(docs_dir))): xl
                for xl in xlsx_to_process
            }
            for xi, future in enumerate(as_completed(future_to_xl)):
                if stop_event and stop_event.is_set():
                    for f in future_to_xl:
                        f.cancel()
                    _save_checkpoint(conn, session_id, files_done_total, total_files,
                                     _metrics["pages"], total_budget_rows, 0, "", "interrupted")
                    conn.commit()
                    _progress("stopped", xi, len(xlsx_to_process), "Stopped — resume with --resume")
                    conn.close()
                    return

                xl = future_to_xl[future]
                rel_path = str(xl.relative_to(docs_dir))
                files_done_total += 1
                try:
                    result = future.result(timeout=120)
                except Exception as _xl_err:
                    print(f"  ERROR: {xl.name}: {_xl_err}")
                    _failures.append(FailedFileEntry(
                        file_path=rel_path,
                        error_type=type(_xl_err).__name__,
                        error_detail=str(_xl_err),
                    ))
                    conn.execute(
                        "INSERT OR REPLACE INTO ingested_files "
                        "(file_path, file_type, file_size, file_modified,"
                        " ingested_at, row_count, status) "
                        "VALUES (?,?,?,?,datetime('now'),?,?)",
                        (rel_path, "xlsx", xl.stat().st_size, xl.stat().st_mtime,
                         0, f"error: {_xl_err}"))
                    continue

                if result.get("error"):
                    _failures.append(FailedFileEntry(
                        file_path=rel_path,
                        error_type="ParseError",
                        error_detail=result["error"],
                    ))
                    continue

                rows = result["rows"]
                fy_cols = result["columns"]
                if fy_cols:
                    _ensure_fy_columns(conn, fy_cols)

                if rows:
                    # Reconstruct the INSERT dynamically based on extracted columns
                    _fixed_c = (
                        "source_file, exhibit_type, sheet_name, fiscal_year, "
                        "account, account_title, organization, organization_name, "
                        "budget_activity, budget_activity_title, "
                        "sub_activity, sub_activity_title, "
                        "line_item, line_item_title, classification"
                    )
                    _fy_c = ", ".join(sorted(fy_cols)) if fy_cols else ""
                    _tail_c = (
                        "extra_fields, pe_number, currency_year, "
                        "appropriation_code, appropriation_title, "
                        "amount_unit, budget_type, amount_type"
                    )
                    all_c = ", ".join(filter(None, [_fixed_c, _fy_c, _tail_c]))
                    n_p = len(all_c.split(","))
                    ph = ", ".join(["?"] * n_p)
                    conn.executemany(f"INSERT INTO budget_lines ({all_c}) VALUES ({ph})", rows)
                    total_budget_rows += len(rows)
                    _metrics["rows"] = total_budget_rows

                stat = xl.stat()
                conn.execute(
                    "INSERT OR REPLACE INTO ingested_files "
                    "(file_path, file_type, file_size, file_modified,"
                    " ingested_at, row_count, status) "
                    "VALUES (?,?,?,?,datetime('now'),?,?)",
                    (rel_path, "xlsx", stat.st_size, stat.st_mtime, len(rows), "ok"))
                _mark_file_processed(conn, session_id, rel_path, "excel", rows_count=len(rows))
                elapsed = time.time() - t_excel_start
                excel_file_times.append(elapsed / max(xi + 1, 1))
                print(f"  [{xi+1}/{len(xlsx_to_process)}] {xl.name}: {len(rows)} rows")
                _progress("excel", xi + 1 + skipped_xlsx, len(xlsx_files),
                          f"Done: {xl.name} ({len(rows)} rows)",
                          {"rows": total_budget_rows,
                           "files_remaining": total_files - files_done_total})
        conn.commit()
        if skipped_xlsx:
            print(f"\n  Skipped {skipped_xlsx} unchanged Excel file(s)")
        print(f"  Ingested budget line items: {total_budget_rows:,}")

    if not _excel_use_parallel:
     for xi, xlsx in enumerate(xlsx_to_process):
        # Check for graceful shutdown request
        if stop_event and stop_event.is_set():
            print("\n  Graceful stop requested — saving checkpoint...")
            _save_checkpoint(conn, session_id, files_done_total, total_files,
                             _metrics["pages"], _metrics["rows"],
                             0, str(xlsx), "interrupted")
            conn.commit()
            _progress("stopped", xi, len(xlsx_files),
                      f"Stopped at {xlsx.name} — resume with --resume",
                      {"files_remaining": total_files - files_done_total})
            conn.close()
            return

        rel_path = str(xlsx.relative_to(docs_dir))

        # Skip if already processed in a resumed session
        if rel_path in already_processed:
            skipped_xlsx += 1
            files_done_total += 1
            _progress("excel", xi + 1, len(xlsx_files),
                      f"Resumed (skipped): {xlsx.name}",
                      {"files_remaining": total_files - files_done_total})
            continue

        if not rebuild and not _file_needs_update(
                conn, rel_path, xlsx, cache=_ingested_cache):
            skipped_xlsx += 1
            files_done_total += 1
            _progress("excel", xi + 1, len(xlsx_files),
                      f"Skipped (unchanged): {xlsx.name}",
                      {"files_remaining": total_files - files_done_total})
            continue

        _progress("excel", xi + 1, len(xlsx_files),
                  f"Processing: {xlsx.name}",
                  {"files_remaining": total_files - files_done_total})
        print(f"  Processing: {xlsx.name}...", end=" ", flush=True)

        _remove_file_data(conn, rel_path, "xlsx")

        t0 = time.time()
        try:
            rows = ingest_excel_file(conn, xlsx, docs_dir=docs_dir)
        except Exception as _xl_err:
            file_elapsed = time.time() - t0
            print(f"ERROR ({file_elapsed:.1f}s): {_xl_err}")
            _failures.append(FailedFileEntry(
                file_path=rel_path,
                error_type=type(_xl_err).__name__,
                error_detail=str(_xl_err),
            ))
            conn.execute(
                "INSERT OR REPLACE INTO ingested_files "
                "(file_path, file_type, file_size, file_modified,"
                " ingested_at, row_count, status) "
                "VALUES (?,?,?,?,datetime('now'),?,?)",
                (rel_path, "xlsx", xlsx.stat().st_size, xlsx.stat().st_mtime,
                 0, f"error: {_xl_err}")
            )
            files_done_total += 1
            continue
        file_elapsed = time.time() - t0
        print(f"{rows} rows ({file_elapsed:.1f}s)")

        # Update speed tracking
        if file_elapsed > 0 and rows > 0:
            _update_speed("speed_rows", rows / file_elapsed)
        excel_file_times.append(file_elapsed)

        stat = xlsx.stat()
        conn.execute(
            "INSERT OR REPLACE INTO ingested_files "
            "(file_path, file_type, file_size, file_modified,"
            " ingested_at, row_count, status) "
            "VALUES (?,?,?,?,datetime('now'),?,?)",
            (rel_path, "xlsx", stat.st_size, stat.st_mtime, rows, "ok")
        )
        total_budget_rows += rows
        files_done_total += 1
        _metrics["rows"] = total_budget_rows

        # Post-file progress update with updated row count
        _progress("excel", xi + 1, len(xlsx_files),
                  f"Done: {xlsx.name} ({rows} rows)",
                  {"rows": total_budget_rows,
                   "files_remaining": total_files - files_done_total})

        # Track file as processed; defer commit to batch boundary below
        _mark_file_processed(conn, session_id, rel_path, "excel", rows_count=rows)
        if files_done_total % checkpoint_interval == 0:
            remaining = total_files - files_done_total
            # ETA: average per-file time × files remaining
            avg_time = sum(excel_file_times) / len(excel_file_times) if excel_file_times else 0
            _metrics["eta_sec"] = avg_time * remaining
            _metrics["files_remaining"] = remaining
            _save_checkpoint(conn, session_id, files_done_total, total_files,
                             _metrics["pages"], total_budget_rows,
                             stat.st_size, rel_path, "ok")
        elif xi % 5 == 4 or xi == len(xlsx_files) - 1:
            # Commit every 5 Excel files even outside checkpoint interval
            conn.commit()

    if not _excel_use_parallel:
        conn.commit()

    # Rebuild budget_lines FTS5 index in one pass and restore triggers,
    # mirroring what we do for pdf_pages after bulk PDF ingestion.
    final_bl_count = conn.execute(
        "SELECT COUNT(*) FROM budget_lines"
    ).fetchone()[0]
    if final_bl_count > initial_bl_count:
        print("  Rebuilding budget_lines FTS index...")
        conn.execute(
            "INSERT INTO budget_lines_fts(budget_lines_fts) VALUES('rebuild')"
        )
    conn.executescript("""
        CREATE TRIGGER IF NOT EXISTS budget_lines_ai
        AFTER INSERT ON budget_lines BEGIN
            INSERT INTO budget_lines_fts(
                rowid, account_title, budget_activity_title,
                sub_activity_title, line_item_title,
                organization_name, pe_number)
            VALUES (
                new.id, new.account_title, new.budget_activity_title,
                new.sub_activity_title, new.line_item_title,
                new.organization_name, new.pe_number);
        END;
        CREATE TRIGGER IF NOT EXISTS budget_lines_ad
        AFTER DELETE ON budget_lines BEGIN
            INSERT INTO budget_lines_fts(
                budget_lines_fts, rowid, account_title,
                budget_activity_title, sub_activity_title,
                line_item_title, organization_name, pe_number)
            VALUES (
                'delete', old.id, old.account_title,
                old.budget_activity_title, old.sub_activity_title,
                old.line_item_title, old.organization_name,
                old.pe_number);
        END;
    """)
    conn.commit()

    if not _excel_use_parallel:
        if skipped_xlsx:
            print(f"\n  Skipped {skipped_xlsx} unchanged Excel file(s)")
        print(f"  Ingested budget line items: {total_budget_rows:,}")

    # ── Ingest PDF files ───────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("  INGESTING PDF FILES")
    print(f"{'='*60}")

    total_pdf_pages = 0
    skipped_pdf = 0
    processed_pdf = 0
    last_commit_time = time.time()
    initial_page_count = conn.execute("SELECT COUNT(*) FROM pdf_pages").fetchone()[0]
    pdf_file_times: list[float] = []  # per-file elapsed for speed calc
    _pdf_stopped = False  # set when a graceful stop fires inside the PDF loop
    # num_workers already set above (shared for both Excel and PDF phases)

    # ── Filter PDFs that need processing ──────────────────────────────
    pdfs_to_process: list[Path] = []
    for pdf in pdf_files:
        rel_path = str(pdf.relative_to(docs_dir))
        if rel_path in already_processed:
            skipped_pdf += 1
            files_done_total += 1
            continue
        if not rebuild and not _file_needs_update(
                conn, rel_path, pdf, cache=_ingested_cache):
            skipped_pdf += 1
            files_done_total += 1
            continue
        pdfs_to_process.append(pdf)

    _metrics["files_remaining"] = total_files - files_done_total
    _progress("pdf", 0, len(pdf_files),
              f"Found {len(pdfs_to_process)} PDFs to process, "
              f"{skipped_pdf} skipped (unchanged/resumed)",
              {"files_remaining": total_files - files_done_total})

    if skipped_pdf:
        print(f"  Skipping {skipped_pdf} unchanged/resumed PDF file(s)")

    # Drop FTS5 triggers before pre-clean to avoid per-row FTS5 updates on DELETE
    print("  Dropping FTS5 triggers for bulk insert optimization...")
    try:
        conn.execute("DROP TRIGGER IF EXISTS pdf_pages_ai")
        conn.execute("DROP TRIGGER IF EXISTS pdf_pages_ad")
        conn.commit()
    except Exception as e:
        print(f"    (Warning: {e})")

    # Pre-clean: remove old data for files being re-processed (triggers already dropped)
    # Use a single batched DELETE via a temp table to avoid 4000+ individual statements
    if pdfs_to_process:
        print(f"  Pre-cleaning {len(pdfs_to_process)} PDF(s) from DB...")
        rel_paths = [str(pdf.relative_to(docs_dir)) for pdf in pdfs_to_process]
        conn.execute("CREATE TEMP TABLE IF NOT EXISTS _pdf_preclean (path TEXT PRIMARY KEY)")
        conn.executemany("INSERT OR IGNORE INTO _pdf_preclean VALUES (?)",
                         [(p,) for p in rel_paths])
        conn.execute(
            "DELETE FROM pdf_pages WHERE source_file IN (SELECT path FROM _pdf_preclean)"
        )
        # LION-103: Also clean junction table to prevent stale PE-to-page links
        try:
            conn.execute(
                "DELETE FROM pdf_pe_numbers WHERE source_file IN "
                "(SELECT path FROM _pdf_preclean)"
            )
        except Exception:
            pass  # Table may not exist in older schemas
        conn.execute("DROP TABLE _pdf_preclean")
        conn.commit()

    try:
      if num_workers > 1 and len(pdfs_to_process) > 1:
        # ── Parallel PDF extraction ───────────────────────────────────
        print(f"  Processing {len(pdfs_to_process)} PDFs with {num_workers} "
              f"parallel workers...")
        pdf_phase_start = time.time()

        # Sliding window: keep at most (num_workers * 2) futures in-flight at once.
        # Submitting all tasks upfront would hold every completed result in memory
        # simultaneously — potentially GBs of page text across thousands of PDFs.
        window_size = num_workers * 2
        pdf_iter = iter(pdfs_to_process)
        active: dict = {}  # future -> pdf

        def _submit_next():
            """Submit the next PDF from the iterator into the pool."""
            pdf = next(pdf_iter, None)
            if pdf is None:
                return
            f = pool.submit(_extract_pdf_data, (str(pdf), str(docs_dir), pdf_timeout))
            active[f] = pdf

        with ProcessPoolExecutor(max_workers=num_workers) as pool:
            # Seed the window
            for _ in range(window_size):
                _submit_next()

            while active:
                if stop_event and stop_event.is_set():
                    print("\n  Graceful stop requested — cancelling workers...")
                    for f in active:
                        f.cancel()
                    _save_checkpoint(conn, session_id, files_done_total,
                                     total_files, total_pdf_pages,
                                     total_budget_rows, 0, "", "interrupted")
                    conn.commit()
                    _pdf_stopped = True
                    _progress("stopped", processed_pdf, len(pdfs_to_process),
                              "Stopped — resume with --resume",
                              {"files_remaining": total_files - files_done_total})
                    break

                # Wait for the next completed future
                done, _ = wait(active, return_when=FIRST_COMPLETED)
                for future in done:
                    pdf = active.pop(future)
                    rel_path = str(pdf.relative_to(docs_dir))
                    processed_pdf += 1
                    files_done_total += 1

                    # Immediately submit a replacement to keep window full
                    _submit_next()

                    try:
                        result = future.result(timeout=300)  # 5-min safety timeout
                    except Exception as e:
                        print(f"  ERROR: {pdf.name}: {e}")
                        stat = pdf.stat()
                        conn.execute(
                            "INSERT OR REPLACE INTO ingested_files "
                            "(file_path, file_type, file_size, file_modified,"
                            " ingested_at, row_count, status) "
                            "VALUES (?,?,?,?,datetime('now'),?,?)",
                            (rel_path, "pdf", stat.st_size, stat.st_mtime,
                             0, f"error: {e}"))
                        _failures.append(FailedFileEntry(
                            file_path=rel_path,
                            error_type=type(e).__name__,
                            error_detail=str(e),
                        ))
                        _mark_file_processed(conn, session_id, rel_path, "pdf",
                                             pages_count=0)
                        continue

                    pages_data = result["pages_data"]
                    issues = result["issues"]
                    error = result["error"]

                    if error:
                        print(f"  ERROR: {pdf.name}: {error}")
                        stat = pdf.stat()
                        conn.execute(
                            "INSERT OR REPLACE INTO ingested_files "
                            "(file_path, file_type, file_size, file_modified,"
                            " ingested_at, row_count, status) "
                            "VALUES (?,?,?,?,datetime('now'),?,?)",
                            (rel_path, "pdf", stat.st_size, stat.st_mtime,
                             0, f"error: {error}"))
                        _failures.append(FailedFileEntry(
                            file_path=rel_path,
                            error_type="ExtractionError",
                            error_detail=str(error),
                        ))
                        _mark_file_processed(conn, session_id, rel_path, "pdf",
                                             pages_count=0)
                        continue

                    # Batch insert all pages for this file
                    pages = len(pages_data)
                    if pages_data:
                        conn.executemany("""
                            INSERT INTO pdf_pages (source_file, source_category,
                                fiscal_year, exhibit_type,
                                page_number, page_text, has_tables, table_data)
                            VALUES (?,?,?,?,?,?,?,?)
                        """, pages_data)

                    # Record extraction issues
                    if issues:
                        conn.executemany(
                            "INSERT INTO extraction_issues"
                            " (file_path, page_number, issue_type, issue_detail)"
                            " VALUES (?,?,?,?)", issues)

                    # LION-103: Insert PE-to-PDF junction rows from parallel worker
                    pe_mentions = result.get("pe_mentions", [])
                    pdf_fy = result.get("fiscal_year")
                    if pe_mentions:
                        # We need pdf_page_id but pages were just inserted;
                        # look up by source_file + page_number
                        page_id_map = {}
                        for row in conn.execute(
                            "SELECT id, page_number FROM pdf_pages "
                            "WHERE source_file = ?", (rel_path,)
                        ):
                            page_id_map[row[1]] = row[0]
                        junction_rows = []
                        for pe, page_num in pe_mentions:
                            pid = page_id_map.get(page_num)
                            if pid is not None:
                                junction_rows.append((
                                    pid, pe, page_num, rel_path, pdf_fy))
                        if junction_rows:
                            conn.executemany(
                                "INSERT INTO pdf_pe_numbers "
                                "(pdf_page_id, pe_number, page_number, "
                                "source_file, fiscal_year) "
                                "VALUES (?,?,?,?,?)", junction_rows)

                    file_status = "ok_with_issues" if issues else "ok"
                    stat = pdf.stat()
                    conn.execute(
                        "INSERT OR REPLACE INTO ingested_files "
                        "(file_path, file_type, file_size, file_modified,"
                        " ingested_at, row_count, status) "
                        "VALUES (?,?,?,?,datetime('now'),?,?)",
                        (rel_path, "pdf", stat.st_size, stat.st_mtime,
                         pages, file_status))

                    total_pdf_pages += pages
                    _metrics["pages"] = total_pdf_pages
                    _metrics["files_remaining"] = total_files - files_done_total

                    # Overall throughput: total pages / wall-clock elapsed
                    elapsed = time.time() - pdf_phase_start
                    if elapsed > 0 and total_pdf_pages > 0:
                        _metrics["speed_pages"] = total_pdf_pages / elapsed

                    # ETA: remaining files * avg time per file so far
                    if processed_pdf > 0:
                        avg_per_file = elapsed / processed_pdf
                        remaining = len(pdfs_to_process) - processed_pdf
                        _metrics["eta_sec"] = avg_per_file * remaining

                    print(f"  [{processed_pdf}/{len(pdfs_to_process)}] "
                          f"{pdf.name}: {pages} pages")
                    _progress("pdf", processed_pdf, len(pdfs_to_process),
                              f"[{processed_pdf}/{len(pdfs_to_process)}] "
                              f"{pdf.name}: {pages} pages",
                              dict(_metrics))

                    _mark_file_processed(conn, session_id, rel_path, "pdf",
                                         pages_count=pages)

                    if files_done_total % checkpoint_interval == 0:
                        _save_checkpoint(conn, session_id, files_done_total,
                                         total_files, total_pdf_pages,
                                         total_budget_rows, stat.st_size,
                                         rel_path, file_status,
                                         commit=False)  # time-based commit below

                    if time.time() - last_commit_time > 2.0:
                        conn.commit()
                        last_commit_time = time.time()

      else:
        # ── Sequential PDF extraction (workers=1) ─────────────────────
        if pdfs_to_process:
            print(f"  Processing {len(pdfs_to_process)} PDFs sequentially...")
        for idx, pdf in enumerate(pdfs_to_process):
            if stop_event and stop_event.is_set():
                print("\n  Graceful stop requested — saving checkpoint...")
                _save_checkpoint(conn, session_id, files_done_total, total_files,
                                 total_pdf_pages, total_budget_rows,
                                 0, str(pdf), "interrupted")
                conn.commit()
                _pdf_stopped = True
                _progress("stopped", idx, len(pdfs_to_process),
                          f"Stopped at {pdf.name} — resume with --resume",
                          {"files_remaining": total_files - files_done_total})
                break

            rel_path = str(pdf.relative_to(docs_dir))
            processed_pdf += 1
            files_done_total += 1
            _metrics["files_remaining"] = total_files - files_done_total

            _progress("pdf", processed_pdf, len(pdfs_to_process),
                      f"[{processed_pdf}] {pdf.name}",
                      {"files_remaining": total_files - files_done_total,
                       "current_pages": 0,
                       "current_total_pages": 0})
            print(f"  [{processed_pdf}/{len(pdfs_to_process)}] {pdf.name}...",
                  end=" ", flush=True)

            def _page_cb(pages_done: int, page_total: int,
                         _proc=processed_pdf, _name=pdf.name,
                         _idx=processed_pdf, _total=len(pdfs_to_process)) -> None:
                """Per-page progress callback forwarded from ingest_pdf_file."""
                _metrics["current_pages"] = pages_done
                _metrics["current_total_pages"] = page_total
                _progress("pdf", _idx, _total,
                          f"[{_proc}] {_name} — page {pages_done}/{page_total}",
                          {"files_remaining": total_files - files_done_total,
                           "current_pages": pages_done,
                           "current_total_pages": page_total})

            t0 = time.time()
            pages, issue_count = ingest_pdf_file(conn, pdf, page_callback=_page_cb,
                                                 docs_dir=docs_dir,
                                                 pdf_timeout=pdf_timeout)
            file_elapsed = time.time() - t0
            print(f"{pages} pages ({file_elapsed:.1f}s)")

            if file_elapsed > 0 and pages > 0:
                _update_speed("speed_pages", pages / file_elapsed)
            pdf_file_times.append(file_elapsed)

            pdfs_remaining = len(pdfs_to_process) - (idx + 1)
            avg_pdf_time = (sum(pdf_file_times) / len(pdf_file_times)
                            if pdf_file_times else 0)
            _metrics["eta_sec"] = avg_pdf_time * pdfs_remaining
            _metrics["pages"] = total_pdf_pages + pages

            file_status = "ok_with_issues" if issue_count > 0 else "ok"

            stat = pdf.stat()
            conn.execute(
                "INSERT OR REPLACE INTO ingested_files "
                "(file_path, file_type, file_size, file_modified,"
                " ingested_at, row_count, status) "
                "VALUES (?,?,?,?,datetime('now'),?,?)",
                (rel_path, "pdf", stat.st_size, stat.st_mtime, pages,
                 file_status))
            total_pdf_pages += pages
            _metrics["pages"] = total_pdf_pages

            _mark_file_processed(conn, session_id, rel_path, "pdf",
                                 pages_count=pages)

            if files_done_total % checkpoint_interval == 0:
                _save_checkpoint(conn, session_id, files_done_total, total_files,
                                 total_pdf_pages, total_budget_rows,
                                 stat.st_size, rel_path, file_status,
                                 commit=False)  # time-based commit below

            if time.time() - last_commit_time > 2.0:
                conn.commit()
                last_commit_time = time.time()

    finally:
        if not _pdf_stopped:
            # Rebuild FTS5 if new pages were added
            final_page_count = conn.execute("SELECT COUNT(*) FROM pdf_pages").fetchone()[0]
            if final_page_count > initial_page_count:
                print("\n  Rebuilding full-text search indexes...")
                # FTS5 'rebuild' command repopulates the index from the content
                # table in a single optimized pass — faster than DELETE+INSERT.
                conn.execute("INSERT INTO pdf_pages_fts(pdf_pages_fts) VALUES('rebuild')")
                conn.execute("""
                    CREATE TRIGGER IF NOT EXISTS pdf_pages_ai AFTER INSERT ON pdf_pages BEGIN
                        INSERT INTO pdf_pages_fts(rowid, page_text, source_file, table_data)
                        VALUES (new.id, new.page_text, new.source_file, new.table_data);
                    END
                """)
                conn.execute("""
                    CREATE TRIGGER IF NOT EXISTS pdf_pages_ad
                    AFTER DELETE ON pdf_pages BEGIN
                        INSERT INTO pdf_pages_fts(
                            pdf_pages_fts, rowid,
                            page_text, source_file, table_data)
                        VALUES (
                            'delete', old.id, old.page_text,
                            old.source_file, old.table_data);
                    END
                """)
                conn.commit()
                print("  FTS5 rebuild complete and triggers recreated")
            else:
                _recreate_pdf_fts_triggers(conn)
                conn.commit()
                print("  Skipped FTS5 rebuild (no new pages added)")

    # If stopped gracefully during PDF loop, close and exit cleanly
    if _pdf_stopped:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        conn.close()
        return

    if skipped_pdf:
        print(f"\n  Skipped {skipped_pdf} unchanged PDF file(s)")
    print(f"  Ingested PDF pages: {total_pdf_pages:,}")
    if num_workers > 1:
        print(f"  Workers used: {num_workers}")

    # ── Detect removed files ───────────────────────────────────────────────
    all_current = {str(f.relative_to(docs_dir)) for f in xlsx_files}
    all_current |= {str(f.relative_to(docs_dir)) for f in pdf_files}

    orphaned = conn.execute(
        "SELECT file_path, file_type FROM ingested_files"
    ).fetchall()
    removed_count = 0
    for row in orphaned:
        if row[0] not in all_current:
            _remove_file_data(conn, row[0], row[1])
            conn.execute("DELETE FROM ingested_files WHERE file_path = ?", (row[0],))
            print(f"  Removed stale data for: {row[0]}")
            removed_count += 1

    if removed_count:
        conn.commit()
        print(f"  Cleaned up {removed_count} removed file(s)")

    # ── Create indexes ─────────────────────────────────────────────────────
    _progress("index", 0, 1, "Creating indexes...")
    print("\nCreating indexes...")
    conn.executescript("""
        CREATE INDEX IF NOT EXISTS idx_bl_exhibit ON budget_lines(exhibit_type);
        CREATE INDEX IF NOT EXISTS idx_bl_org ON budget_lines(organization_name);
        CREATE INDEX IF NOT EXISTS idx_bl_account ON budget_lines(account);
        CREATE INDEX IF NOT EXISTS idx_bl_fy ON budget_lines(fiscal_year);
        CREATE INDEX IF NOT EXISTS idx_bl_sheet ON budget_lines(sheet_name);
        CREATE INDEX IF NOT EXISTS idx_bl_source ON budget_lines(source_file);
        CREATE INDEX IF NOT EXISTS idx_bl_pe ON budget_lines(pe_number);
        CREATE INDEX IF NOT EXISTS idx_bl_approp ON budget_lines(appropriation_code);
        -- Composite index for PE detail: pe_number + fiscal_year covers
        -- the common WHERE pe_number=? ORDER BY fiscal_year pattern.
        CREATE INDEX IF NOT EXISTS idx_bl_pe_fy
            ON budget_lines(pe_number, fiscal_year);
        -- Composite index for dashboard: organization + amount for top-N queries.
        CREATE INDEX IF NOT EXISTS idx_bl_org_amount
            ON budget_lines(organization_name, amount_fy2026_request);
        CREATE INDEX IF NOT EXISTS idx_pp_source ON pdf_pages(source_file);
        CREATE INDEX IF NOT EXISTS idx_pp_category ON pdf_pages(source_category);
        -- Composite indexes used by enrich_budget_db.py Phase 3 batch queries
        CREATE INDEX IF NOT EXISTS idx_bl_pe_fields
            ON budget_lines(pe_number, budget_activity_title,
                            appropriation_title, organization_name);
        CREATE INDEX IF NOT EXISTS idx_pe_desc_pe_text
            ON pe_descriptions(pe_number)
            WHERE description_text IS NOT NULL;

        -- Composite: exhibit_type + fiscal_year covers dashboard CTE and
        -- aggregation GROUP BY queries filtering on both columns.
        CREATE INDEX IF NOT EXISTS idx_bl_exhibit_fy
            ON budget_lines(exhibit_type, fiscal_year);

        -- Composite: appropriation_code + amount covers dashboard
        -- by-appropriation top-N ORDER BY queries.
        CREATE INDEX IF NOT EXISTS idx_bl_approp_amount
            ON budget_lines(appropriation_code, amount_fy2026_request);

        -- Composite: pdf_pages fiscal_year + exhibit_type covers
        -- PDF filtering queries that filter on both columns.
        CREATE INDEX IF NOT EXISTS idx_pp_fy_exhibit
            ON pdf_pages(fiscal_year, exhibit_type);

        -- Composite: hierarchy endpoint GROUP BY (org, approp, line_item).
        CREATE INDEX IF NOT EXISTS idx_bl_org_approp_line
            ON budget_lines(organization_name, appropriation_code,
                            line_item_title);

        -- Composite: pe_descriptions section header filter
        -- covers GET /pe/{pe}/descriptions?section= queries.
        CREATE INDEX IF NOT EXISTS idx_pe_desc_section
            ON pe_descriptions(pe_number, section_header);

        -- Single-column: budget_type used by aggregation GROUP BY
        -- and dashboard budget-type breakdown queries.
        CREATE INDEX IF NOT EXISTS idx_bl_budget_type
            ON budget_lines(budget_type);

        -- Composite: fiscal_year + organization_name covers dashboard
        -- and aggregation queries filtered on both columns.
        CREATE INDEX IF NOT EXISTS idx_bl_fy_org
            ON budget_lines(fiscal_year, organization_name);

        -- Composite: pe_number + amount covers PE list queries that
        -- ORDER BY amount_fy2026_request for a given PE.
        CREATE INDEX IF NOT EXISTS idx_bl_pe_amount
            ON budget_lines(pe_number, amount_fy2026_request);

        -- Composite: budget_type + amount covers dashboard budget-type
        -- breakdown with ORDER BY amount_fy2026_request DESC.
        CREATE INDEX IF NOT EXISTS idx_bl_budget_type_amount
            ON budget_lines(budget_type, amount_fy2026_request);
    """)
    conn.commit()
    _progress("index", 1, 1, "Indexes created")

    # ── Update data source timestamps ──────────────────────────────────────
    conn.execute("""
        UPDATE data_sources SET last_updated = datetime('now')
        WHERE source_id IN (
            SELECT DISTINCT
                substr(file_path, 1,
                    instr(substr(file_path, 1+instr(file_path, '/')), '/')
                    + instr(file_path, '/') - 1)
            FROM ingested_files
        )
    """)
    conn.commit()

    # ── Mark session complete ──────────────────────────────────────────────
    _mark_session_complete(conn, session_id,
                           f"Processed {files_done_total} files: "
                           f"{total_budget_rows:,} rows, {total_pdf_pages:,} pages")

    # ── Summary ────────────────────────────────────────────────────────────
    total_lines = conn.execute("SELECT COUNT(*) FROM budget_lines").fetchone()[0]
    total_pages = conn.execute("SELECT COUNT(*) FROM pdf_pages").fetchone()[0]
    db_size = db_path.stat().st_size / (1024 * 1024)

    summary = (
        f"Database: {db_path} ({db_size:.1f} MB)\n"
        f"Budget lines: {total_lines:,}\n"
        f"PDF pages: {total_pages:,}\n"
        f"Excel files: {len(xlsx_files)}\n"
        f"PDF files: {len(pdf_files)}"
    )

    print(f"\n{'='*60}")
    print("  BUILD COMPLETE")
    print(f"{'='*60}")
    print(f"  Database:           {db_path} ({db_size:.1f} MB)")
    print(f"  Total budget lines: {total_lines:,}")
    print(f"  Total PDF pages:    {total_pages:,}")
    print(f"  Excel files:        {len(xlsx_files)}")
    print(f"  PDF files:          {len(pdf_files)}")
    if not rebuild:
        new_files = (len(xlsx_files) - skipped_xlsx) + (len(pdf_files) - skipped_pdf)
        print(f"  New/updated files:  {new_files}")
        print(f"  Unchanged (skip):   {skipped_xlsx + skipped_pdf}")
        summary += f"\nNew/updated: {new_files} | Skipped: {skipped_xlsx + skipped_pdf}"

    _progress("done", total_files, total_files, summary,
              {"files_remaining": 0, "eta_sec": 0.0})
    # Final WAL checkpoint: flush all accumulated WAL pages to the main DB file
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.close()

    # BUILD-001: Write failures log
    if _failures:
        with open(_failures_log_path, "w") as _flog:
            json.dump([f.to_dict() for f in _failures], _flog, indent=2)
        print(f"\n  BUILD-001: {len(_failures)} file(s) failed — logged to {_failures_log_path}")
        print("  Re-process with: python build_budget_db.py --retry-failures")
    elif _failures_log_path.exists() and not retry_failures:
        # Clear stale failure log after a clean build
        _failures_log_path.unlink()

    # 1.B6-h / 2.B3-a: Post-build validation + data-quality JSON report
    try:
        from validate_budget_data import (  # noqa: PLC0415
            generate_quality_report,
        )
        report = generate_quality_report(db_path, print_console=True)
        val = report["validation_summary"]
        print(
            f"\n  [QUALITY REPORT] {report['total_budget_lines']:,} budget lines | "
            f"{val['total_checks']} checks | "
            f"{val['total_warnings']} warning(s) | "
            f"{val['total_failures']} failure(s)"
        )
        print("  [QUALITY REPORT] Written to data_quality_report.json")
    except Exception as _val_err:
        print(f"\n  [VALIDATION] Skipped: {_val_err}")


def main():
    """Parse command-line arguments and run the database build pipeline."""
    parser = argparse.ArgumentParser(description="Build DoD budget search database")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH,
                        help=f"Database path (default: {DEFAULT_DB_PATH})")
    parser.add_argument("--docs", type=Path, default=DOCS_DIR,
                        help=f"Documents directory (default: {DOCS_DIR})")
    parser.add_argument("--rebuild", action="store_true",
                        help="Force full rebuild (delete existing database)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from last checkpoint")
    parser.add_argument("--checkpoint-interval", type=int, default=10,
                        metavar="N",
                        help="Save a checkpoint every N files (default: 10)")
    parser.add_argument("--workers", type=int, default=0, metavar="N",
                        help="Parallel workers for PDF extraction "
                             "(default: 0 = auto-detect CPU count, 1 = sequential)")
    # BUILD-003: Configurable PDF extraction timeout
    parser.add_argument("--pdf-timeout", type=int, default=30, metavar="SECS",
                        help="Seconds to wait for table extraction per PDF page "
                             "(default: 30; increase for complex PDFs)")
    # BUILD-001: Failure log and retry
    parser.add_argument("--retry-failures", action="store_true",
                        help="Re-process only files listed in failed_downloads.json")
    parser.add_argument("--failures-log", type=Path, default=Path("failed_downloads.json"),
                        metavar="PATH",
                        help="Path to write/read the failure log "
                             "(default: failed_downloads.json)")
    args = parser.parse_args()

    # ── Graceful shutdown via Ctrl+C ───────────────────────────────────────
    import threading
    stop_event = threading.Event()

    def _sigint_handler(sig, frame):
        """Handle SIGINT (Ctrl+C): request graceful stop on first press, force-quit on second."""
        if not stop_event.is_set():
            print("\n\nKeyboard interrupt — finishing current file and saving checkpoint...")
            print("Resume later with: python build_budget_db.py --resume")
            stop_event.set()
        else:
            print("\nForce-quitting...")
            sys.exit(1)

    signal.signal(signal.SIGINT, _sigint_handler)

    try:
        build_database(args.docs, args.db,
                       rebuild=args.rebuild,
                       resume=args.resume,
                       checkpoint_interval=args.checkpoint_interval,
                       stop_event=stop_event,
                       workers=args.workers,
                       pdf_timeout=args.pdf_timeout,
                       failures_log=args.failures_log,
                       retry_failures=args.retry_failures)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
