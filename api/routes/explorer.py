"""
Keyword Explorer endpoints.

A generalized version of the Hypersonics page that accepts user-supplied
keywords.  Users enter keywords, the system builds a cache (with PDF mining),
and they can preview results and download a customized XLSX.

Endpoints:
  POST /api/v1/explorer/build             — start async cache build
  GET  /api/v1/explorer/status            — poll build progress
  GET  /api/v1/explorer                   — JSON PE-level summary + available columns
  POST /api/v1/explorer/download/xlsx     — XLSX with user-selected columns
  GET  /api/v1/explorer/desc/{pe_number}  — lazy-load description text
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import re
import sqlite3
import threading
import time
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Body, Depends, Query
from fastapi.responses import Response

from api.database import get_db
from api.routes.keyword_search import (
    FY_END,
    FY_START,
    build_cache_table,
    cache_rows_to_dicts,
    load_per_fy_descriptions,
    xlsx_base_styles,
)
from utils.fuzzy_match import expand_keywords

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/explorer", tags=["explorer"])

# ── Input validation constants ────────────────────────────────────────────────

MAX_KEYWORDS = 20
MIN_KEYWORD_LEN = 2
MAX_KEYWORD_LEN = 100
MAX_CACHE_TABLES = 50
_KEYWORD_RE = re.compile(r"^[a-zA-Z0-9\s\-/&.]+$")

# ── Build progress tracking ──────────────────────────────────────────────────

# Module-level dict: keyword_set_id → progress state.
# Protected by a lock for thread safety (BackgroundTasks run in threads).
_build_lock = threading.Lock()
_build_progress: dict[str, dict[str, Any]] = {}

# ── Available columns for XLSX export ─────────────────────────────────────────

_ALL_COLUMNS = [
    "PE Number",
    "Service/Org",
    "Exhibit Type",
    "Line Item Title",
    "Budget Activity",
    "Budget Activity (Normalized)",
    "Appropriation",
    "Color of Money",
    "Keywords (Row)",
    "Keywords (Desc)",
    "Description",
]

# FY columns are added dynamically based on active years

_COL_TO_FIELD: dict[str, str] = {
    "PE Number": "pe_number",
    "Service/Org": "organization_name",
    "Exhibit Type": "exhibit_type",
    "Line Item Title": "line_item_title",
    "Budget Activity": "budget_activity_title",
    "Budget Activity (Normalized)": "budget_activity_norm",
    "Appropriation": "appropriation_title",
    "Color of Money": "color_of_money",
    "Keywords (Row)": "matched_keywords_row",
    "Keywords (Desc)": "matched_keywords_desc",
    "Description": "description_text",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _keyword_set_id(keywords: list[str]) -> str:
    """SHA-256 hash of sorted lowercase keywords."""
    normalized = sorted(set(kw.strip().lower() for kw in keywords if kw.strip()))
    return hashlib.sha256(",".join(normalized).encode()).hexdigest()


def _cache_table_name(kw_id: str) -> str:
    """Return cache table name from keyword set ID."""
    return f"explorer_cache_{kw_id[:16]}"


def _parse_keywords(raw: str) -> list[str]:
    """Parse and validate comma-separated keywords string.

    Returns cleaned keyword list. Raises ValueError on validation failure.
    """
    if not raw or not raw.strip():
        raise ValueError("No keywords provided")

    parts = [k.strip() for k in raw.split(",") if k.strip()]
    if not parts:
        raise ValueError("No keywords provided")
    if len(parts) > MAX_KEYWORDS:
        raise ValueError(f"Too many keywords (max {MAX_KEYWORDS})")

    cleaned: list[str] = []
    for kw in parts:
        if len(kw) < MIN_KEYWORD_LEN:
            raise ValueError(f"Keyword '{kw}' is too short (min {MIN_KEYWORD_LEN} characters)")
        if len(kw) > MAX_KEYWORD_LEN:
            raise ValueError(f"Keyword '{kw}' is too long (max {MAX_KEYWORD_LEN} characters)")
        if not _KEYWORD_RE.match(kw):
            raise ValueError(f"Keyword '{kw}' contains invalid characters")
        cleaned.append(kw)
    return cleaned


def _ensure_meta_table(conn: sqlite3.Connection) -> None:
    """Create the explorer_cache_meta table if it doesn't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS explorer_cache_meta (
            keyword_set_id   TEXT PRIMARY KEY,
            keywords_json    TEXT NOT NULL,
            table_name       TEXT NOT NULL,
            built_at         REAL NOT NULL,
            last_accessed_at REAL NOT NULL,
            row_count        INTEGER
        )
    """)


def _prune_old_caches(conn: sqlite3.Connection) -> None:
    """Remove explorer cache tables exceeding MAX_CACHE_TABLES (oldest first)."""
    _ensure_meta_table(conn)
    rows = conn.execute(
        "SELECT keyword_set_id, table_name FROM explorer_cache_meta "
        "ORDER BY last_accessed_at DESC"
    ).fetchall()
    if len(rows) <= MAX_CACHE_TABLES:
        return
    for kw_id, table_name in rows[MAX_CACHE_TABLES:]:
        conn.execute(f"DROP TABLE IF EXISTS {table_name}")
        conn.execute("DELETE FROM explorer_cache_meta WHERE keyword_set_id = ?", [kw_id])
    conn.commit()
    logger.info("Pruned %d old explorer caches", len(rows) - MAX_CACHE_TABLES)


def _do_build(
    db_path: str,
    kw_id: str,
    keywords: list[str],
    expanded: list[str],
) -> None:
    """Background task: build the explorer cache table.

    Runs in a separate thread via BackgroundTasks. Uses its own DB connection
    since SQLite connections aren't thread-safe.
    """
    cache_table = _cache_table_name(kw_id)

    def _progress(step: str, detail: dict[str, Any]) -> None:
        with _build_lock:
            _build_progress[kw_id] = {
                "state": "building",
                "progress": step,
                **detail,
            }

    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")

        _ensure_meta_table(conn)
        _prune_old_caches(conn)

        row_count = build_cache_table(
            conn, cache_table, expanded, expanded,
            fy_start=FY_START, fy_end=FY_END,
            progress_callback=_progress,
        )

        # Record in metadata
        now = time.time()
        conn.execute(
            "INSERT OR REPLACE INTO explorer_cache_meta "
            "(keyword_set_id, keywords_json, table_name, built_at, last_accessed_at, row_count) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [kw_id, json.dumps(keywords), cache_table, now, now, row_count],
        )
        conn.commit()
        conn.close()

        with _build_lock:
            _build_progress[kw_id] = {
                "state": "ready",
                "progress": "done",
                "row_count": row_count,
            }
    except Exception as e:
        logger.exception("Explorer cache build failed for %s", kw_id)
        with _build_lock:
            _build_progress[kw_id] = {
                "state": "error",
                "progress": str(e),
            }


# ── POST /api/v1/explorer/build ──────────────────────────────────────────────

@router.post(
    "/build",
    summary="Start async cache build for a keyword set",
)
def start_build(
    keywords: str = Query(..., description="Comma-separated keywords"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Kick off a background cache build and return immediately."""
    try:
        keyword_list = _parse_keywords(keywords)
    except ValueError as e:
        return {"error": str(e)}

    expanded = expand_keywords(keyword_list)
    kw_id = _keyword_set_id(keyword_list)

    # Check if cache already exists and is fresh
    _ensure_meta_table(conn)
    meta = conn.execute(
        "SELECT built_at, row_count FROM explorer_cache_meta WHERE keyword_set_id = ?",
        [kw_id],
    ).fetchone()

    if meta:
        age_hours = (time.time() - meta[0]) / 3600
        if age_hours < 24:
            # Cache is fresh — update access time and return ready
            conn.execute(
                "UPDATE explorer_cache_meta SET last_accessed_at = ? WHERE keyword_set_id = ?",
                [time.time(), kw_id],
            )
            conn.commit()
            with _build_lock:
                _build_progress[kw_id] = {
                    "state": "ready",
                    "progress": "done",
                    "row_count": meta[1],
                }
            return {
                "keyword_set_id": kw_id,
                "keywords": keyword_list,
                "expanded_keywords": expanded,
                "state": "ready",
            }

    # Check if build is already in progress
    with _build_lock:
        current = _build_progress.get(kw_id, {})
        if current.get("state") == "building":
            return {
                "keyword_set_id": kw_id,
                "keywords": keyword_list,
                "expanded_keywords": expanded,
                "state": "building",
            }

        _build_progress[kw_id] = {
            "state": "building",
            "progress": "starting",
        }

    # Get DB path for the background thread
    db_path = conn.execute("PRAGMA database_list").fetchone()[2]

    background_tasks.add_task(_do_build, db_path, kw_id, keyword_list, expanded)

    return {
        "keyword_set_id": kw_id,
        "keywords": keyword_list,
        "expanded_keywords": expanded,
        "state": "building",
    }


# ── GET /api/v1/explorer/status ──────────────────────────────────────────────

@router.get(
    "/status",
    summary="Poll cache build progress",
)
def build_status(
    keywords: str = Query(..., description="Comma-separated keywords"),
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Return current build state for a keyword set.

    Checks in-memory progress first; falls back to the database so that
    completed builds survive process restarts / uvicorn reloads.
    """
    try:
        keyword_list = _parse_keywords(keywords)
    except ValueError as e:
        return {"state": "error", "progress": str(e)}

    kw_id = _keyword_set_id(keyword_list)

    with _build_lock:
        status = _build_progress.get(kw_id)

    # If in-memory says "ready" or "error", trust it
    if status and status.get("state") in ("ready", "error"):
        return {"keyword_set_id": kw_id, **status}

    # Check DB — the build may have finished (in this or a previous process)
    _ensure_meta_table(conn)
    meta = conn.execute(
        "SELECT row_count, built_at FROM explorer_cache_meta WHERE keyword_set_id = ?",
        [kw_id],
    ).fetchone()
    if meta is not None:
        # Refresh in-memory state so subsequent polls are fast
        ready = {"state": "ready", "progress": "done", "row_count": meta[0]}
        with _build_lock:
            _build_progress[kw_id] = ready
        return {"keyword_set_id": kw_id, **ready}

    # In-memory says "building" but DB has no result yet — still in progress
    if status and status.get("state") == "building":
        return {"keyword_set_id": kw_id, **status}

    return {"state": "not_started", "keyword_set_id": kw_id}


# ── GET /api/v1/explorer ─────────────────────────────────────────────────────

@router.get(
    "",
    summary="PE-level summary and available columns for a keyword set",
)
def get_explorer_data(
    keywords: str = Query(..., description="Comma-separated keywords"),
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Return PE-level summary of cached results plus available download columns."""
    try:
        keyword_list = _parse_keywords(keywords)
    except ValueError as e:
        return {"error": str(e)}

    kw_id = _keyword_set_id(keyword_list)
    cache_table = _cache_table_name(kw_id)
    expanded = expand_keywords(keyword_list)

    # Try to read from cache
    try:
        sql = f"SELECT * FROM {cache_table} ORDER BY pe_number, exhibit_type, line_item_title"
        raw_rows = conn.execute(sql).fetchall()
    except sqlite3.OperationalError:
        return {"error": "Cache not built yet. Call POST /build first."}

    items = cache_rows_to_dicts(raw_rows)

    # Build PE-level summary
    pe_groups: dict[str, list[dict]] = {}
    for row in items:
        pe = row["pe_number"]
        pe_groups.setdefault(pe, []).append(row)

    year_range = list(range(FY_START, FY_END + 1))
    active_years = [
        yr for yr in year_range
        if any(r.get(f"fy{yr}") is not None for r in items)
    ] if items else []

    pe_summary: list[dict] = []
    for pe, children in pe_groups.items():
        # Find PE title from R-1 row or first child
        r1_titles = [c["line_item_title"] for c in children if c.get("exhibit_type") == "r1"]
        title = r1_titles[-1] if r1_titles else (children[0].get("line_item_title") or pe)
        service = next((c.get("organization_name") for c in children if c.get("organization_name")), "")

        # Count matching sub-elements (those with any keyword match)
        matching = sum(
            1 for c in children
            if c.get("matched_keywords_row") or c.get("matched_keywords_desc")
        )
        pe_summary.append({
            "pe_number": pe,
            "pe_title": title,
            "service": service,
            "total_sub_elements": len(children),
            "matching_sub_elements": matching,
        })

    # Build available columns list (static + dynamic FY columns).
    # Per-year columns are offered as an interleaved [value, source, description]
    # triple so selecting them in order keeps the output columns grouped by year.
    available_columns = list(_ALL_COLUMNS)
    for yr in active_years:
        available_columns.append(f"FY{yr} ($K)")
        available_columns.append(f"FY{yr} Source")
        available_columns.append(f"FY{yr} Description")

    # Default selected columns — fixed metadata, then an interleaved FY triple
    # (value / source / description) per active year so the sheet reads left to
    # right in chronological order with descriptions adjacent to amounts.
    default_columns = [
        "PE Number", "Service/Org", "Exhibit Type", "Line Item Title",
        "Color of Money",
    ]
    for yr in active_years:
        default_columns.append(f"FY{yr} ($K)")
        default_columns.append(f"FY{yr} Source")
        default_columns.append(f"FY{yr} Description")

    return {
        "keyword_set_id": kw_id,
        "keywords": keyword_list,
        "expanded_keywords": expanded,
        "pe_summary": pe_summary,
        "total_pes": len(pe_summary),
        "total_rows": len(items),
        "active_years": active_years,
        "available_columns": available_columns,
        "default_columns": default_columns,
    }


# ── POST /api/v1/explorer/download/xlsx ──────────────────────────────────────

@router.post(
    "/download/xlsx",
    summary="Download explorer results as XLSX with selected columns",
    response_class=Response,
)
def download_explorer_xlsx(
    keywords: str = Body(..., description="Comma-separated keywords"),
    columns: list[str] = Body(..., description="Ordered list of column names to include"),
    matching_only: bool = Body(False, description="Only include directly matching sub-elements"),
    conn: sqlite3.Connection = Depends(get_db),
) -> Response:
    """Generate XLSX with user-selected columns in chosen order."""
    import openpyxl
    from openpyxl.styles import Alignment

    try:
        keyword_list = _parse_keywords(keywords)
    except ValueError as e:
        return Response(content=str(e).encode(), media_type="text/plain", status_code=400)

    kw_id = _keyword_set_id(keyword_list)
    cache_table = _cache_table_name(kw_id)

    try:
        sql = f"SELECT * FROM {cache_table} ORDER BY pe_number, exhibit_type, line_item_title"
        raw_rows = conn.execute(sql).fetchall()
    except sqlite3.OperationalError:
        return Response(content=b"Cache not built", media_type="text/plain", status_code=400)

    items = cache_rows_to_dicts(raw_rows)

    # Filter to matching-only if requested
    if matching_only:
        items = [
            r for r in items
            if r.get("matched_keywords_row") or r.get("matched_keywords_desc")
        ]

    if not items:
        return Response(content=b"No rows to export", media_type="text/plain", status_code=400)

    year_range = list(range(FY_START, FY_END + 1))

    # Per-(pe, fy) descriptions so each year's column can display narrative text
    # from that year's submission (often a different source than the amount).
    desc_by_pe_fy = load_per_fy_descriptions(
        conn, {r.get("pe_number", "") for r in items}
    )

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Explorer"

    sty = xlsx_base_styles()
    header_fill = sty["header_fill"]
    header_font_white = sty["header_font"]
    normal_font = sty["base_font"]
    italic_font = sty["italic_font"]
    total_font = sty["total_font"]
    money_fmt = sty["money_fmt"]

    # Ensure an "In Totals" column is present so users can see which rows feed the
    # per-FY totals and so SUMIF formulas have something to key on. We append it
    # (without dropping anything the user explicitly asked for) if absent.
    export_columns: list[str] = list(columns)
    if "In Totals" not in export_columns:
        export_columns.append("In Totals")
    in_totals_col_idx = export_columns.index("In Totals") + 1  # 1-indexed

    # Headers
    for col_idx, col_name in enumerate(export_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    first_data_row = 2
    fy_value_columns: dict[str, int] = {}  # column name → 1-indexed column number
    for row_num, row in enumerate(items, first_data_row):
        has_match = bool(
            row.get("matched_keywords_row") or row.get("matched_keywords_desc")
        )
        font = normal_font if has_match else italic_font
        for col_idx, col_name in enumerate(export_columns, 1):
            value = _extract_column_value(
                row, col_name, year_range, desc_by_pe_fy, has_match
            )
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = font
            if col_name.endswith("($K)") and isinstance(value, (int, float)):
                cell.number_format = money_fmt
                fy_value_columns.setdefault(col_name, col_idx)

    last_data_row = first_data_row + len(items) - 1

    # Totals row — live SUMIF formulas keyed off the "In Totals" column so users
    # can edit flags in Excel and have the totals recalculate.
    if fy_value_columns and last_data_row >= first_data_row:
        totals_row = last_data_row + 1
        ws.cell(row=totals_row, column=1, value="TOTALS").font = total_font
        if in_totals_col_idx != 1:
            ws.cell(row=totals_row, column=in_totals_col_idx, value="Sum").font = total_font
        in_totals_letter = openpyxl.utils.get_column_letter(in_totals_col_idx)
        in_totals_range = f"${in_totals_letter}${first_data_row}:${in_totals_letter}${last_data_row}"
        for col_name, col_idx in fy_value_columns.items():
            val_letter = openpyxl.utils.get_column_letter(col_idx)
            val_range = f"${val_letter}${first_data_row}:${val_letter}${last_data_row}"
            formula = f'=SUMIF({in_totals_range},"Yes",{val_range})'
            cell = ws.cell(row=totals_row, column=col_idx, value=formula)
            cell.font = total_font
            cell.number_format = money_fmt

    # Column widths
    _WIDTH_MAP = {
        "Description": 60, "Line Item Title": 50,
        "Budget Activity": 20, "Budget Activity (Normalized)": 20,
        "PE Number": 14, "Service/Org": 14, "Exhibit Type": 8,
        "Color of Money": 12, "Appropriation": 30, "In Totals": 10,
    }
    for col_idx, col_name in enumerate(export_columns, 1):
        if col_name in _WIDTH_MAP:
            width = _WIDTH_MAP[col_name]
        elif col_name.endswith("($K)"):
            width = 14
        elif col_name.endswith("Source"):
            width = 30
        elif col_name.endswith("Description"):
            width = 40
        else:
            width = max(14, len(col_name) + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row + first 4 columns (PE, Service, Exhibit, Title)
    freeze_col = min(5, len(columns) + 1)
    ws.freeze_panes = f"{openpyxl.utils.get_column_letter(freeze_col)}2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="explorer_results.xlsx"'},
    )


def _extract_column_value(
    row: dict,
    col_name: str,
    year_range: list[int],
    desc_by_pe_fy: dict[tuple[str, str], str] | None = None,
    in_totals: bool | None = None,
) -> Any:
    """Extract a cell value from a cache row dict for a given column name."""
    if col_name == "In Totals":
        if in_totals is None:
            in_totals = bool(
                row.get("matched_keywords_row") or row.get("matched_keywords_desc")
            )
        return "Yes" if in_totals else ""

    # Static columns
    field = _COL_TO_FIELD.get(col_name)
    if field:
        val = row.get(field, "")
        # Format list fields
        if field in ("matched_keywords_row", "matched_keywords_desc"):
            if isinstance(val, list):
                return ", ".join(val)
        return val if val is not None else ""

    # FY amount columns: "FY2024 ($K)"
    m = re.match(r"FY(\d{4})\s*\(\$K\)", col_name)
    if m:
        yr = int(m.group(1))
        return row.get(f"fy{yr}")

    # FY source columns: "FY2024 Source"
    m = re.match(r"FY(\d{4})\s*Source", col_name)
    if m:
        yr = int(m.group(1))
        return row.get("refs", {}).get(f"fy{yr}", "")

    # FY description columns: "FY2024 Description"
    m = re.match(r"FY(\d{4})\s*Description", col_name)
    if m:
        yr = int(m.group(1))
        if desc_by_pe_fy is None:
            return ""
        pe = row.get("pe_number", "")
        return desc_by_pe_fy.get((pe, str(yr)), "")

    return ""


# ── GET /api/v1/explorer/desc/{pe_number} ────────────────────────────────────

@router.get(
    "/desc/{pe_number}",
    summary="Get description text for a PE from explorer cache",
)
def get_description(
    pe_number: str,
    keywords: str = Query(..., description="Comma-separated keywords"),
    project: str | None = None,
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Return description_text for a PE from the explorer cache."""
    try:
        keyword_list = _parse_keywords(keywords)
    except ValueError:
        return {"description": None}

    kw_id = _keyword_set_id(keyword_list)
    cache_table = _cache_table_name(kw_id)

    try:
        if project:
            row = conn.execute(
                f"SELECT description_text FROM {cache_table} "
                "WHERE pe_number = ? AND line_item_title = ? AND description_text IS NOT NULL LIMIT 1",
                [pe_number, project],
            ).fetchone()
        else:
            row = conn.execute(
                f"SELECT description_text FROM {cache_table} "
                "WHERE pe_number = ? AND description_text IS NOT NULL LIMIT 1",
                [pe_number],
            ).fetchone()
        return {"description": row[0] if row else None}
    except sqlite3.OperationalError:
        return {"description": None}
