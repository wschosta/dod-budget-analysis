"""
GET /api/v1/download endpoint (Step 2.C4-a).

Streams large result sets as CSV, JSON, or Excel without loading everything
into memory. Accepts the same filter parameters as /budget-lines.

DL-001: Excel (.xlsx) export via openpyxl write_only mode.
DL-002: Keyword search filter (q) for FTS-filtered downloads.
DL-003: X-Total-Count header for client progress tracking.
OPT-DL-001: Uses shared WHERE builder from utils/query.py.
"""

import csv
import io
import json
import sqlite3
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse

from api.database import get_db
from utils import sanitize_fts5_query
from utils.query import build_where_clause

router = APIRouter(prefix="/download", tags=["download"])

_DOWNLOAD_COLUMNS = [
    "id", "source_file", "exhibit_type", "sheet_name", "fiscal_year",
    "account", "account_title", "organization_name",
    "budget_activity_title", "sub_activity_title",
    "line_item", "line_item_title", "pe_number",
    "appropriation_code", "appropriation_title", "budget_type",
    "amount_fy2024_actual", "amount_fy2025_enacted", "amount_fy2025_supplemental",
    "amount_fy2025_total", "amount_fy2026_request", "amount_fy2026_reconciliation",
    "amount_fy2026_total", "amount_type", "amount_unit", "currency_year",
]

_ALLOWED_SORT = {"id", "source_file", "exhibit_type", "fiscal_year",
                 "organization_name", "amount_fy2026_request", "budget_type",
                 "pe_number", "appropriation_code"}


def _iter_rows(conn: sqlite3.Connection, sql: str, params: list[Any]):
    """Yield rows one at a time to enable streaming."""
    cur = conn.execute(sql, params)
    while True:
        batch = cur.fetchmany(500)
        if not batch:
            break
        yield from batch


def _build_download_sql(
    fiscal_year: list[str] | None,
    service: list[str] | None,
    exhibit_type: list[str] | None,
    pe_number: list[str] | None,
    appropriation_code: list[str] | None,
    q: str | None,
    conn: sqlite3.Connection,
    limit: int,
    export_cols: list[str],
    sort_by: str = "id",
    sort_dir: str = "asc",
    min_amount: float | None = None,
    max_amount: float | None = None,
    budget_type: list[str] | None = None,
) -> tuple[str, list[Any], int]:
    """Build the download SQL with all filters applied.

    Returns:
        (sql, params, total_count)
    """
    # DL-002: Handle FTS keyword filter
    fts_ids: list[int] | None = None
    if q and q.strip():
        try:
            safe_q = sanitize_fts5_query(q.strip())
            fts_rows = conn.execute(
                "SELECT rowid FROM budget_lines_fts WHERE budget_lines_fts MATCH ?",
                (safe_q,),
            ).fetchall()
            fts_ids = [r[0] for r in fts_rows]
        except Exception:
            fts_ids = []

    # OPT-DL-001: Use shared WHERE builder
    where, params = build_where_clause(
        fiscal_year=fiscal_year,
        service=service,
        exhibit_type=exhibit_type,
        pe_number=pe_number,
        appropriation_code=appropriation_code,
        budget_type=budget_type,
        min_amount=min_amount,
        max_amount=max_amount,
        fts_ids=fts_ids,
    )

    # DL-003: Count first for X-Total-Count header
    count_sql = f"SELECT COUNT(*) FROM budget_lines {where}"
    total = conn.execute(count_sql, params).fetchone()[0]

    col_list = ", ".join(export_cols)
    sort_col = sort_by if sort_by in _ALLOWED_SORT else "id"
    direction = "DESC" if sort_dir.lower() == "desc" else "ASC"
    sql = (f"SELECT {col_list} FROM budget_lines {where} "
           f"ORDER BY {sort_col} {direction} LIMIT {limit}")

    return sql, params, total



@router.get("", summary="Download search results as CSV, JSON, or Excel")
def download(
    request: Request,
    fmt: str = Query("csv", pattern="^(csv|json|xlsx)$", description="Output format"),
    fiscal_year: list[str] | None = Query(None),
    service: list[str] | None = Query(None),
    exhibit_type: list[str] | None = Query(None),
    pe_number: list[str] | None = Query(None),
    appropriation_code: list[str] | None = Query(None, description="Filter by appropriation code(s)"),
    budget_type: list[str] | None = Query(None, description="Filter by budget type (RDT&E, Procurement, etc.)"),
    # DL-002: keyword search filter
    q: str | None = Query(None, description="Keyword search filter (FTS5)"),
    min_amount: float | None = Query(None, description="Min FY2026 request amount"),
    max_amount: float | None = Query(None, description="Max FY2026 request amount"),
    sort_by: str = Query("id", description="Column to sort by"),
    sort_dir: str = Query("asc", pattern="^(asc|desc)$", description="Sort direction"),
    # FIX-017: Download a single item by ID (used by detail panel)
    item_id: int | None = Query(None, description="Download a specific budget line by ID"),
    limit: int = Query(
        10_000, ge=1, le=100_000,
        description="Max rows to export (default 10,000)",
    ),
    columns: list[str] | None = Query(None, description="FE-011: Subset of columns to export"),
    conn: sqlite3.Connection = Depends(get_db),
) -> StreamingResponse:
    """Stream budget line items as CSV, JSON (newline-delimited), or Excel."""
    # FE-011: filter columns to those requested (and valid)
    export_cols = (
        [c for c in columns if c in _DOWNLOAD_COLUMNS]
        if columns
        else _DOWNLOAD_COLUMNS
    ) or _DOWNLOAD_COLUMNS

    # FIX-017: If item_id is specified, download just that single row
    if item_id is not None:
        col_list = ", ".join(export_cols)
        sql = f"SELECT {col_list} FROM budget_lines WHERE id = ?"
        params = [item_id]
        total_count = 1
    else:
        sql, params, total_count = _build_download_sql(
            fiscal_year=fiscal_year,
            service=service,
            exhibit_type=exhibit_type,
            pe_number=pe_number,
            appropriation_code=appropriation_code,
            q=q,
            conn=conn,
            limit=limit,
            export_cols=export_cols,
            sort_by=sort_by,
            sort_dir=sort_dir,
            min_amount=min_amount,
            max_amount=max_amount,
            budget_type=budget_type,
        )

    # DL-003: X-Total-Count header
    extra_headers: dict[str, str] = {"X-Total-Count": str(total_count)}

    # EAGLE-6: Source attribution metadata
    export_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    export_url = str(request.url)
    # Build a human-readable filter summary
    active_filters: list[str] = []
    if fiscal_year:
        active_filters.append(f"fiscal_year={','.join(fiscal_year)}")
    if service:
        active_filters.append(f"service={','.join(service)}")
    if exhibit_type:
        active_filters.append(f"exhibit_type={','.join(exhibit_type)}")
    if pe_number:
        active_filters.append(f"pe_number={','.join(pe_number)}")
    if appropriation_code:
        active_filters.append(f"appropriation_code={','.join(appropriation_code)}")
    if budget_type:
        active_filters.append(f"budget_type={','.join(budget_type)}")
    if q:
        active_filters.append(f"q={q}")
    if min_amount is not None:
        active_filters.append(f"min_amount={min_amount}")
    if max_amount is not None:
        active_filters.append(f"max_amount={max_amount}")
    filter_summary = "; ".join(active_filters) if active_filters else "none"

    if fmt == "csv":
        def csv_stream():
            buf = io.StringIO()
            # EAGLE-6: Source attribution header rows
            writer_raw = csv.writer(buf)
            writer_raw.writerow([f"# Source: DoD Budget Explorer"])
            writer_raw.writerow([f"# Export Date: {export_date}"])
            writer_raw.writerow([f"# Filters: {filter_summary}"])
            writer_raw.writerow([f"# URL: {export_url}"])
            writer_raw.writerow([f"# Total Records: {total_count}"])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate()
            writer = csv.DictWriter(buf, fieldnames=export_cols)
            writer.writeheader()
            yield buf.getvalue()
            for row in _iter_rows(conn, sql, params):
                buf.seek(0)
                buf.truncate()
                writer.writerow(dict(zip(export_cols, row)))
                yield buf.getvalue()

        return StreamingResponse(
            csv_stream(),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=budget_lines.csv",
                **extra_headers,
            },
        )

    if fmt == "xlsx":
        # DL-001/JS-001: Excel export using openpyxl write_only mode
        import openpyxl

        def xlsx_bytes() -> bytes:
            wb = openpyxl.Workbook(write_only=True)
            # EAGLE-6: Metadata sheet with source attribution
            meta_ws = wb.create_sheet("Metadata")
            meta_ws.append(["Source", "DoD Budget Explorer"])
            meta_ws.append(["Export Date", export_date])
            meta_ws.append(["Filters", filter_summary])
            meta_ws.append(["URL", export_url])
            meta_ws.append(["Total Records", total_count])
            # Data sheet
            ws = wb.create_sheet("Budget Lines")
            ws.append(export_cols)  # header row (FE-011: respects column subset)
            for row in _iter_rows(conn, sql, params):
                ws.append(list(row))
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        content = xlsx_bytes()
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=budget_lines.xlsx",
                "Content-Length": str(len(content)),
                **extra_headers,
            },
        )

    # JSON newline-delimited (NDJSON)
    # EAGLE-6: First line is metadata object
    def json_stream():
        metadata = {
            "_metadata": {
                "source": "DoD Budget Explorer",
                "export_date": export_date,
                "filters": filter_summary,
                "url": export_url,
                "total_records": total_count,
            }
        }
        yield json.dumps(metadata, default=str) + "\n"
        for row in _iter_rows(conn, sql, params):
            d = dict(zip(export_cols, row))
            yield json.dumps(d, default=str) + "\n"

    return StreamingResponse(
        json_stream(),
        media_type="application/x-ndjson",
        headers={
            "Content-Disposition": "attachment; filename=budget_lines.ndjson",
            **extra_headers,
        },
    )
