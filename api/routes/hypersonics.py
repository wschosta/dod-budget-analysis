"""
Hypersonics PE lines endpoints.

Returns a pivoted view of all budget lines related to hypersonics programs,
FY2015 onward. Sources: budget_lines keyword search + pe_descriptions narrative search.

Matching logic (PE-level): a PE number is included if ANY of its budget_lines rows match
a keyword OR if any pe_descriptions row for that PE matches a keyword. Once a PE is
matched, ALL of its sub-elements (line_item_title rows) are returned — not just the
rows that individually match a keyword.

Data is served from a pre-computed ``hypersonics_cache`` table for instant reads.
Call ``rebuild_hypersonics_cache(conn)`` after pipeline/enrichment runs, or hit the
POST /api/v1/hypersonics/rebuild endpoint to refresh.

Endpoints:
  GET  /api/v1/hypersonics          — JSON, pivoted table data
  GET  /api/v1/hypersonics/download — streaming CSV of pivoted table
  GET  /api/v1/hypersonics/debug    — data-quality checks
  POST /api/v1/hypersonics/rebuild  — rebuild the materialized cache table
"""

from __future__ import annotations

import csv
import io
import json
import logging
import sqlite3
from collections import Counter, OrderedDict
from typing import Any

from fastapi import APIRouter, Body, Depends, Query
from fastapi.responses import Response

from api.database import get_db
from api.routes.keyword_search import (
    FY_END,
    FY_START,
    apply_filters,
    build_cache_table,
    cache_rows_to_dicts,
    ensure_cache,
    load_per_fy_descriptions,
    xlsx_base_styles,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/hypersonics", tags=["hypersonics"])

# ── Keywords ──────────────────────────────────────────────────────────────────

_HYPERSONICS_KEYWORDS = [
    # ── Generic / cross-program ───────────────────────────────────────────
    "hypersonic",           # hypersonics, hypersonic glide, hypersonic weapon…
    "boost glide",          # boost-glide vehicles (all services)
    "glide body",           # Common Hypersonic Glide Body (C-HGB)
    "glide vehicle",        # generic glide vehicle references
    "scramjet",             # air-breathing hypersonic propulsion

    # ── Offensive — Air Force ──────────────────────────────────────────────
    "ARRW",                 # Air-Launched Rapid Response Weapon (AGM-183A)
    "AGM-183",              # ARRW missile designation
    "HACM",                 # Hypersonic Attack Cruise Missile
    "HCSW",                 # Hypersonic Conventional Strike Weapon (cancelled FY20)

    # ── Offensive — Army ───────────────────────────────────────────────────
    "LRHW",                 # Long Range Hypersonic Weapon / Dark Eagle battery
    "Dark Eagle",           # LRHW battery name
    "OpFires",              # Operational Fires (hypersonic component)

    # ── Offensive — Navy / Joint ───────────────────────────────────────────
    "C-HGB",                # Common Hypersonic Glide Body (joint Army/Navy)
    "CHGB",                 # alternate abbreviation
    "conventional prompt strike",   # Navy CPS program
    "prompt strike",        # catches "Intermediate Range Conventional Prompt Strike"

    # ── Offensive — Navy / SM-6 / OASUW ──────────────────────────────────
    "offensive anti",       # Offensive Anti-Surface Warfare references
    "oasuw",                # Offensive Anti-Surface Warfare (OASUW)
    "standard missile 6",   # Standard Missile 6 (SM-6)
    "sm-6",                 # SM-6 abbreviation
    "blk ib",               # Block IB variant designation
    "increment ii",         # program increment references

    # ── Generic speed / regime ─────────────────────────────────────────────
    "high speed",           # high-speed strike / high-speed weapons
    "mach",                 # Mach-regime references (broad)
    "conventional prompt",  # broader catch for conventional prompt programs

    # ── Defensive / tracking ───────────────────────────────────────────────
    "Glide Phase Interceptor",  # GPI — MDA program to defeat HGVs in glide
    "HBTSS",                # Hypersonic and Ballistic Tracking Space Sensor
]

# Keywords excluded from description search (too noisy for narrative matching).
_DESC_EXCLUDE = {"glide body", "AGM-183"}
_DESC_KEYWORDS = [kw for kw in _HYPERSONICS_KEYWORDS if kw not in _DESC_EXCLUDE]

_CACHE_TABLE = "hypersonics_cache"

# Fixed (non-FY) columns in the XLSX export, in order.
_XLSX_FIXED_HEADERS: list[str] = [
    "PE Number", "Service/Org", "Exhibit", "Line Item / Sub-Program",
    "Budget Activity", "Color of Money",
]
_COLS_PER_FY = 4  # value, in-total, source, description


def _fy_quad(
    fy_col_idx: int, fixed_col_count: int = len(_XLSX_FIXED_HEADERS)
) -> tuple[int, int, int, int]:
    """Return 1-indexed (value_col, intotal_col, source_col, desc_col) for an FY column index."""
    base = fixed_col_count + (fy_col_idx * _COLS_PER_FY) + 1
    return base, base + 1, base + 2, base + 3


# PEs to always include regardless of keyword matching
_EXTRA_PES = [
    "0101101F", "0210600A", "0601102F", "0601153N", "0602102F",
    "0602114N", "0602235N", "0602602F", "0602750N", "0603032F",
    "0603183D8Z", "0603273F", "0603467E", "0603601F", "0603673N",
    "0603680D8Z", "0603680F", "0603941D8Z", "0603945D8Z", "0604250D8Z",
    "0604331D8Z", "0604940D8Z", "0605456A", "0607210D8Z", "0902199D8Z",
]


# ── Cache management ──────────────────────────────────────────────────────────

def rebuild_hypersonics_cache(conn: sqlite3.Connection) -> int:
    """Rebuild the hypersonics_cache table from budget_lines + pe_descriptions."""
    logger.info("Rebuilding hypersonics cache table...")
    return build_cache_table(
        conn, _CACHE_TABLE, _HYPERSONICS_KEYWORDS, _DESC_KEYWORDS,
        fy_start=FY_START, fy_end=FY_END, extra_pes=_EXTRA_PES,
    )


def _ensure_cache(conn: sqlite3.Connection) -> bool:
    """Ensure hypersonics_cache exists and is populated."""
    return ensure_cache(
        conn, _CACHE_TABLE, _HYPERSONICS_KEYWORDS, _DESC_KEYWORDS,
        fy_start=FY_START, fy_end=FY_END, extra_pes=_EXTRA_PES,
    )


def _query_cache(
    conn: sqlite3.Connection,
    service: str | None = None,
    exhibit: str | None = None,
) -> tuple[list[dict], list[int]]:
    """Query the hypersonics cache with optional filters. Returns (items, year_range)."""
    _ensure_cache(conn)
    extra_where, extra_params = apply_filters(service, exhibit, None, None)
    where = f"WHERE {extra_where}" if extra_where else ""
    sql = f"SELECT * FROM {_CACHE_TABLE} {where} ORDER BY pe_number, exhibit_type, line_item_title"
    rows = conn.execute(sql, extra_params).fetchall()
    return cache_rows_to_dicts(rows), list(range(FY_START, FY_END + 1))


# ── GET /api/v1/hypersonics ───────────────────────────────────────────────────

@router.get(
    "",
    summary="Pivoted hypersonics PE lines from materialized cache",
)
def get_hypersonics(
    service: str | None = Query(None, description="Filter by service/org name (substring)"),
    exhibit: str | None = Query(None, description="Filter by exhibit type (exact, e.g. 'r2')"),
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Return all hypersonics-related budget line sub-elements as a pivoted table."""
    items, year_range = _query_cache(conn, service, exhibit)
    active_years = (
        [yr for yr in year_range if any(r.get(f"fy{yr}") is not None for r in items)]
        if items else year_range
    )

    return {
        "count": len(items),
        "fiscal_years": active_years,
        "keywords": _HYPERSONICS_KEYWORDS,
        "rows": items,
    }


# ── GET /api/v1/hypersonics/download ─────────────────────────────────────────

@router.get(
    "/download",
    summary="Download hypersonics PE lines as CSV",
    response_class=Response,
)
def download_hypersonics(
    service: str | None = Query(None),
    exhibit: str | None = Query(None),
    conn: sqlite3.Connection = Depends(get_db),
) -> Response:
    """Download pivoted hypersonics PE lines as CSV."""
    items, year_range = _query_cache(conn, service, exhibit)

    fy_headers: list[str] = []
    for yr in year_range:
        fy_headers.extend([f"FY{yr} ($K)", f"FY{yr} Source"])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "PE Number", "Service/Org", "Exhibit", "Line Item / Sub-Program",
        "Budget Activity", "Budget Activity (Normalized)", "Appropriation",
        "Color of Money", "Keywords (Row)", "Keywords (Desc)", "Description",
        *fy_headers,
    ])
    for r in items:
        fy_cells: list[Any] = []
        for yr in year_range:
            fy_cells.append(r.get(f"fy{yr}"))
            fy_cells.append(r.get("refs", {}).get(f"fy{yr}", ""))
        writer.writerow([
            r["pe_number"], r["organization_name"], r["exhibit_type"],
            r["line_item_title"], r["budget_activity_title"],
            r["budget_activity_norm"], r["appropriation_title"],
            r["color_of_money"],
            ", ".join(r.get("matched_keywords_row", [])),
            ", ".join(r.get("matched_keywords_desc", [])),
            r.get("description_text", ""),
            *fy_cells,
        ])

    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="hypersonics_pe_lines.csv"'},
    )


# ── Summary sheet builder ────────────────────────────────────────────────────


def _build_summary_sheet(
    wb: Any,
    selected_rows: list[tuple[str, dict]],
    active_years: list[int],
    total_set: set[str],
    first_data_row: int,
    last_data_row: int,
    val_letters: list[str] | None = None,
    intotal_letters: list[str] | None = None,
) -> None:
    """Build a Summary sheet with PE pivot tables and dimension breakdowns.

    All value cells use SUMIFS formulas referencing the Hypersonics data sheet
    so totals stay live when users edit Y/N/P flags.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Summary")
    get_col_letter = openpyxl.utils.get_column_letter

    sty = xlsx_base_styles()
    header_fill = sty["header_fill"]
    header_font = sty["header_font"]
    total_font = sty["total_font"]
    base_font = sty["base_font"]
    money_fmt = sty["money_fmt"]
    title_font = Font(bold=True, size=12)

    # Data-sheet column letters for each active year (reuse from caller if provided)
    if val_letters is None or intotal_letters is None:
        val_letters = []
        intotal_letters = []
        for fi in range(len(active_years)):
            vc, ic, _, _ = _fy_quad(fi)
            val_letters.append(get_col_letter(vc))
            intotal_letters.append(get_col_letter(ic))

    # Collect sorted unique dimension values from selected rows.
    def _uniq(key: str) -> list[str]:
        return sorted(v for v in dict.fromkeys(row.get(key, "") for _, row in selected_rows) if v)

    unique_pes = _uniq("pe_number")
    unique_svcs = _uniq("organization_name")
    unique_bas = _uniq("budget_activity_norm")
    unique_coms = _uniq("color_of_money")

    # Precompute column letters for PE sections (col 2+) and dim table (col 4+)
    pe_fy_letters = [get_col_letter(2 + yi) for yi in range(len(active_years))]
    dim_fy_letters = [get_col_letter(4 + yi) for yi in range(len(active_years))]

    # ── Helper: write a PE-summary section (Y, P, or Grand Total) ──

    def _write_pe_section(
        start_row: int,
        title: str,
        criteria: str | None,
        y_data_start: int = 0,
        p_data_start: int = 0,
    ) -> tuple[int, int]:
        """Write a PE summary section. Returns (next_row, data_start_row)."""
        r = start_row
        # Title
        ws.cell(row=r, column=1, value=title).font = title_font
        ws.merge_cells(
            start_row=r, start_column=1,
            end_row=r, end_column=1 + len(active_years),
        )
        r += 1
        # Header
        ws.cell(row=r, column=1, value="PE Number").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        for yi, yr in enumerate(active_years):
            c = ws.cell(row=r, column=2 + yi, value=f"FY{yr}")
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        r += 1
        data_start = r

        # Data rows
        for pe in unique_pes:
            ws.cell(row=r, column=1, value=pe).font = base_font
            for yi in range(len(active_years)):
                if criteria:
                    vr = f"Hypersonics!${val_letters[yi]}${first_data_row}:${val_letters[yi]}${last_data_row}"
                    pr = f"Hypersonics!$A${first_data_row}:$A${last_data_row}"
                    ir = f"Hypersonics!${intotal_letters[yi]}${first_data_row}:${intotal_letters[yi]}${last_data_row}"
                    formula = f'=SUMIFS({vr},{pr},$A{r},{ir},"{criteria}")'
                else:
                    offset = r - data_start
                    cl = pe_fy_letters[yi]
                    formula = f"={cl}{y_data_start + offset}+{cl}{p_data_start + offset}"
                c = ws.cell(row=r, column=2 + yi, value=formula)
                c.font = base_font
                c.number_format = money_fmt
            r += 1

        # Total row
        ws.cell(row=r, column=1, value="Total").font = total_font
        for yi in range(len(active_years)):
            cl = pe_fy_letters[yi]
            c = ws.cell(row=r, column=2 + yi, value=f"=SUM({cl}{data_start}:{cl}{r - 1})")
            c.font = total_font
            c.number_format = money_fmt
        r += 1
        return r, data_start

    # ── Sections A/B/C: PE Summary — Y / P / Grand Total ──
    cur = 1
    cur, y_ds = _write_pe_section(cur, "PE Summary \u2014 Y Values", "Y")
    cur += 1  # blank row
    cur, p_ds = _write_pe_section(cur, "PE Summary \u2014 P Values", "P")
    cur += 1
    cur, _ = _write_pe_section(
        cur, "PE Summary \u2014 Grand Total", None,
        y_data_start=y_ds, p_data_start=p_ds,
    )
    cur += 2  # spacing before dimension table

    # ── Section D: Summary by Dimension ──
    ws.cell(row=cur, column=1, value="Summary by Service/Agency, Budget Activity, Color of Money").font = title_font
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=3 + len(active_years))
    cur += 1

    # Dimension table header
    for ci, h in enumerate(["Category", "Value", "Type"]):
        c = ws.cell(row=cur, column=ci + 1, value=h)
        c.font = header_font
        c.fill = header_fill
    for yi, yr in enumerate(active_years):
        c = ws.cell(row=cur, column=4 + yi, value=f"FY{yr}")
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    cur += 1

    # Data-sheet column letters per dimension
    dim_groups: list[tuple[str, list[str], str]] = [
        ("Service/Agency", unique_svcs, "$B"),
        ("Budget Activity", unique_bas, "$E"),
        ("Color of Money", unique_coms, "$F"),
    ]

    for category, values, data_col in dim_groups:
        group_start = cur
        for val in values:
            for type_label, crit in [("Y", "Y"), ("P", "P"), ("Grand Total", None)]:
                ws.cell(row=cur, column=1, value=category).font = base_font
                ws.cell(row=cur, column=2, value=val).font = base_font
                fnt = total_font if type_label == "Grand Total" else base_font
                ws.cell(row=cur, column=3, value=type_label).font = fnt
                for yi in range(len(active_years)):
                    if crit:
                        vr = f"Hypersonics!${val_letters[yi]}${first_data_row}:${val_letters[yi]}${last_data_row}"
                        mr = f"Hypersonics!{data_col}${first_data_row}:{data_col}${last_data_row}"
                        ir = f"Hypersonics!${intotal_letters[yi]}${first_data_row}:${intotal_letters[yi]}${last_data_row}"
                        formula = f'=SUMIFS({vr},{mr},$B{cur},{ir},"{crit}")'
                    else:
                        cl = dim_fy_letters[yi]
                        formula = f"={cl}{cur - 2}+{cl}{cur - 1}"
                    c = ws.cell(row=cur, column=4 + yi, value=formula)
                    c.font = fnt
                    c.number_format = money_fmt
                cur += 1

        # Group subtotal rows (Y / P / Grand Total)
        for type_label, crit in [("Y", "Y"), ("P", "P"), ("Grand Total", None)]:
            ws.cell(row=cur, column=1, value=f"{category} Total").font = total_font
            ws.cell(row=cur, column=2, value="").font = total_font
            ws.cell(row=cur, column=3, value=type_label).font = total_font
            for yi in range(len(active_years)):
                cl = dim_fy_letters[yi]
                if crit:
                    offset = 0 if crit == "Y" else 1
                    refs = [f"{cl}{group_start + i * 3 + offset}" for i in range(len(values))]
                    formula = "=" + "+".join(refs)
                else:
                    formula = f"={cl}{cur - 2}+{cl}{cur - 1}"
                c = ws.cell(row=cur, column=4 + yi, value=formula)
                c.font = total_font
                c.number_format = money_fmt
            cur += 1
        cur += 1  # blank row between groups

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    for yi in range(len(active_years)):
        ws.column_dimensions[pe_fy_letters[yi]].width = 14
        ws.column_dimensions[dim_fy_letters[yi]].width = 14

    ws.freeze_panes = "B2"


# ── POST /api/v1/hypersonics/download/xlsx ────────────────────────────────────

@router.post(
    "/download/xlsx",
    summary="Download selected hypersonics rows as XLSX with formatting",
    response_class=Response,
)
def download_hypersonics_xlsx(
    show_ids: list[str] = Body(..., description="data-idx values of SHOW-checked rows"),
    total_ids: list[str] = Body(..., description="data-idx values of TOTAL-checked rows"),
    conn: sqlite3.Connection = Depends(get_db),
) -> Response:
    """Download an XLSX with all SHOW-checked rows.

    Rows that are SHOW-only (not TOTAL-checked) are rendered in italics.
    Rows that are TOTAL-checked get normal formatting.
    A "Totals" row is appended at the bottom summing the TOTAL-checked FY columns.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation

    _ensure_cache(conn)

    year_range = list(range(FY_START, FY_END + 1))

    sql = f"SELECT * FROM {_CACHE_TABLE} ORDER BY pe_number, exhibit_type, line_item_title"
    all_rows = conn.execute(sql).fetchall()
    items = cache_rows_to_dicts(all_rows)

    pe_groups: OrderedDict[str, list[dict]] = OrderedDict()
    for item in items:
        pe = item["pe_number"]
        pe_groups.setdefault(pe, []).append(item)

    show_set = set(show_ids)
    total_set = set(total_ids)

    selected_rows: list[tuple[str, dict]] = []
    for pe, children in pe_groups.items():
        for i, child in enumerate(children):
            idx = f"{pe}-{i}"
            if idx in show_set:
                selected_rows.append((idx, child))

    if not selected_rows:
        return Response(
            content=b"No rows selected",
            media_type="text/plain",
            status_code=400,
        )

    # Detect which FY columns have any data in selected rows
    active_years = [
        yr for yr in year_range
        if any(row.get(f"fy{yr}") is not None for _, row in selected_rows)
    ]

    desc_by_pe_fy = load_per_fy_descriptions(
        conn, {row.get("pe_number", "") for _, row in selected_rows}
    )

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hypersonics"

    # Styles
    sty = xlsx_base_styles()
    header_fill = sty["header_fill"]
    header_font = sty["header_font"]
    base_font = sty["base_font"]
    total_font = sty["total_font"]
    money_fmt = sty["money_fmt"]
    source_font = sty["source_font"]
    desc_font = sty["desc_font"]

    # Conditional formatting styles (FormulaRule accepts font/fill directly)
    cf_bold = Font(bold=True)
    cf_italic_gray = Font(italic=True, color="888888")
    cf_green = PatternFill(bgColor="C6EFCE")
    cf_yellow = PatternFill(bgColor="FFEB9C")
    cf_red = PatternFill(bgColor="FFC7CE")

    headers: list[str] = list(_XLSX_FIXED_HEADERS)
    for yr in active_years:
        headers.extend([
            f"FY{yr} ($K)", f"FY{yr} In Total",
            f"FY{yr} Source", f"FY{yr} Description",
        ])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    first_data_row = 2
    row_num = first_data_row
    for data_idx, row in selected_rows:
        is_total = data_idx in total_set

        vals = [
            row.get("pe_number", ""),
            row.get("organization_name", ""),
            row.get("exhibit_type", ""),
            row.get("line_item_title", ""),
            row.get("budget_activity_norm", ""),
            row.get("color_of_money", ""),
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row_num, column=i, value=v).font = base_font

        pe = row.get("pe_number", "")
        refs_map = row.get("refs", {}) or {}
        for fy_col_idx, yr in enumerate(active_years):
            val_col, intotal_col, src_col, desc_col = _fy_quad(fy_col_idx)

            amount = row.get(f"fy{yr}")
            val_cell = ws.cell(row=row_num, column=val_col, value=amount)
            val_cell.font = base_font
            if amount is not None:
                val_cell.number_format = money_fmt

            ws.cell(
                row=row_num, column=intotal_col,
                value="Y" if is_total else "N",
            ).font = base_font

            source_ref = refs_map.get(f"fy{yr}", "")
            if source_ref:
                src_cell = ws.cell(row=row_num, column=src_col, value=source_ref)
                src_cell.font = source_font
                src_cell.alignment = Alignment(wrap_text=True, vertical="top")

            desc_text = desc_by_pe_fy.get((pe, str(yr)), "")
            if desc_text:
                desc_cell = ws.cell(row=row_num, column=desc_col, value=desc_text)
                desc_cell.font = desc_font
                desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_num += 1

    last_data_row = row_num - 1

    # Precompute FY column numbers and letters once for validation, CF, and totals
    get_col_letter = openpyxl.utils.get_column_letter
    fy_cols: list[tuple[int, int, int, int, str, str, str, str]] = []
    for fi in range(len(active_years)):
        vc, ic, sc, dc = _fy_quad(fi)
        fy_cols.append((vc, ic, sc, dc,
                        get_col_letter(vc), get_col_letter(ic),
                        get_col_letter(sc), get_col_letter(dc)))
    intotal_letters = [fc[5] for fc in fy_cols]

    # ── Data validation (Y/N/P dropdown on In Total columns) ──
    dv = DataValidation(type="list", formula1='"Y,N,P"', allow_blank=False)
    dv.error = "Please enter Y, N, or P"
    dv.errorTitle = "Invalid value"
    for _, _, _, _, _, il, _, _ in fy_cols:
        dv.add(f"{il}{first_data_row}:{il}{last_data_row}")
    ws.add_data_validation(dv)

    # ── Conditional formatting ──
    if last_data_row >= first_data_row and active_years:
        for _, _, _, _, vl, il, sl, dl in fy_cols:
            for col_letter in [vl, sl, dl]:
                rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'=${il}{first_data_row}="Y"'],
                    font=cf_bold, stopIfTrue=True,
                ))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'=${il}{first_data_row}="N"'],
                    font=cf_italic_gray, stopIfTrue=True,
                ))
            it_rng = f"{il}{first_data_row}:{il}{last_data_row}"
            for val, cf_fill in [("Y", cf_green), ("P", cf_yellow), ("N", cf_red)]:
                ws.conditional_formatting.add(it_rng, FormulaRule(
                    formula=[f'=${il}{first_data_row}="{val}"'],
                    fill=cf_fill, stopIfTrue=True,
                ))

        fixed_last = get_col_letter(len(_XLSX_FIXED_HEADERS))
        fixed_rng = f"A{first_data_row}:{fixed_last}{last_data_row}"
        or_y = ",".join(f'${lt}{first_data_row}="Y"' for lt in intotal_letters)
        and_n = ",".join(f'${lt}{first_data_row}="N"' for lt in intotal_letters)
        ws.conditional_formatting.add(fixed_rng, FormulaRule(
            formula=[f"=OR({or_y})"], font=cf_bold, stopIfTrue=True,
        ))
        ws.conditional_formatting.add(fixed_rng, FormulaRule(
            formula=[f"=AND({and_n})"], font=cf_italic_gray, stopIfTrue=True,
        ))

    # ── Totals rows (Y / P / Grand Total) ──
    if last_data_row >= first_data_row and active_years:
        y_row, p_row, grand_row = row_num, row_num + 1, row_num + 2
        ws.cell(row=y_row, column=1, value="Y TOTALS").font = total_font
        ws.cell(row=p_row, column=1, value="P TOTALS").font = total_font
        ws.cell(row=grand_row, column=1, value="GRAND TOTAL").font = total_font
        for vc, ic, _, _, vl, il, _, _ in fy_cols:
            val_rng = f"${vl}${first_data_row}:${vl}${last_data_row}"
            it_rng = f"${il}${first_data_row}:${il}${last_data_row}"
            for tr, label, criteria in [(y_row, "Y Sum", "Y"), (p_row, "P Sum", "P")]:
                ws.cell(row=tr, column=ic, value=label).font = total_font
                c = ws.cell(
                    row=tr, column=vc,
                    value=f'=SUMIF({it_rng},"{criteria}",{val_rng})',
                )
                c.font = total_font
                c.number_format = money_fmt
            ws.cell(row=grand_row, column=ic, value="Grand Sum").font = total_font
            c = ws.cell(row=grand_row, column=vc, value=f"={vl}{y_row}+{vl}{p_row}")
            c.font = total_font
            c.number_format = money_fmt
        row_num = grand_row + 1

    # ── Column widths ──
    col_widths: list[int] = [14, 14, 8, 50, 20, 12]
    for _ in active_years:
        col_widths.extend([14, 10, 30, 40])
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_col_letter(i)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ──
    _build_summary_sheet(
        wb, selected_rows, active_years, total_set,
        first_data_row, last_data_row,
        val_letters=[fc[4] for fc in fy_cols],
        intotal_letters=intotal_letters,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="hypersonics_selected.xlsx"'},
    )


# ── POST /api/v1/hypersonics/rebuild ──────────────────────────────────────────

@router.post(
    "/rebuild",
    summary="Rebuild the materialized hypersonics cache table",
)
def rebuild_cache(conn: sqlite3.Connection = Depends(get_db)) -> dict:
    """Rebuild the hypersonics_cache table from budget_lines + pe_descriptions."""
    count = rebuild_hypersonics_cache(conn)
    return {"status": "ok", "rows": count}


# ── GET /api/v1/hypersonics/desc ──────────────────────────────────────────────

@router.get(
    "/desc/{pe_number}",
    summary="Get description text for a PE or R-2 project",
)
def get_description(
    pe_number: str,
    project: str | None = None,
    conn: sqlite3.Connection = Depends(get_db),
) -> dict:
    """Return description_text for a PE (R-1 level) or specific R-2 project row."""
    try:
        if project:
            row = conn.execute(
                f"SELECT description_text FROM {_CACHE_TABLE} "
                "WHERE pe_number = ? AND line_item_title = ? AND description_text IS NOT NULL LIMIT 1",
                [pe_number, project],
            ).fetchone()
        else:
            row = conn.execute(
                f"SELECT description_text FROM {_CACHE_TABLE} "
                "WHERE pe_number = ? AND exhibit_type = 'r2' AND description_text IS NOT NULL LIMIT 1",
                [pe_number],
            ).fetchone()
            if not row:
                row = conn.execute(
                    f"SELECT description_text FROM {_CACHE_TABLE} "
                    "WHERE pe_number = ? AND description_text IS NOT NULL LIMIT 1",
                    [pe_number],
                ).fetchone()
        return {"description": row[0] if row else None}
    except sqlite3.OperationalError:
        logger.warning("OperationalError fetching description for PE %s", pe_number, exc_info=True)
        return {"description": None}


# ── GET /api/v1/hypersonics/debug ─────────────────────────────────────────────

@router.get(
    "/debug",
    summary="Pre-flight data quality checks for the hypersonics view",
)
def debug_hypersonics(conn: sqlite3.Connection = Depends(get_db)) -> dict:
    """Surface data-quality stats to validate the hypersonics view against real data."""
    result: dict = {}

    # 1. Cache status
    try:
        cache_count = conn.execute(f"SELECT COUNT(*) FROM {_CACHE_TABLE}").fetchone()[0]
        distinct_pes = conn.execute(
            f"SELECT COUNT(DISTINCT pe_number) FROM {_CACHE_TABLE}"
        ).fetchone()[0]
        result["cache"] = {
            "table_exists": True,
            "row_count": cache_count,
            "distinct_pe_numbers": distinct_pes,
        }
    except sqlite3.OperationalError:
        result["cache"] = {"table_exists": False, "row_count": 0}

    # 2. pe_descriptions status
    try:
        conn.execute("SELECT 1 FROM pe_descriptions LIMIT 0")
        row_count = conn.execute("SELECT COUNT(*) FROM pe_descriptions").fetchone()[0]
        pe_count = conn.execute(
            "SELECT COUNT(DISTINCT pe_number) FROM pe_descriptions"
        ).fetchone()[0]
        result["pe_descriptions"] = {
            "table_exists": True,
            "row_count": row_count,
            "distinct_pe_numbers": pe_count,
            "populated": row_count > 0,
        }
    except sqlite3.OperationalError:
        result["pe_descriptions"] = {"table_exists": False, "populated": False}

    # 3. Keyword hit summary from cache
    if result.get("cache", {}).get("table_exists"):
        try:
            kw_rows = conn.execute(
                f"SELECT matched_keywords_row, matched_keywords_desc FROM {_CACHE_TABLE}"
            ).fetchall()
            row_counts: Counter[str] = Counter()
            desc_counts: Counter[str] = Counter()
            for r in kw_rows:
                row_counts.update(json.loads(r[0] or "[]"))
                desc_counts.update(json.loads(r[1] or "[]"))
            all_kw_counts = dict(row_counts + desc_counts)
            result["keyword_hit_counts"] = {"row": row_counts, "desc": desc_counts, "combined": all_kw_counts}
            result["keywords_with_zero_hits"] = [
                kw for kw in _HYPERSONICS_KEYWORDS if kw not in all_kw_counts
            ]
        except Exception:
            pass

    return result
