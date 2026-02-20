#!/usr/bin/env python3
"""
Exhibit Type Inventory Script — Step 1.B1-a

Walks the DoD_Budget_Documents directory, opens Excel files, and
generates a report of unique exhibit types, sheet names, and header patterns.

Usage:
    python exhibit_type_inventory.py                    # Scan default directory
    python exhibit_type_inventory.py --docs /path/to/docs  # Custom docs directory
    python exhibit_type_inventory.py --output report.txt    # Save to file
    python exhibit_type_inventory.py --verbose              # Show details
"""

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class ExhibitInventory:
    """Scans Excel files and inventories exhibit types."""

    def __init__(self, docs_dir: Path, verbose=False):
        """Initialize the scanner with a documents directory path and verbosity flag.

        Args:
            docs_dir: Path to the directory containing Excel budget documents.
            verbose: If True, print each sheet as it is scanned.
        """
        self.docs_dir = Path(docs_dir)
        self.verbose = verbose
        self.exhibits = defaultdict(lambda: {"files": [], "sheets": set(), "headers": set()})
        self.total_files = 0
        self.total_sheets = 0
        self.errors = []

    def _detect_exhibit_type(self, filename: str) -> str:
        """Simple exhibit type detection from filename."""
        name = filename.lower().replace("_display", "").replace(".xlsx", "")
        exhibit_types = ["p1r", "p1", "o1", "r1", "m1", "c1", "rf1"]
        for e_type in exhibit_types:
            if e_type in name:
                return e_type
        return "unknown"

    def _extract_headers(self, worksheet) -> list:
        """Extract header row from worksheet."""
        for row_idx, row in enumerate(worksheet.iter_rows(max_row=5, values_only=True)):
            # Look for "Account" or similar header indicator
            row_str = " ".join(str(v or "").lower() for v in row)
            if "account" in row_str or "budget" in row_str:
                # Found likely header row
                return [str(v or "").strip() for v in row if v]
        return []

    def scan(self):
        """Scan all Excel files in the documents directory."""
        if not self.docs_dir.exists():
            print(f"ERROR: Directory not found: {self.docs_dir}")
            sys.exit(1)

        xlsx_files = list(self.docs_dir.rglob("*.xlsx"))
        if not xlsx_files:
            print(f"WARNING: No .xlsx files found in {self.docs_dir}")
            return

        print(f"Scanning {len(xlsx_files)} Excel files...")

        for file_path in xlsx_files:
            try:
                self.total_files += 1
                exhibit_type = self._detect_exhibit_type(file_path.name)

                wb = openpyxl.load_workbook(str(file_path), read_only=True)
                for sheet_name in wb.sheetnames:
                    self.total_sheets += 1
                    ws = wb[sheet_name]

                    headers = self._extract_headers(ws)
                    header_sig = " | ".join(headers[:10])  # First 10 columns as signature

                    self.exhibits[exhibit_type]["files"].append(str(file_path.relative_to(self.docs_dir)))
                    self.exhibits[exhibit_type]["sheets"].add(sheet_name)
                    self.exhibits[exhibit_type]["headers"].add(header_sig[:100])

                    if self.verbose:
                        print(f"  ✓ {file_path.name} / {sheet_name}")

                wb.close()

            except Exception as e:
                self.errors.append({
                    "file": str(file_path.relative_to(self.docs_dir)),
                    "error": str(e),
                })
                print(f"  ⚠ ERROR in {file_path.name}: {e}")

    def report(self):
        """Generate summary report."""
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("EXHIBIT TYPE INVENTORY REPORT")
        report_lines.append("=" * 80)
        report_lines.append("")
        report_lines.append(f"Directory:      {self.docs_dir}")
        report_lines.append(f"Total Files:    {self.total_files}")
        report_lines.append(f"Total Sheets:   {self.total_sheets}")
        report_lines.append("")

        for exhibit_type in sorted(self.exhibits.keys()):
            data = self.exhibits[exhibit_type]
            report_lines.append(f"\nEXHIBIT TYPE: {exhibit_type.upper()}")
            report_lines.append("-" * 80)
            report_lines.append(f"  Files:       {len(set(data['files']))}")
            report_lines.append(f"  Sheets:      {len(data['sheets'])}")
            report_lines.append(f"  Header Sigs: {len(data['headers'])}")

            if self.verbose:
                report_lines.append("\n  Sheet Names:")
                for sheet in sorted(data["sheets"]):
                    report_lines.append(f"    - {sheet}")

                report_lines.append("\n  Header Signatures (first 3):")
                for i, header_sig in enumerate(sorted(data["headers"])[:3]):
                    report_lines.append(f"    [{i+1}] {header_sig[:70]}")

        if self.errors:
            report_lines.append("\n" + "=" * 80)
            report_lines.append("ERRORS")
            report_lines.append("=" * 80)
            for error in self.errors:
                report_lines.append(f"  {error['file']}: {error['error']}")

        report_lines.append("")
        return "\n".join(report_lines)

    def export_json(self, output_path: Path):
        """Export inventory as JSON."""
        data = {}
        for exhibit_type, info in self.exhibits.items():
            data[exhibit_type] = {
                "files": sorted(set(info["files"])),
                "sheets": sorted(info["sheets"]),
                "header_signatures": sorted(info["headers"]),
            }

        with open(output_path, "w") as f:
            json.dump(data, f, indent=2)

        print(f"✓ Exported JSON to {output_path}")

    def export_csv(self, output_path: Path):
        """Export inventory as CSV."""
        with open(output_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "ExhibitType",
                "FileCount",
                "SheetCount",
                "HeaderSignatures",
                "SampleFiles"
            ])

            for exhibit_type in sorted(self.exhibits.keys()):
                data = self.exhibits[exhibit_type]
                sample_files = " | ".join(sorted(set(data["files"]))[:2])
                writer.writerow([
                    exhibit_type,
                    len(set(data["files"])),
                    len(data["sheets"]),
                    len(data["headers"]),
                    sample_files,
                ])

        print(f"✓ Exported CSV to {output_path}")


def main():
    """Parse CLI arguments and run the exhibit type inventory scan."""
    parser = argparse.ArgumentParser(
        description="Inventory exhibit types in Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python exhibit_type_inventory.py
  python exhibit_type_inventory.py --output inventory.txt --verbose
  python exhibit_type_inventory.py --export-json inventory.json
        """,
    )
    parser.add_argument(
        "--docs",
        type=Path,
        default=Path("DoD_Budget_Documents"),
        help="Documents directory (default: DoD_Budget_Documents)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Save text report to file",
    )
    parser.add_argument(
        "--export-json",
        type=Path,
        default=None,
        help="Export structured data as JSON",
    )
    parser.add_argument(
        "--export-csv",
        type=Path,
        default=None,
        help="Export summary as CSV",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Verbose output with details",
    )

    args = parser.parse_args()

    inventory = ExhibitInventory(args.docs, verbose=args.verbose)
    inventory.scan()

    # Print report
    report = inventory.report()
    print(report)

    # Save report if requested
    if args.output:
        with open(args.output, "w") as f:
            f.write(report)
        print(f"✓ Saved report to {args.output}")

    # Export if requested
    if args.export_json:
        inventory.export_json(args.export_json)

    if args.export_csv:
        inventory.export_csv(args.export_csv)


if __name__ == "__main__":
    main()
