"""
Tests for exhibit_type_inventory.py — ExhibitInventory class

Verifies exhibit type detection from filenames, header extraction,
report generation, and JSON/CSV export using synthetic Excel fixtures.
"""
import csv
import json
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from exhibit_type_inventory import ExhibitInventory


def _create_xlsx(path: Path, headers: list[str], rows: list[list] = None):
    """Helper to create a minimal .xlsx file with given headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in (rows or []):
        ws.append(row)
    wb.save(str(path))
    wb.close()


@pytest.fixture()
def docs_dir(tmp_path):
    """Create a docs directory with sample Excel files."""
    docs = tmp_path / "docs"
    docs.mkdir()

    _create_xlsx(
        docs / "army_p1_fy2026.xlsx",
        ["Account", "Budget Activity", "FY2026 Request"],
        [[2035, "Aircraft", 1000]],
    )
    _create_xlsx(
        docs / "navy_r1_fy2026.xlsx",
        ["Account", "Budget Line", "FY2026 Request"],
        [[1319, "RDT&E", 2000]],
    )
    _create_xlsx(
        docs / "af_o1_fy2026.xlsx",
        ["Organization", "Budget Activity", "FY2026 Request"],
    )
    return docs


# ── _detect_exhibit_type ──────────────────────────────────────────────────────

class TestDetectExhibitType:
    def test_p1(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("army_p1_fy2026.xlsx") == "p1"

    def test_r1(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("navy_r1_display.xlsx") == "r1"

    def test_o1(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("AF_O1_FY2025.xlsx") == "o1"

    def test_m1(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("army_m1_fy2026.xlsx") == "m1"

    def test_c1(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("mc_c1_display.xlsx") == "c1"

    def test_unknown(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("random_data.xlsx") == "unknown"

    def test_p1r(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("army_p1r_fy2026.xlsx") == "p1r"

    def test_strips_display_suffix(self, tmp_path):
        inv = ExhibitInventory(tmp_path)
        assert inv._detect_exhibit_type("navy_r1_display.xlsx") == "r1"


# ── scan & report ─────────────────────────────────────────────────────────────

class TestScan:
    def test_scan_counts_files(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        assert inv.total_files == 3

    def test_scan_counts_sheets(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        assert inv.total_sheets == 3

    def test_scan_detects_exhibit_types(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        assert "p1" in inv.exhibits
        assert "r1" in inv.exhibits
        assert "o1" in inv.exhibits

    def test_scan_verbose(self, docs_dir, capsys):
        inv = ExhibitInventory(docs_dir, verbose=True)
        inv.scan()
        out = capsys.readouterr().out
        assert "army_p1_fy2026.xlsx" in out

    def test_scan_empty_dir(self, tmp_path, capsys):
        empty = tmp_path / "empty"
        empty.mkdir()
        inv = ExhibitInventory(empty)
        inv.scan()
        out = capsys.readouterr().out
        assert "No .xlsx files" in out
        assert inv.total_files == 0


class TestReport:
    def test_report_contains_exhibit_types(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        report = inv.report()
        assert "P1" in report
        assert "R1" in report
        assert "EXHIBIT TYPE INVENTORY REPORT" in report

    def test_report_contains_stats(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        report = inv.report()
        assert "Total Files:" in report
        assert "Total Sheets:" in report


# ── export_json / export_csv ──────────────────────────────────────────────────

class TestExportJson:
    def test_exports_valid_json(self, docs_dir, tmp_path):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        out_path = tmp_path / "inventory.json"
        inv.export_json(out_path)
        data = json.loads(out_path.read_text())
        assert "p1" in data
        assert "files" in data["p1"]
        assert "sheets" in data["p1"]


class TestExportCsv:
    def test_exports_valid_csv(self, docs_dir, tmp_path):
        inv = ExhibitInventory(docs_dir)
        inv.scan()
        out_path = tmp_path / "inventory.csv"
        inv.export_csv(out_path)
        with open(out_path) as f:
            reader = csv.reader(f)
            headers = next(reader)
            assert "ExhibitType" in headers
            rows = list(reader)
            assert len(rows) >= 2  # at least p1, r1


# ── _extract_headers ──────────────────────────────────────────────────────────

class TestExtractHeaders:
    def test_finds_account_header(self, docs_dir):
        inv = ExhibitInventory(docs_dir)
        wb = openpyxl.load_workbook(str(docs_dir / "army_p1_fy2026.xlsx"), read_only=True)
        ws = wb.active
        headers = inv._extract_headers(ws)
        wb.close()
        assert "Account" in headers

    def test_returns_empty_for_no_match(self, tmp_path):
        xlsx_path = tmp_path / "noheader.xlsx"
        _create_xlsx(xlsx_path, ["X", "Y", "Z"])
        inv = ExhibitInventory(tmp_path)
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
        headers = inv._extract_headers(ws)
        wb.close()
        assert headers == []
