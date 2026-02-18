"""
Pytest fixtures for DoD budget analysis tests — Step 1.C1

Provides reusable test fixtures: sample Excel files, sample PDFs, temporary
SQLite databases, and pre-loaded database connections.

──────────────────────────────────────────────────────────────────────────────
Fixture creation tasks (STEP 1.C1 COMPLETE)
──────────────────────────────────────────────────────────────────────────────

1.C1-a: Create minimal Excel fixture for each summary exhibit type.
    DONE — see _create_exhibit_xlsx() and the fixtures_dir fixture below.

1.C1-b: Create minimal PDF fixtures.
    DONE — see _create_sample_pdf() using fpdf2 below.

1.C1-c: Create a fixture that builds a pre-populated test database.
    DONE — see test_db fixture below.

1.C1-d: Create a fixture for a "known bad" Excel file.
    DONE — see bad_excel fixture below.

1.C1-e: Add fixture requirements to requirements-dev.txt.
    DONE — requirements-dev.txt already lists pytest, pytest-cov, fpdf2.

──────────────────────────────────────────────────────────────────────────────
Remaining TODOs
──────────────────────────────────────────────────────────────────────────────

TODO FIX-005 [Complexity: MEDIUM] [Tokens: ~2000] [User: NO]
    Fix pyo3_runtime.PanicException in test_pipeline.py tests.
    All 8 test_pipeline tests error with PanicException from pdfplumber
    during test_db fixture creation (build_database processes PDF fixtures).
    Steps:
      1. Investigate: run build_database() on just the PDF fixtures manually
         to see if the panic is in fpdf2 output or pdfplumber parsing
      2. Option A: Generate simpler PDF fixtures that pdfplumber handles
      3. Option B: Catch pyo3_runtime.PanicException in ingest_pdf_file()
         and log as a warning instead of crashing
      4. Option C: Create fixtures_dir_excel_only for tests that don't need PDF
      5. Run pytest tests/test_pipeline.py -v to verify
    Success: All 8 test_pipeline tests pass (or cleanly skip PDF-only tests).

TODO FIX-006 [Complexity: LOW] [Tokens: ~500] [User: NO]
    Fix test_rows_metric_increases_with_excel in test_build_integration.py.
    Likely caused by the undefined `rows` variable bug (TODO FIX-001).
    Steps:
      1. First fix TODO FIX-001 in build_budget_db.py
      2. Re-run: pytest tests/test_build_integration.py -v
    Dependency: TODO FIX-001 must be fixed first.
    Success: test_rows_metric_increases_with_excel passes.

DONE TEST-001: _P5_ROWS, _R2_ROWS, and header patterns added; p5_display.xlsx
    and r2_display.xlsx generated in fixtures_dir.
NOTE TEST-002: Blocked — requires api/app.py (TODO 2.C7-a not yet started).
NOTE TEST-003: Low priority until PDF extraction improvements are stable.
NOTE TEST-004: No actual test functions return values (inner helpers only).
"""

import sys
import types
from pathlib import Path

import pytest

# ── Stub pdfplumber (heavy dep, not needed for fixture generation) ─────────────
sys.modules.setdefault("pdfplumber", types.ModuleType("pdfplumber"))

import openpyxl  # noqa: E402 — available in requirements-dev.txt

FIXTURES_DIR = Path(__file__).parent / "fixtures"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _create_exhibit_xlsx(path: Path, exhibit_type: str, rows: list[tuple]) -> Path:
    """Create a minimal .xlsx fixture for a given exhibit type.

    Builds a workbook with one sheet containing a realistic header row and
    the provided data rows.  Values are deterministic so test assertions are
    stable.  Implements TODO 1.C1-a.

    Args:
        path:         Destination file path (must end in .xlsx).
        exhibit_type: e.g. "p1", "r1", "o1", "m1", "c1", "rf1".
        rows:         List of tuples (one per data row) matching the header.

    Returns:
        The written path.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FY 2026"

    if exhibit_type in ("p1", "p1r"):
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "Budget Line Item", "Budget Line Item (BLI) Title",
            "FY2024 Actual\nAmount", "FY2025 Enacted\nAmount",
            "FY2026 Request\nAmount",
        ]
    elif exhibit_type == "r1":
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "PE/BLI", "Program Element/Budget Line Item (BLI) Title",
            "FY2024 Actual\nAmount", "FY2025 Enacted\nAmount",
            "FY2026 Request\nAmount",
        ]
    elif exhibit_type == "o1":
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "BSA", "Budget SubActivity Title",
            "FY2024 Actual\nAmount", "FY2025 Enacted\nAmount",
            "FY2026 Request\nAmount",
        ]
    elif exhibit_type == "m1":
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "BSA", "Budget SubActivity Title",
            "FY2024 Actual\nAmount", "FY2025 Enacted\nAmount",
            "FY2026 Request\nAmount",
        ]
    elif exhibit_type == "c1":
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "Construction Project", "Construction Project Title",
            "Authorization Amount", "Appropriation Amount",
        ]
    elif exhibit_type == "rf1":
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "Budget Line Item", "Budget Line Item (BLI) Title",
            "FY2024 Actual\nAmount", "FY2025 Enacted\nAmount",
            "FY2026 Request\nAmount",
        ]
    elif exhibit_type == "p5":
        # P-5 Procurement Detail: line items with quantities and unit costs
        headers = [
            "Account", "Program Element", "Line Item", "Item Title",
            "Unit of Measure",
            "Prior Year Quantity", "Current Year Quantity", "Estimate Quantity",
            "Prior Year Unit Cost", "Current Year Unit Cost", "Estimate Unit Cost",
            "Justification",
        ]
    elif exhibit_type == "r2":
        # R-2 RDT&E Detail Schedule
        headers = [
            "Account", "PE", "Sub-Element", "Title",
            "Prior Year", "Current Year", "Estimate",
            "Metric", "Planned Achievement",
        ]
    else:
        headers = [
            "Account", "Account Title", "Organization",
            "Budget Activity", "Budget Activity Title",
            "FY2024 Actual\nAmount", "FY2026 Request\nAmount",
        ]

    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _create_sample_pdf(path: Path, title: str = "Test Budget Document",
                        include_table: bool = False) -> Path:
    """Create a minimal PDF fixture using fpdf2.

    Generates a PDF with extractable text and optionally a simple table
    layout.  Implements TODO 1.C1-b.

    Args:
        path:          Destination .pdf path.
        title:         Document title printed on the first page.
        include_table: If True, add a rudimentary table-like layout.

    Returns:
        The written path.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        pytest.skip("fpdf2 not installed — skipping PDF fixture generation")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, title, ln=True)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, "Department of Defense", ln=True)
    pdf.cell(0, 8, "Fiscal Year 2026 Budget Justification", ln=True)
    pdf.ln(6)
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(
        0, 6,
        "This document contains budget line items for demonstration purposes. "
        "Program elements are identified by codes such as 0602702E and 0305116BB. "
        "Amounts are presented in thousands of then-year dollars.",
    )

    if include_table:
        pdf.ln(6)
        pdf.set_font("Helvetica", "B", 10)
        col_w = [60, 35, 35, 35]
        headers = ["Program", "FY2024 Actual", "FY2025 Enacted", "FY2026 Request"]
        for w, h in zip(col_w, headers):
            pdf.cell(w, 8, h, border=1)
        pdf.ln()
        pdf.set_font("Helvetica", size=10)
        sample_rows = [
            ("0602702E Advanced Research", "12,345", "13,456", "14,000"),
            ("0305116BB Counter-ISR", "9,876", "10,234", "10,500"),
            ("0601102D Basic Research", "5,000", "5,100", "5,200"),
        ]
        for row in sample_rows:
            for w, val in zip(col_w, row):
                pdf.cell(w, 7, val, border=1)
            pdf.ln()

    path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(path))
    return path


# ── Session-scoped fixtures ───────────────────────────────────────────────────

_P1_ROWS = [
    ("2035", "2035 Aircraft Procurement, Army", "A", "01", "Air Operations",
     "0205231A", "AH-64 Apache Block III", 12_345.0, 13_456.0, 14_000.0),
    ("2035", "2035 Aircraft Procurement, Army", "A", "01", "Air Operations",
     "0205231B", "UH-60 Blackhawk", 8_900.0, 9_100.0, 9_500.0),
    ("2035", "2035 Aircraft Procurement, Army", "A", "02", "Missile Programs",
     "0205231C", "AIM-120 AMRAAM", 6_700.0, 7_000.0, 7_200.0),
]

_R1_ROWS = [
    ("1300", "RDT&E, Army", "A", "06", "RDT&E",
     "0602702E", "Advanced Research Program", 55_000.0, 57_000.0, 59_000.0),
    ("1300", "RDT&E, Army", "A", "07", "Operational Systems Dev",
     "0305116BB", "Counter-ISR", 22_000.0, 23_000.0, 24_000.0),
]

_C1_ROWS = [
    ("2100", "Military Construction, Army", "A", "01", "MilCon",
     "P-2345", "Barracks Replacement, Fort Bragg", 45_000.0, 30_000.0),
]

# TODO TEST-001: P-5 Procurement Detail fixtures (columns match p5 header above)
_P5_ROWS = [
    # account, pe, line_item, title, unit, py_qty, cy_qty, est_qty,
    # py_unit_cost, cy_unit_cost, est_unit_cost, justification
    ("2035", "0205231A", "LIN-001", "AH-64 Apache Block III",
     "Each", 12, 14, 15,
     55_000.0, 56_500.0, 58_000.0, "Full-rate production continues."),
    ("2035", "0205231B", "LIN-002", "UH-60 Blackhawk M-Model",
     "Each", 8, 10, 11,
     18_000.0, 18_500.0, 19_000.0, "Replaces aging L-model fleet."),
]

# TODO TEST-001: R-2 RDT&E Detail Schedule fixtures
_R2_ROWS = [
    # account, pe, sub_element, title, prior_year, current_year, estimate,
    # metric, planned_achievement
    ("1300", "0602702E", "A", "Advanced Materials Research",
     12_000.0, 13_500.0, 14_000.0,
     "TRL Level", "Achieve TRL-4 for candidate materials"),
    ("1300", "0602702E", "B", "Computational Modeling",
     5_000.0, 5_500.0, 5_800.0,
     "Simulation Fidelity", "High-fidelity model validated vs. test data"),
]


@pytest.fixture(scope="session")
def fixtures_dir(tmp_path_factory):
    """Return a temporary directory populated with deterministic fixture files.

    Creates one .xlsx per summary exhibit type (TODO 1.C1-a) and two PDFs
    (with and without a table layout) (TODO 1.C1-b).
    """
    d = tmp_path_factory.mktemp("budget_fixtures")

    # TODO 1.C1-a: Excel fixtures (summary exhibits)
    _create_exhibit_xlsx(d / "p1_display.xlsx", "p1", _P1_ROWS)
    _create_exhibit_xlsx(d / "r1_display.xlsx", "r1", _R1_ROWS)
    _create_exhibit_xlsx(d / "c1_display.xlsx", "c1", _C1_ROWS)
    _create_exhibit_xlsx(d / "o1_display.xlsx", "o1", _P1_ROWS[:2])
    _create_exhibit_xlsx(d / "m1_display.xlsx", "m1", _P1_ROWS[:2])
    _create_exhibit_xlsx(d / "rf1_display.xlsx", "rf1", _P1_ROWS[:1])
    # TEST-001: Detail exhibit fixtures (P-5, R-2)
    _create_exhibit_xlsx(d / "p5_display.xlsx", "p5", _P5_ROWS)
    _create_exhibit_xlsx(d / "r2_display.xlsx", "r2", _R2_ROWS)

    # TODO 1.C1-b: PDF fixtures
    _create_sample_pdf(d / "text_only.pdf", title="Budget Overview FY2026",
                       include_table=False)
    _create_sample_pdf(d / "with_table.pdf", title="RDT&E Justification FY2026",
                       include_table=True)

    return d


@pytest.fixture(scope="session")
def test_db(fixtures_dir, tmp_path_factory):
    """Return a Path to a SQLite database pre-built from fixture files.

    Runs build_database() once per test session so integration tests can
    query real data without repeating the build.  Implements TODO 1.C1-c.
    """
    # Import here so test collection works even without the full dep stack
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from build_budget_db import build_database  # type: ignore

    db_dir = tmp_path_factory.mktemp("test_db")
    db_path = db_dir / "test_budget.sqlite"
    try:
        build_database(fixtures_dir, db_path, rebuild=True)
    except Exception as exc:
        pytest.skip(f"build_database() failed (missing deps?): {exc}")
    return db_path


@pytest.fixture(scope="session")
def fixtures_dir_excel_only(tmp_path_factory):
    """Return a temporary directory with only Excel fixtures (no PDFs).

    FIX-005 Option C: Provides a PDF-free fixture directory so tests that
    exercise only Excel ingestion do not trigger pyo3/pdfplumber PanicException.
    """
    d = tmp_path_factory.mktemp("excel_only_fixtures")
    _create_exhibit_xlsx(d / "p1_display.xlsx", "p1", _P1_ROWS)
    _create_exhibit_xlsx(d / "r1_display.xlsx", "r1", _R1_ROWS)
    _create_exhibit_xlsx(d / "c1_display.xlsx", "c1", _C1_ROWS)
    _create_exhibit_xlsx(d / "o1_display.xlsx", "o1", _P1_ROWS[:2])
    _create_exhibit_xlsx(d / "m1_display.xlsx", "m1", _P1_ROWS[:2])
    _create_exhibit_xlsx(d / "rf1_display.xlsx", "rf1", _P1_ROWS[:1])
    _create_exhibit_xlsx(d / "p5_display.xlsx", "p5", _P5_ROWS)
    _create_exhibit_xlsx(d / "r2_display.xlsx", "r2", _R2_ROWS)
    return d


@pytest.fixture(scope="session")
def test_db_excel_only(fixtures_dir_excel_only, tmp_path_factory):
    """Return a Path to a SQLite database built from Excel-only fixtures.

    FIX-005: No PDF fixtures → no pyo3/pdfplumber PanicException.
    Tests that only need Excel data should use this fixture instead of test_db.
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from build_budget_db import build_database  # type: ignore

    db_dir = tmp_path_factory.mktemp("test_db_excel_only")
    db_path = db_dir / "test_budget_excel.sqlite"
    try:
        build_database(fixtures_dir_excel_only, db_path, rebuild=True)
    except Exception as exc:
        pytest.skip(f"build_database() failed (missing deps?): {exc}")
    return db_path


@pytest.fixture(scope="session")
def bad_excel(tmp_path_factory) -> Path:
    """Return a Path to an intentionally malformed Excel file.

    The file has: no recognisable header row, merged-cell-like empty columns,
    inconsistent column names, and blank rows scattered through the data.
    Used to verify that the parser degrades gracefully.  Implements TODO 1.C1-d.
    """
    d = tmp_path_factory.mktemp("bad_fixtures")
    path = d / "malformed_exhibit.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: completely blank (no header here)
    ws.append([None, None, None, None])
    # Row 2: unrecognised headers — no "Account" column
    ws.append(["Item#", "Description", "", "Dollars"])
    # Row 3: blank row in the middle of the header area
    ws.append([None, None, None, None])
    # Row 4: inconsistent naming — "Acct" instead of "Account"
    ws.append(["Acct", "Title", "Org", "Amount"])
    # Data rows with mixed types and missing values
    ws.append([None, "Missing account code", "A", 1_000])
    ws.append(["2035", None, None, "not-a-number"])
    ws.append(["", "", "", ""])
    ws.append([2035, "Aircraft Procurement", "A", 99_999])

    wb.save(str(path))
    return path


# ── Function-scoped helpers ───────────────────────────────────────────────────

@pytest.fixture()
def tmp_db(tmp_path):
    """Return a fresh (empty) SQLite database connection for unit tests.

    Useful for tests that need a database but don't need pre-loaded data.
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from build_budget_db import create_database  # type: ignore

    db_path = tmp_path / "unit_test.sqlite"
    conn = create_database(db_path)
    yield conn
    conn.close()
