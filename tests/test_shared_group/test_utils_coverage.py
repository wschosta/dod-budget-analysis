"""
Additional utility function tests — coverage gap fill

Tests for utility functions that were previously untested:
  - utils.manifest.compute_file_hash()
  - utils.database.disable_fts5_triggers() / enable_fts5_triggers()
  - scripts/generate_expected_output.py functions
"""
import hashlib
import json
import sqlite3
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "scripts"))

from utils.manifest import compute_file_hash
from utils.database import disable_fts5_triggers, enable_fts5_triggers
from generate_expected_output import (
    create_fixture_xlsx,
    create_expected_json,
    FIXTURES,
)


# ── compute_file_hash tests ──────────────────────────────────────────────────

class TestComputeFileHash:
    def test_known_content(self, tmp_path):
        """SHA-256 of known content matches expected digest."""
        path = tmp_path / "test.txt"
        path.write_bytes(b"hello world\n")
        expected = hashlib.sha256(b"hello world\n").hexdigest()
        assert compute_file_hash(path) == expected

    def test_empty_file(self, tmp_path):
        """Empty file returns the SHA-256 of empty bytes."""
        path = tmp_path / "empty.bin"
        path.write_bytes(b"")
        expected = hashlib.sha256(b"").hexdigest()
        assert compute_file_hash(path) == expected

    def test_binary_file(self, tmp_path):
        """Binary content is hashed correctly."""
        data = bytes(range(256)) * 100  # 25.6 KB of binary data
        path = tmp_path / "binary.bin"
        path.write_bytes(data)
        expected = hashlib.sha256(data).hexdigest()
        assert compute_file_hash(path) == expected

    def test_returns_hex_string(self, tmp_path):
        """Return value is a 64-character hex string."""
        path = tmp_path / "data.txt"
        path.write_text("test data")
        result = compute_file_hash(path)
        assert isinstance(result, str)
        assert len(result) == 64
        assert all(c in "0123456789abcdef" for c in result)

    def test_different_content_different_hash(self, tmp_path):
        """Different file contents produce different hashes."""
        path_a = tmp_path / "a.txt"
        path_b = tmp_path / "b.txt"
        path_a.write_text("content A")
        path_b.write_text("content B")
        assert compute_file_hash(path_a) != compute_file_hash(path_b)

    def test_same_content_same_hash(self, tmp_path):
        """Identical content in different files produces the same hash."""
        path_a = tmp_path / "copy1.txt"
        path_b = tmp_path / "copy2.txt"
        path_a.write_text("identical content")
        path_b.write_text("identical content")
        assert compute_file_hash(path_a) == compute_file_hash(path_b)


# ── FTS5 trigger helpers tests ───────────────────────────────────────────────

@pytest.fixture()
def fts_db():
    """Create an in-memory database with a simple table and FTS5 index."""
    conn = sqlite3.connect(":memory:")
    conn.execute("""
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            content TEXT
        )
    """)
    conn.execute("""
        CREATE VIRTUAL TABLE documents_fts USING fts5(
            content, content=documents
        )
    """)
    # Create initial triggers
    conn.execute("""
        CREATE TRIGGER documents_ai AFTER INSERT ON documents BEGIN
            INSERT INTO documents_fts(rowid, content)
            VALUES (new.id, new.content);
        END
    """)
    conn.execute("""
        CREATE TRIGGER documents_ad AFTER DELETE ON documents BEGIN
            INSERT INTO documents_fts(documents_fts, rowid, content)
            VALUES ('delete', old.id, old.content);
        END
    """)
    conn.execute("""
        CREATE TRIGGER documents_au AFTER UPDATE ON documents BEGIN
            INSERT INTO documents_fts(documents_fts, rowid, content)
            VALUES ('delete', old.id, old.content);
            INSERT INTO documents_fts(rowid, content)
            VALUES (new.id, new.content);
        END
    """)
    conn.commit()
    yield conn
    conn.close()


def _trigger_names(conn: sqlite3.Connection) -> set[str]:
    """Return the set of trigger names in the database."""
    rows = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='trigger'"
    ).fetchall()
    return {r[0] for r in rows}


class TestDisableFts5Triggers:
    def test_drops_all_three_triggers(self, fts_db):
        """All three trigger variants (ai, ad, au) are dropped."""
        assert "documents_ai" in _trigger_names(fts_db)
        assert "documents_ad" in _trigger_names(fts_db)
        assert "documents_au" in _trigger_names(fts_db)

        disable_fts5_triggers(fts_db, "documents")

        remaining = _trigger_names(fts_db)
        assert "documents_ai" not in remaining
        assert "documents_ad" not in remaining
        assert "documents_au" not in remaining

    def test_idempotent(self, fts_db):
        """Calling disable twice does not error."""
        disable_fts5_triggers(fts_db, "documents")
        disable_fts5_triggers(fts_db, "documents")  # should not raise
        assert "documents_ai" not in _trigger_names(fts_db)

    def test_no_error_for_nonexistent_table(self, fts_db):
        """Disabling triggers for a table with no triggers does not error."""
        disable_fts5_triggers(fts_db, "nonexistent_table")


class TestEnableFts5Triggers:
    def test_recreates_triggers(self, fts_db):
        """After disabling, enable recreates all triggers."""
        disable_fts5_triggers(fts_db, "documents")
        assert "documents_ai" not in _trigger_names(fts_db)

        enable_fts5_triggers(fts_db, "documents", "documents_fts")

        triggers = _trigger_names(fts_db)
        assert "documents_ai" in triggers
        assert "documents_ad" in triggers
        assert "documents_au" in triggers

    def test_insert_trigger_works_after_reenable(self, fts_db):
        """After re-enabling, inserts propagate to FTS."""
        disable_fts5_triggers(fts_db, "documents")
        enable_fts5_triggers(fts_db, "documents", "documents_fts")

        fts_db.execute(
            "INSERT INTO documents (content) VALUES (?)",
            ("budget justification for Army procurement",),
        )
        fts_db.commit()

        results = fts_db.execute(
            "SELECT content FROM documents_fts WHERE documents_fts MATCH 'Army'"
        ).fetchall()
        assert len(results) >= 1


# ── generate_expected_output function tests ──────────────────────────────────

class TestCreateFixtureXlsx:
    def test_creates_file(self, tmp_path):
        """create_fixture_xlsx creates an .xlsx file at the specified path."""
        fixture = FIXTURES[0]  # army_p1_fy2026
        path = create_fixture_xlsx(fixture, tmp_path)
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_file_has_correct_rows(self, tmp_path):
        """Created .xlsx has header + expected data row count."""
        import openpyxl

        fixture = FIXTURES[0]
        path = create_fixture_xlsx(fixture, tmp_path)

        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Header + data rows
        assert len(rows) == 1 + len(fixture["rows"])

    def test_headers_match_fixture_def(self, tmp_path):
        """Header row matches the fixture definition."""
        import openpyxl

        fixture = FIXTURES[0]
        path = create_fixture_xlsx(fixture, tmp_path)

        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(max_row=1, values_only=True))
        wb.close()

        assert list(header_row) == fixture["headers"]


class TestCreateExpectedJson:
    def test_creates_json_file(self, tmp_path):
        """create_expected_json creates a valid JSON file."""
        fixture = FIXTURES[0]
        path = create_expected_json(fixture, tmp_path)
        assert path.exists()
        assert path.suffix == ".json"

        data = json.loads(path.read_text())
        assert "exhibit_type" in data
        assert "row_count" in data

    def test_json_contains_fixture_metadata(self, tmp_path):
        """JSON file includes source_file, exhibit_type, and service."""
        fixture = FIXTURES[0]
        path = create_expected_json(fixture, tmp_path)
        data = json.loads(path.read_text())

        assert data["source_file"] == fixture["filename"]
        assert data["exhibit_type"] == fixture["exhibit_type"]
        assert data["service"] == fixture["service"]

    def test_expected_dir_created(self, tmp_path):
        """The expected/ subdirectory is created automatically."""
        fixture = FIXTURES[0]
        create_expected_json(fixture, tmp_path)
        assert (tmp_path / "expected").is_dir()

    def test_all_fixtures_have_definitions(self):
        """Every entry in FIXTURES has required keys."""
        for fixture in FIXTURES:
            assert "filename" in fixture
            assert "exhibit_type" in fixture
            assert "headers" in fixture
            assert "rows" in fixture
            assert "expected" in fixture
            assert "row_count" in fixture["expected"]
