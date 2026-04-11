"""Tests for api/routes/hypersonics.py — hypersonics PE lines endpoints."""

import csv
import io

import openpyxl
import pytest

from api.routes.hypersonics import _HYPERSONICS_KEYWORDS, _DESC_KEYWORDS


@pytest.fixture(scope="module", autouse=True)
def _rebuild_hypersonics_cache(client):
    """Rebuild the hypersonics cache once for the entire module."""
    client.post("/api/v1/hypersonics/rebuild")


class TestRebuildCache:
    def test_rebuild_returns_200(self, client):
        resp = client.post("/api/v1/hypersonics/rebuild")
        assert resp.status_code == 200
        body = resp.json()
        assert body["status"] == "ok"
        assert "rows" in body


class TestGetHypersonics:
    def test_get_returns_200(self, client):
        resp = client.get("/api/v1/hypersonics")
        assert resp.status_code == 200

    def test_response_structure(self, client):
        body = client.get("/api/v1/hypersonics").json()
        assert "count" in body
        assert "fiscal_years" in body
        assert "keywords" in body
        assert "rows" in body
        assert isinstance(body["rows"], list)
        assert isinstance(body["fiscal_years"], list)
        assert isinstance(body["keywords"], list)

    def test_filter_by_service(self, client):
        resp = client.get("/api/v1/hypersonics", params={"service": "Army"})
        assert resp.status_code == 200

    def test_filter_by_exhibit(self, client):
        resp = client.get("/api/v1/hypersonics", params={"exhibit": "r1"})
        assert resp.status_code == 200


class TestDownloadCSV:
    def test_csv_download(self, client):
        resp = client.get("/api/v1/hypersonics/download")
        assert resp.status_code == 200
        assert "text/csv" in resp.headers.get("content-type", "")

    def test_csv_has_content_disposition(self, client):
        resp = client.get("/api/v1/hypersonics/download")
        assert "attachment" in resp.headers.get("content-disposition", "")
        assert "csv" in resp.headers.get("content-disposition", "")

    def test_csv_parseable(self, client):
        resp = client.get("/api/v1/hypersonics/download")
        text = resp.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader)
        assert "PE Number" in headers
        assert "Service/Org" in headers


class TestGetDescription:
    def test_desc_existing_pe(self, client):
        resp = client.get("/api/v1/hypersonics/desc/0602702E")
        assert resp.status_code == 200
        body = resp.json()
        assert "description" in body

    def test_desc_nonexistent_pe(self, client):
        resp = client.get("/api/v1/hypersonics/desc/ZZZZZZZZ")
        assert resp.status_code == 200
        assert resp.json()["description"] is None


class TestDebug:
    def test_debug_returns_200(self, client):
        body = client.get("/api/v1/hypersonics/debug").json()
        assert "cache" in body

    def test_debug_cache_status(self, client):
        cache = client.get("/api/v1/hypersonics/debug").json()["cache"]
        assert "table_exists" in cache
        assert "row_count" in cache


class TestKeywordLists:
    """Validate that required search terms are present in keyword lists."""

    # Every term listed here must appear in _HYPERSONICS_KEYWORDS (case-insensitive).
    REQUIRED_KEYWORDS = [
        # Generic / cross-program
        "hypersonic", "boost glide", "glide body", "glide vehicle", "scramjet",
        # Offensive — Air Force
        "ARRW", "AGM-183", "HACM", "HCSW",
        # Offensive — Army
        "LRHW", "Dark Eagle", "OpFires",
        # Offensive — Navy / Joint
        "C-HGB", "CHGB", "conventional prompt strike", "prompt strike",
        # SM-6 / OASUW
        "offensive anti", "oasuw", "standard missile 6", "sm-6", "blk ib",
        "increment ii",
        # Generic speed / regime
        "high speed", "mach", "conventional prompt",
        # Defensive / tracking
        "Glide Phase Interceptor", "HBTSS",
    ]

    # Subset that must also be in _DESC_KEYWORDS
    REQUIRED_DESC_KEYWORDS = [
        "hypersonic", "boost glide", "glide vehicle", "scramjet",
        "ARRW", "HACM", "HCSW", "LRHW", "Dark Eagle",
        "C-HGB", "CHGB", "conventional prompt strike", "conventional prompt",
        "prompt strike", "Glide Phase Interceptor", "HBTSS", "OpFires",
        "offensive anti", "oasuw", "standard missile 6", "sm-6",
        "blk ib", "increment ii", "high speed", "mach",
    ]

    @pytest.mark.parametrize("required, actual, label", [
        ("REQUIRED_KEYWORDS", _HYPERSONICS_KEYWORDS, "_HYPERSONICS_KEYWORDS"),
        ("REQUIRED_DESC_KEYWORDS", _DESC_KEYWORDS, "_DESC_KEYWORDS"),
    ])
    def test_all_required_present(self, required, actual, label):
        required_list = getattr(self, required) if isinstance(required, str) else required
        actual_lower = [kw.lower() for kw in actual]
        missing = [kw for kw in required_list if kw.lower() not in actual_lower]
        assert not missing, f"Missing from {label}: {missing}"

    @pytest.mark.parametrize("keywords, label", [
        (_HYPERSONICS_KEYWORDS, "_HYPERSONICS_KEYWORDS"),
        (_DESC_KEYWORDS, "_DESC_KEYWORDS"),
    ])
    def test_no_duplicates(self, keywords, label):
        seen: set[str] = set()
        dupes = [kw for kw in keywords if (low := kw.lower()) in seen or seen.add(low)]  # type: ignore[func-returns-value]
        assert not dupes, f"Duplicate entries in {label}: {dupes}"

    def test_desc_keywords_subset_of_main(self):
        """Every desc keyword should also be in the main keyword list."""
        main_lower = {kw.lower() for kw in _HYPERSONICS_KEYWORDS}
        extras = [kw for kw in _DESC_KEYWORDS if kw.lower() not in main_lower]
        assert not extras, f"In _DESC_KEYWORDS but not _HYPERSONICS_KEYWORDS: {extras}"


class TestDownloadXLSX:
    def test_no_rows_selected_returns_400(self, client):
        resp = client.post(
            "/api/v1/hypersonics/download/xlsx",
            json={"show_ids": [], "total_ids": []},
        )
        assert resp.status_code == 400

    @staticmethod
    def _pick_show_ids(client, limit: int = 4) -> list[str]:
        """Read hypersonics JSON and construct ``pe-{idx}`` identifiers."""
        body = client.get("/api/v1/hypersonics").json()
        rows = body.get("rows", [])
        if not rows:
            pytest.skip("No hypersonics rows in test DB")
        ids: list[str] = []
        pe_counts: dict[str, int] = {}
        for row in rows:
            pe = row.get("pe_number")
            if not pe:
                continue
            idx = pe_counts.get(pe, 0)
            ids.append(f"{pe}-{idx}")
            pe_counts[pe] = idx + 1
            if len(ids) >= limit:
                break
        return ids

    @staticmethod
    def _download_xlsx(client) -> "openpyxl.Workbook":
        """POST the XLSX endpoint with sample IDs and return the loaded workbook."""
        show_ids = TestDownloadXLSX._pick_show_ids(client, limit=4)
        resp = client.post(
            "/api/v1/hypersonics/download/xlsx",
            json={"show_ids": show_ids, "total_ids": show_ids[:2]},
        )
        assert resp.status_code == 200
        return openpyxl.load_workbook(io.BytesIO(resp.content))

    def test_interleaved_fy_headers(self, client):
        """Each FY should produce a quad: ($K), In Total, Source, Description."""
        wb = self._download_xlsx(client)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        assert headers[:6] == [
            "PE Number", "Service/Org", "Exhibit", "Line Item / Sub-Program",
            "Budget Activity", "Color of Money",
        ]

        remainder = headers[6:]
        assert len(remainder) % 4 == 0, f"FY columns not in quads: {remainder}"
        for i in range(0, len(remainder), 4):
            val_h, it_h, src_h, desc_h = remainder[i : i + 4]
            assert val_h.endswith("($K)"), f"Expected value header at {i}: {val_h}"
            year = val_h.replace("FY", "").replace(" ($K)", "").strip()
            assert it_h == f"FY{year} In Total"
            assert src_h == f"FY{year} Source"
            assert desc_h == f"FY{year} Description"

    def test_totals_rows_y_p_grand(self, client):
        """Bottom should have Y TOTALS, P TOTALS, GRAND TOTAL rows with SUMIF formulas."""
        wb = self._download_xlsx(client)
        ws = wb.active
        last_row = ws.max_row

        assert ws.cell(row=last_row - 2, column=1).value == "Y TOTALS"
        assert ws.cell(row=last_row - 1, column=1).value == "P TOTALS"
        assert ws.cell(row=last_row, column=1).value == "GRAND TOTAL"

        y_formula = ws.cell(row=last_row - 2, column=7).value
        assert isinstance(y_formula, str) and y_formula.startswith("=SUMIF("), (
            f"Expected SUMIF formula in Y totals cell, got: {y_formula!r}"
        )
        assert '"Y"' in y_formula

        p_formula = ws.cell(row=last_row - 1, column=7).value
        assert isinstance(p_formula, str) and p_formula.startswith("=SUMIF("), (
            f"Expected SUMIF formula in P totals cell, got: {p_formula!r}"
        )
        assert '"P"' in p_formula

        grand_formula = ws.cell(row=last_row, column=7).value
        assert isinstance(grand_formula, str) and grand_formula.startswith("="), (
            f"Expected formula in Grand Total cell, got: {grand_formula!r}"
        )

    def test_summary_sheets_exist(self, client):
        """Workbook should have Y Summary, P Summary, Grand Total, and Dimensions sheets."""
        wb = self._download_xlsx(client)
        for name in ["Y Summary", "P Summary", "Grand Total", "Dimensions"]:
            assert name in wb.sheetnames, f"Missing {name} sheet: {wb.sheetnames}"

        # Y Summary should have spill formula in A2
        ws_y = wb["Y Summary"]
        a2 = ws_y.cell(row=2, column=1).value
        assert a2 and isinstance(a2, str) and "UNIQUE" in a2, (
            f"Expected UNIQUE spill formula in Y Summary A2, got: {a2!r}"
        )

    def test_data_validation_on_intotal_cells(self, client):
        """In Total cells should have Y/N/P data validation."""
        wb = self._download_xlsx(client)
        ws = wb.active
        validations = ws.data_validations.dataValidation
        assert len(validations) > 0, "No data validations found"
        dv = validations[0]
        assert dv.type == "list"
        assert "Y" in dv.formula1 and "N" in dv.formula1 and "P" in dv.formula1
