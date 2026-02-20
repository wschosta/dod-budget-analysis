"""
Integration tests for fixture file parsing — Step 1.C1

Verifies that the generated fixture .xlsx files in tests/fixtures/ can be
correctly parsed by the column detection logic and that the results match
the expected output JSON files.
"""
import json
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Stub pdfplumber to avoid heavy import
import types
sys.modules.setdefault("pdfplumber", types.ModuleType("pdfplumber"))

from build_budget_db import _detect_exhibit_type, _map_columns  # noqa: E402
from exhibit_catalog import find_matching_columns  # noqa: E402

import openpyxl  # noqa: E402

FIXTURES_DIR = Path(__file__).parent / "fixtures"
EXPECTED_DIR = FIXTURES_DIR / "expected"


def _load_expected(fixture_stem: str) -> dict | None:
    """Load the expected JSON for a given fixture stem."""
    json_path = EXPECTED_DIR / f"{fixture_stem}_expected.json"
    if not json_path.exists():
        return None
    return json.loads(json_path.read_text())


def _read_headers(xlsx_path: Path) -> list[str]:
    """Read the first row (headers) from an Excel file."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(max_row=1, values_only=True):
        headers = list(row)
        break
    wb.close()
    return headers


def _get_fixture_files() -> list[Path]:
    """Return all .xlsx files in the fixtures directory."""
    if not FIXTURES_DIR.exists():
        return []
    return sorted(FIXTURES_DIR.glob("*.xlsx"))


# Skip all tests if fixtures haven't been generated yet
pytestmark = pytest.mark.skipif(
    not _get_fixture_files(),
    reason="No fixture .xlsx files found — run scripts/generate_expected_output.py first",
)


class TestFixtureExhibitDetection:
    """Verify that _detect_exhibit_type correctly identifies each fixture file."""

    @pytest.mark.parametrize("xlsx_path", _get_fixture_files(),
                             ids=lambda p: p.name)
    def test_exhibit_type_detected(self, xlsx_path):
        expected = _load_expected(xlsx_path.stem)
        if expected is None:
            pytest.skip(f"No expected JSON for {xlsx_path.name}")

        detected = _detect_exhibit_type(xlsx_path.name)
        assert detected == expected["exhibit_type"], (
            f"Expected exhibit type '{expected['exhibit_type']}' "
            f"for {xlsx_path.name}, got '{detected}'"
        )


class TestFixtureColumnMapping:
    """Verify that _map_columns detects the expected columns for each fixture."""

    @pytest.mark.parametrize("xlsx_path", _get_fixture_files(),
                             ids=lambda p: p.name)
    def test_expected_columns_detected(self, xlsx_path):
        expected = _load_expected(xlsx_path.stem)
        if expected is None:
            pytest.skip(f"No expected JSON for {xlsx_path.name}")

        headers = _read_headers(xlsx_path)
        exhibit_type = expected["exhibit_type"]
        mapping = _map_columns(headers, exhibit_type)

        expected_cols = set(expected.get("columns_detected", []))
        mapped_cols = set(mapping.keys())

        # All expected columns should be present in the mapping
        missing = expected_cols - mapped_cols
        assert not missing, (
            f"Missing expected columns for {xlsx_path.name}: {missing}\n"
            f"Detected: {sorted(mapped_cols)}"
        )


class TestFixtureRowCount:
    """Verify that fixture files have the expected number of data rows."""

    @pytest.mark.parametrize("xlsx_path", _get_fixture_files(),
                             ids=lambda p: p.name)
    def test_row_count(self, xlsx_path):
        expected = _load_expected(xlsx_path.stem)
        if expected is None:
            pytest.skip(f"No expected JSON for {xlsx_path.name}")

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True,
                                    data_only=True)
        ws = wb.active
        # Count all rows minus the header
        row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert row_count == expected["row_count"], (
            f"Expected {expected['row_count']} data rows in {xlsx_path.name}, "
            f"got {row_count}"
        )


class TestFixtureCatalogMatching:
    """Verify that find_matching_columns works for detail exhibit fixtures."""

    def test_p5_fixture_catalog_match(self):
        xlsx_path = FIXTURES_DIR / "army_p5_fy2026.xlsx"
        if not xlsx_path.exists():
            pytest.skip("P-5 fixture not found")

        headers = _read_headers(xlsx_path)
        matched = find_matching_columns("p5", headers)
        fields = set(matched.values())

        assert "program_element" in fields
        assert "line_item_number" in fields
        assert "line_item_title" in fields
        assert "unit" in fields

    def test_r2_fixture_catalog_match(self):
        xlsx_path = FIXTURES_DIR / "army_r2_fy2026.xlsx"
        if not xlsx_path.exists():
            pytest.skip("R-2 fixture not found")

        headers = _read_headers(xlsx_path)
        matched = find_matching_columns("r2", headers)
        fields = set(matched.values())

        assert "program_element" in fields
        assert "sub_element" in fields
        assert "title" in fields
